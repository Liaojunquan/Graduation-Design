import tkinter as tk
from tkinter import ttk   #使用Notebook模块
from tkinter import *
import tkinter.messagebox
import threading
import selenium as se
from selenium import webdriver
from selenium.webdriver.remote.command import Command              #浏览器状态命令
from selenium.webdriver.common.action_chains import ActionChains   #用于操控鼠标在网页上移动
import time
import os
import openpyxl
import requests     #网页请求模块
import bs4          #html解析模块
import jieba
import copy
from pyecharts.charts import Map, Bar, Pie, Line
from pyecharts import options as opts
from wordcloud import WordCloud
import multiprocessing
import psutil  #获取内存信息


#全局变量
driver = None            #浏览器驱动变量
sleepTime = 5            #用于间隔休眠的时间，单位秒
startFlag = False        #是否未停止，用于控制爬虫是否停止。一旦按下开始键，是否暂停中或运行中，一律当爬虫未停止。
runFlag = False          #是否运行中，用于开始或暂停爬虫
stopFlag = False         #传递到爬虫内部，用于彻底停止爬虫
chromeOpen = False       #谷歌浏览器是否已打开
file_name = os.path.abspath('..') + r'\data.xlsx'           #数据保存文件名
wb = None                #Excel工作本Workbook
ws = None                #Excel活动表格
android_kw = ['Android', 'android', '安卓', 'ANDROID']
animate_kw = ['动画', '动效', '动作', 'flash', 'Flash', 'FLASH', 'MG', 'Mg', 'mg', '动漫','Motion','motion','animate','Animate','Animator','animator']
cs_kw = ['C#', 'c#', 'NET', 'Net', 'net']
cpp_kw = ['C++', 'c++']
game_kw = ['端游', '手游', '网游', '游戏', 'UE4', 'ue4', 'Ue4', '虚幻', 'cocos', 'Cocos', 'COCOS', 'TA', 'Ta', '技术美术', '技美', 'Laya', 'laya', 'LAYA', 'Egret', 'egret', 'EGRET', '主美', '场景', '数值', '关卡', '系统策划', '剧情策划', '玩法', '主策划', '主程']
graphic_kw = ['UI', 'ui', 'Ui', '界面', '平面设计', '美工', '插画', '图案', '图标', '修图', '修片', '抠图', 'PS', 'ps', 'Ps', 'photoshop', 'Photoshop', '图片', 'PHOTOSHOP',
                '漫画', 'Graphic', 'graphic', '排版', '板式', '封面', '原画']
java_kw = ['Java', 'java', 'JAVA']
media_kw = ['音频', '视频', '导演', '编导', '分镜', '记者', '电影', '电视', '媒体', '文案', '公众号', '新闻', '广播', '编辑', '广告', '小编', '小红书', '微博', '抖音', 'B站', 'b站']
model_kw = ['建模', '模型', 'maya', 'Maya', 'MAYA', '3dmax', '3DMAX', '3DMax', '3Dmax', '3dMax', '3dMAX', 'ZB', 'Zb', 'zb', 'ZBrush', 'zbrush', 'Zbrush', 'ZBRUSH', '犀牛', 'Blender',
              'blender', '3ds', '3Ds', '贴图', '材质', '绑骨', '绑定', 'c4d', 'C4d', 'C4D', 'c4D']
photography_kw = ['摄影', '摄像', '拍摄', '录像', 'photograph', '航拍', '飞手']
post_production_kw = ['渲染', '特效', '影视', '后期', '剪辑', '合成', 'CG', 'cg', 'Cg', '达芬奇', 'Davinci', '调色', '调光', '字幕', '包装', '地编', 'Pipeline', 'pipeline', '地编']
sql_kw = ['sql', 'Sql', 'SQL', '数据库', 'DB', 'DataBase', 'Database', 'database', 'dba']
technology_kw = ['算法', '数字信号', '编码', '解码', '图像', '图形', '数据结构', '程序']
unity_kw = ['Unity', 'U3D', 'U3d', 'u3d', 'u3D', 'unity', 'UNITY', 'Shader', 'SHADER', 'shader']
virtual_kw = ['AR', 'VR', 'Vr', '虚拟现实', 'Ar']
web_kw = ['WEB', 'web', 'Web', '网页', 'html', 'HTML', 'Html', 'css', 'CSS', 'Css', 'JavaScript', 'javascript', 'Javascript', '网站', '前端', 'JS', 'js', 'Js', 'Angular', 'angular', 'php', 'PHP']
smkw = android_kw + animate_kw + cs_kw + cpp_kw + game_kw + graphic_kw + java_kw + media_kw + model_kw + photography_kw + post_production_kw + sql_kw + technology_kw + unity_kw + virtual_kw + web_kw
absolute_kw = ['flash', 'Flash', 'FLASH', 'PS', 'Ps', 'ps', 'CG', 'cg', 'Cg', 'VR', 'Vr', 'vr', 'AR', 'Ar', 'ar', 'TA', 'Ta', 'UI', 'Ui', 'ui', 'NET', 'net', 'Net', 'php', 'PHP']
#KeyWord = """引擎 客户端 开发 端游 手游 小游戏 Unity U3D u3d U3d UE4 ue4 Ue4 UE 虚幻 cocos Cocos COCOS 2d 平面设计 三维 建模 渲染 视觉 MAYA Maya maya C4D c4d C4d 3DMAX 3dmax ZBrush 
#2D 3d 3D 测试 策划 美工 美术 设计 UI ui Ui 特效 动画 动作 服务器 维护 维稳 脚本 数据 前端 WEB web Web 贴图 材质 影视 视频 拍摄 摄影 摄像 导演 编导 前期 后期 剪辑 分镜 AR VR 
#小程序 数据库 数据管理 数据库管理 DBA sql SQL Sql JAVA Java java C++ c++ C c Windows windows Android android 安卓 Python python 图像 音频 音视频 VR  Vr vr 虚拟现实 AR ar 
#Ar Unreal 图形 Flash flash FLASH MG mg Mg 二维 分镜 达芬奇 多媒体 Davinci 调色 修图 调光 图片 媒体 算法 CG Cg cg 航拍 飞手 原画 html HTML Html JS js Js 网页 
#数据结构 .NET .net .Net C# 游戏 TA ZB 技术美术 技美 zb"""     #关键词

StopWord = """销售 推广 地产 经理 主管 客服 人事 教育 讲师 分销 电商 投放 运营 翻译 英语 商务 教师 教研 老师 玩游戏 陪 质检 猎头 游戏机 装配 美主 gm gs GM GS Gm Gs 生产 * 激光 
扫描 机械 公关 理财 译员 发行 主播 直播 玩家 试玩 帮派 治疗 市场 营销 游戏店 服务员 体验 助教 经纪人 管培 投资 顾问 合伙人 投标 审核 制片人 演员 自动驾驶 售前 售后 院长 企业文化 体验馆 店长 
接待 财务 店员 营业员 内容生态 看房 看楼 讲解 ARM arm 收账 会计 半导体 工艺 交通 能源 射频 SAE 标注 漆 涂料 化工 甲油胶 建筑 建材 油墨 印刷 环保 化学 热转印 汽车 胶 粉体 灯光 CAE FAE 
硬件 嵌入式 自动化 芯片 AE工程师 供应链 AE应用工程师 PR工程师 仪器 AE技术 驱动 电路 免费 调色师傅 调色员 美甲 印 电子 鞋 调色工程师 调色技术 电镀 技工 粉 妆 涂 油 配色 普工 操作员 剂 分子 
调色员 墨 瓷砖 工厂 家居 SMT 外贸 带头 FPGA 审核 智能检测 咨询 设备 保险 家庭 邀约 招生 KA 社区 置业 司机 配送 孵化 机器人 需求 微商 业务 寒假 货 蓝牙 幼 营运 电玩 店 打游戏 送餐 
辅导 单片机 保育 厨 网吧 秘书 总裁 家教 主持人 物流 收银 咖啡 物料 装修 灯 室内 仓 保健 干部 封装 采购 导玩 乐园 代驾 班主任 跟单 茶 篮球 引导员 达人 法务 解说 信用 婴 团长 爱好 电话 HR 
人力资源 催收 吃 住 人寿 前台 出纳 金融 电竞 骑手 保安 交易员 贷 代练 拓展 渠道 人力 地推 赚钱 风控 投诉 游戏学徒 公会 党 游戏专员 电销 银行 测评 管理培训 文员 政 大使 底层 政策 电话 网销 
导师 出口 操作 游戏助理 游戏管理员 快递 临时工 保姆 嫂 模特 SEO SEM seo sem Seo Sem 交易 证券 时工 上门 分拣 教练 美容 礼仪 清洁 维修 检员 打字 塑 房 预算 造价 社 融 赛事 金牌 乐有家 车场 
收费 组装 课程 短期工 托 内训 服装设计 招聘专员 校长 装机 本地 医药 物业 客流 监控 员工 知识产权 税务 安保 医生 医师 ASO 课程设计 药师 医学 家装 整装 工业 电工 项目助理 网络管理 IT管理 网管 
网络管理 导师 施工 亚马逊 广告优化 电梯 PE 客户代表 结构 审计 商业化 商业 硬件 客户 媒介 葡萄牙 法语 俄语 商业 意大利 西班牙 德语 韩语 用户研究 信息流 Amazon 产品开发 专柜 导购 绩效 调酒 
音效 学历规划 招聘 大客户 质量检测 aso 通讯 IE 日语 厂 阿拉伯 手机包装 网络工程 选址 音效 录音 艺人 配音 教务 地理 演出 品牌 服装 招商 保洁 放映 经纪 歌手 星探 选片 DJ 主持 品牌专员 
学术 策划执行 策划总监 执行总监 回访 舆情 聊天 品牌策划 活动策划 园林 出差 股票 资料 楼宇 换画 企宣 策划专员 IT 培训专员 版权 企划 五金 培训 电气 收发 焊 险 招标 零售 CEO 品质 护士 董事长 
话务 技术工程 统筹 录入 报价 情感分析 后勤 结算 预算 结预算 教务 木工 舞台 律师 舞台 演出 授权 网络安全 音响 买手 造型 考研 暖通 订单 博主 演艺 音乐制作 舞 文旅 花艺 园艺 编曲 制片 办公室 
模具 软装 硬装 项目管理 打包 IQC CNC 听写 听译 城市 馆长 安装 甜品 点心 阿姨 看房 照顾 老人 书法 场务 招募专员 洗碗 照明设计 试衣 YY 小丑 申报 声乐 作曲 车间 QC 项目总监 项目策划 展厅策划 
乐队 学习管理 礼服 车间 演奏 发单 监察 PE PMC 喷绘 话务 TDDI 机联网 物联网 CTO 技术官 BSP 无线 协议 MES CTO QA 上位机 报表 数字IC 项目专员 移民 光学 OA 通讯 IC数字 数字前端 供应商 QE 电源 
动力 安防 BI 弱电 资金 理赔 QT 消防 产品专员 5G 网路 生信 信息安全 渗透 OTA 竞价 Facebook IC前端 外卖 速卖通 船务 带团 游戏平台 生物 关务 商标 治具 发动机 电机 全屋 家具 钣金 LED 制冷 成本 
液压 热处理 报关 稽查 刀 拆单 水处理 专利 车 电池 底盘 农业 夹具 线束 给排水 连接 PCB 磨 铣 散热 电器 钳 火花 学管 冲压 抛光 打磨 土 ME 下单 底盘 热管理 液 桥 天线 仿真 电池 稽 土建 背光 
省模 夹 物控 材料研发 EHS 营销 护理 实验室 电控 CMF BMS 空调 督导 生态 营养 诉讼 安检 活动执行 机电 留学 链家 谈判 验布 规划 工程监理 质量安全 海运 文件员 商品专员 物探 变频 拆迁 水电 估价 
面料 发卡 加盟 升学 直营 旧改 维保 保修 测绘 测量 勘测 公共关系 会务 缝纫 董秘 传单"""  #停用词

TagStopWord = "漆 涂 化工 甲油胶 建筑 建材 油 墨 印刷 环保 化学 汽车 胶 粉 妆 灯光 调研 营销 市场 客户 乙方 射频 销售 公关 电商 创意 推广 广告 运营 项目 执行 策划 创意 玩具 装饰 装修 物流"  #标签停用词
sw = StopWord.replace('\n','').split(' ')       #停用词列表
#kw = KeyWord.replace('\n','').split(' ')        #关键词列表
tsw = TagStopWord.replace('\n','').split(' ')   #标签停用词列表
isChecking = False             #是否查重中标志
isStopChecking = False         #是否停止查重
root = None                    #tkinter的root
force_exit = False             #用于强制退出程序的标志
checkGD = False                #是否去除广东以外的数据
isVisualization = False        #是否可视化
isStopVisualization = False    #是否停止可视化

def Do_Destroy():          #全局函数用于监听窗口是否关闭
 global runFlag
 global startFlag
 global stopFlag
 global wb
 global file_name
 global isChecking
 global isStopChecking
 global root
 global force_exit
 global isVisualization
 if startFlag or runFlag or isChecking or isVisualization:   #爬虫或查重或可视化运行中
  if tk.messagebox.askyesno(title = '警告', message = '程序正在运行，是否确定要退出？'):  #抛出提示
   force_exit = True
   stopFlag = True
   isStopChecking = True
   tk.messagebox.showinfo('提示', '点击确认3秒后退出程序')
   time.sleep(3)
   wb.close()
   root.destroy()  #关闭软件窗口
  else:
   return
 elif tk.messagebox.askyesno(title = '警告', message = '是否确定要退出？'):  #抛出提示
  if wb != None:
   wb.close()
  root.destroy()  #关闭软件窗口

#-------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Excel工作簿的初始化
class initWB():       #初始化工作本和工作表，无则创建
 def __init__(self, ui):
  self.ui = ui
  
 def open_data_file(self, create_file):
  global wb
  global ws
  self.ui.log('打开data.xlsx文件中...')
  try:
   wb = openpyxl.open(file_name)      #打开Excel文件
  except FileNotFoundError:
   if create_file:
    self.ui.log('找不到Excel文件data!')
    self.ui.log('创建一个新的Excel文件')
    wb = openpyxl.Workbook()       #创建一个新的工作本
    ws = wb.active
    ws.append(['职位名称', '最低薪酬(元/月)', '最高薪酬', '平均薪酬', '公司所在地', '经验要求', '学历要求', '公司福利', '公司名称', '链接地址', '公司类型', '公司大小', '业务定位方向', '发布时间', '职位要求和描述', '标记']) #首行标题
    ws.freeze_panes = 'A2'  #冻结首行
    wb.save(file_name)      #保存文件
    self.ui.log('Excel文件data创建完成!')
    return True
   else:
    return False
  else:
   self.ui.log('成功找到并打开Excel文件data')
   ws = wb.active
   return True

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#爬虫实现
class Crawling():
 def __init__(self, ui):
  self.ui = ui                #UI root
  self.current_url = ''       #当前URL
  self.isFirst = True         #是否第一次进入爬取循环
  self.delWord = ['急聘', '双休', '急招', '周末双休', '13薪', '14薪', '高薪', '年终奖', '奖金', '()', '（）', '大小周', '年底双薪', '带薪年假', '餐补', '朝九晚五', '大小休', '包餐', '包三餐',
                  '福利好', '福利全', '地铁上盖', '地铁旁边', '近地铁', '地铁', '大平台', '平台大', '统招本科', '年以上经验', '不加班', '五天制', '全勤奖', '假期多', '诚聘', '直招', '直聘',
                  '上市公司','上市', '补贴多', '坐班', '白班', '无经验', '周末', '早九晚五', '六天工作制', '外包', '（+）', '新人可带', '食宿补贴', '广州', '（广州）', '(广州)', '深圳',
                  '（深圳）', '(深圳)', '视频面试', '远程面试', '三餐', '3餐', '包3餐', '知名网络科技公司']

 def pause(self):     #暂停函数
  global runFlag
  global stopFlag
  global startFlag
  while startFlag and not runFlag and not stopFlag:
   #self.ui.log('暂停中')
   time.sleep(1)      #休眠实现暂停

 def mStop(self):
  global stopFlag
  global startFlag
  global runFlag
  global wb
  global file_name
  global force_exit
  if force_exit:
   return True
  if startFlag and stopFlag:
   self.ui.log('结束爬取循环!')
   self.ui.log('保存data文件中...')
   wb.save(file_name)
   self.isFirst = True  #重置第一次循环标志
   stopFlag = False
   startFlag = False
   runFlag = False
   self.ui.log("爬虫已被停止!")
   self.ui.openBossButton.config(state=NORMAL)
   self.ui.openJobButton.config(state=NORMAL)
   self.ui.notGDButton.config(state=NORMAL)
   self.ui.checkButton.config(state=NORMAL)
   self.ui.startStopVirtualBtn.config(state=NORMAL)
   self.ui.startButton.config(state=NORMAL)
   self.ui.stopButton.config(state=NORMAL)
   return True
  else:
   return False

 def de_illegal(self, o_str):    #去非法字符函数
  illegal_str = ['\000','\001','\002','\003','\004','\005','\006','\007','\010','\013','\014','\016','\017','\020','\021',
                 '\022','\023','\024','\025','\026','\027','\030','\031','\032','\033','\034','\035','\036','\037']  #excel无法识别的非法字符
  n_str = o_str    #字符串浅拷贝
  for each_i_s in illegal_str:
   n_str = n_str.replace(each_i_s, '')   #去除非法字符
  return n_str
  
 def open_URL(self, url, type_get):                #使用requests模块获取前程无忧网站各工作的职位描述和要求
  #print(url)
  tmp_time = 3           #最多重连3次
  bmsg = None
  tmp_sleep_time = 0     #休眠时间
  while tmp_time > 0:
   self.pause()         #提供暂停
   if self.mStop():     #提供停止
    return ""
   headers = {'User-Agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.89 Safari/537.36'}
   while True:
    self.pause()         #提供暂停
    try:
     res = requests.get(url, headers=headers ,timeout = 8)   #8秒后超时
    except requests.exceptions.ConnectTimeout:
     tmp_sleep_time += 5
     self.ui.log('连接超时！ ' + str(tmp_sleep_time) + '秒后重连')
     if self.mStop():    #提供停止
      return ""
     time.sleep(tmp_sleep_time)
    except (requests.exceptions.MissingSchema, requests.exceptions.InvalidURL):
     self.ui.log('URL参数错误！')
     return ""
    except requests.exceptions.ConnectionError:
     self.ui.log('连接出错！')
     return ""
    else:
     break
   self.pause()         #提供暂停
   res.encoding = 'gbk'      #字符编码为gbk
   soup = bs4.BeautifulSoup(res.text,'html.parser')
   if type_get == 0:
    bmsg = soup.find('p', class_='msg')    #获取公司位置
   elif type_get == 1:
    bmsg = soup.find('span', class_='i_flag')    #获取公司类型
   elif type_get == 2:
    bmsg = soup.find('span', class_='i_people')  #获取公司人数
   elif type_get == 3:
    bmsg = soup.find('span', class_='i_trade')   #获取公司业务方向
   elif type_get == 4:
    bmsg = soup.find('div', class_='bmsg')    #获取职位描述
   elif type_get == 5:
    if soup.find('div', class_='research') != None:   #暂停招聘
     return True
    else:
     return False
   if bmsg == None:
    tmp_sleep_time += 5   #休眠间隔加5s
    if type_get == 0:
     self.ui.log('bs4找不到公司位置元素！ ' + str(tmp_sleep_time) + '秒后重试')  #bs4找不到该元素
    elif type_get == 1:
     self.ui.log('bs4找不到公司类型元素！ ' + str(tmp_sleep_time) + '秒后重试')  #bs4找不到该元素
    elif type_get == 2:
     self.ui.log('bs4找不到公司人数元素！ ' + str(tmp_sleep_time) + '秒后重试')  #bs4找不到该元素
    elif type_get == 3:
     self.ui.log('bs4找不到公司业务方向元素！ ' + str(tmp_sleep_time) + '秒后重试')  #bs4找不到该元素
    elif type_get == 4:
     self.ui.log('bs4找不到职位描述元素！ ' + str(tmp_sleep_time) + '秒后重试')  #bs4找不到该元素
    if self.mStop():   #提供停止
     return ""
    time.sleep(tmp_sleep_time)
    tmp_time -= 1     #次数减一
   else:
    break
  self.pause()         #提供暂停
  if tmp_time == 0:
   return ""
  elif type_get == 0:
   dizhi_tmp =  bmsg.text.replace('\xa0','').replace(' ','').split('|')[0]    #获取公司位置
   if dizhi_tmp.find('-') == -1:
    qu_tmp = soup.find('div', class_='bmsg inbox')
    if qu_tmp != None:    #上班地址元素含有xx区
     qu_tmp = qu_tmp.text.replace('\n','').replace('上班地址：','').replace('工作地址','').replace('\xa0','')
     if qu_tmp.find('区') != -1 and qu_tmp.find('区')+1 < len(qu_tmp):
      if qu_tmp.find('市') == -1:
       return dizhi_tmp + '-' + qu_tmp[:qu_tmp.find('区')+1]   #xx区写入公司位置
      elif qu_tmp.find('市') < len(qu_tmp)/2:
       return dizhi_tmp + '-' + qu_tmp[qu_tmp.find('市')+1 : qu_tmp.find('区')+1]   #xx区写入公司位置
     elif qu_tmp.find('镇') != -1 and qu_tmp.find('镇')+1 < len(qu_tmp):
      if qu_tmp.find('市') == -1:
       return dizhi_tmp + '-' + qu_tmp[:qu_tmp.find('镇')+1]   #xx镇写入公司位置
      elif qu_tmp.find('市') < len(qu_tmp)/2:
       return dizhi_tmp + '-' + qu_tmp[qu_tmp.find('市')+1 : qu_tmp.find('镇')+1]   #xx镇写入公司位置
   return dizhi_tmp
  elif type_get == 1:
   return bmsg.parent.text.replace('\n','').replace(' ','').replace('\xa0','')    #获取公司类型
  elif type_get == 2:
   return bmsg.parent.text.replace('\n','').replace(' ','').replace('\xa0','')   #获取公司人数
  elif type_get == 3:
   return bmsg.parent.text.replace('\n','').replace(' ','').replace('\xa0','')   #获取公司业务方向
  elif type_get == 4:
   return bmsg.text.replace('\n',' ').replace('\xa0',' ').replace('微信分享','').replace('【',' ').replace('】',' ').replace('ü','').replace('●','').replace('★','').replace('◆','').replace('*','').replace('■','')    #获取职位描述

 def append_list_job(self, ee, ee_l, ee_d, ee_fl, ee_c, ee_link, ee_t_s, ee_b, ee_time, ee_msg):  #职位名称  薪资  地区|经验|学历  福利  公司名称  链接  公司类型|大小  业务方向 发布时间 职位描述和要求
  l = []
  sal_low = 0
  sal_hight = 0
  sal_avg = 0
  self.pause()         #提供暂停

  l.append(ee.strip())     #职位名称   ee_tmp.strip()
  ee_l = ee_l.replace(' ','')
  if ee_l.find('-') == -1:
   if ee_l.find('元/日') != -1:
    sal_low = int(ee_l.split('元/日')[0].strip()) * 22              #统一标准单位元/月
    sal_hight = int(ee_l.split('元/日')[0].strip()) * 26
    sal_avg = (sal_low + sal_hight) // 2
    l.append(sal_low)
    l.append(sal_hight)
    l.append(sal_avg)
   elif ee_l.find('元/天') != -1:
    sal_low = int(ee_l.split('元/天')[0].strip()) * 22              #统一标准单位元/月
    sal_hight = int(ee_l.split('元/天')[0].strip()) * 26
    sal_avg = (sal_low + sal_hight) // 2
    l.append(sal_low)
    l.append(sal_hight)
    l.append(sal_avg)
   elif ee_l.find('万/月') != -1:
    sal_low = int(float(ee_l.split('万/月')[0].strip()) * 10000)             #统一标准单位元/月
    sal_hight = sal_low
    sal_avg = sal_low
    l.append(sal_low)
    l.append(sal_hight)
    l.append(sal_avg)
   elif ee_l.find('千/月') != -1:
    sal_low = int(float(ee_l.split('千/月')[0].strip()) * 1000)             #统一标准单位元/月
    sal_hight = sal_low
    sal_avg = sal_low
    l.append(sal_low)
    l.append(sal_hight)
    l.append(sal_avg)
   elif ee_l.find('万/年') != -1:
    sal_low = int(float(ee_l.split('万/年')[0].strip()) * 10000 // 12)             #统一标准单位元/月
    sal_hight = sal_low
    sal_avg = sal_low
    l.append(sal_low)
    l.append(sal_hight)
    l.append(sal_avg)
   else:
    l.append(ee_l.strip())
    l.append("null")
    l.append("null")
  else:
   if ee_l.find('万/月') != -1:
    try:
     sal_low = int(float(ee_l.split('万/月')[0].split('-')[0]) * 10000)      #统一标准单位元/月
     sal_hight = int(float(ee_l.split('万/月')[0].split('-')[1]) * 10000)
     sal_avg = (sal_low + sal_hight) // 2
     l.append(sal_low)
     l.append(sal_hight)
     l.append(sal_avg)
    except:
     l.append(ee_l.strip())
     l.append("null")
     l.append("null")
   elif ee_l.find('千/月') != -1:
    try:
     sal_low = int(float(ee_l.split('千/月')[0].split('-')[0]) * 1000)            #统一标准单位元/月
     sal_hight = int(float(ee_l.split('千/月')[0].split('-')[1]) * 1000)
     sal_avg = (sal_low + sal_hight) // 2
     l.append(sal_low)
     l.append(sal_hight)
     l.append(sal_avg)
    except:
     l.append(ee_l.strip())
     l.append("null")
     l.append("null")
   elif ee_l.find('万/年') != -1:
    try:
     sal_low = int(float(ee_l.split('万/年')[0].split('-')[0]) * 10000 // 12)            #统一标准单位元/月
     sal_hight = int(float(ee_l.split('万/年')[0].split('-')[1]) * 10000 // 12)
     sal_avg = (sal_low + sal_hight) // 2
     l.append(sal_low)
     l.append(sal_hight)
     l.append(sal_avg)
    except:
     l.append(ee_l.strip())
     l.append("null")
     l.append("null")
   elif ee_l.find('元/日') != -1:
    try:
     sal_low = int(float(ee_l.split('元/日')[0].split('-')[0]) * 26)            #统一标准单位元/月
     sal_hight = int(float(ee_l.split('元/日')[0].split('-')[1]) * 26)
     sal_avg = (sal_low + sal_hight) // 2
     l.append(sal_low)
     l.append(sal_hight)
     l.append(sal_avg)
    except:
     l.append(ee_l.strip())
     l.append("null")
     l.append("null")
   elif ee_l.find('元/天') != -1:
    try:
     sal_low = int(float(ee_l.split('元/天')[0].split('-')[0]) * 26)            #统一标准单位元/月
     sal_hight = int(float(ee_l.split('元/天')[0].split('-')[1]) * 26)
     sal_avg = (sal_low + sal_hight) // 2
     l.append(sal_low)
     l.append(sal_hight)
     l.append(sal_avg)
    except:
     l.append(ee_l.strip())
     l.append("null")
     l.append("null")
   else:
    l.append(ee_l.strip())
    l.append("null")
    l.append("null")
        
  self.pause()         #提供暂停
  if ee_d.find('|') == -1:
   l.append(ee_d.strip())
   l.append("")
   l.append("")
  else:
   ee_d = ee_d.replace(' ','')       #去空格
   tmp_index = ee_d.rfind('招')      #从右往左找
   if tmp_index != -1:
    ee_d = ee_d[0:tmp_index-1]        #去掉招收人数
   tmp_index = 0
   while tmp_index < len(ee_d):
    if ee_d[tmp_index].isdigit():      #判断字符串中是否含有数字
     break
    else:
     tmp_index += 1
   if len(ee_d.split('|')) == 3:
    if ee_d.split('|')[0].strip() == '异地招聘' or ee_d.split('|')[0].find('-') == -1:  #地址为异地招聘或不含区
     l.append(self.open_URL(ee_link, 0))        #bs4打开网页查找公司地址
    else:
     l.append(ee_d.split('|')[0].strip())       #公司所在地
    l.append(ee_d.split('|')[1].strip())       #经验
    l.append(ee_d.split('|')[2].strip())       #学历
   elif len(ee_d.split('|')) == 2:
    if tmp_index < len(ee_d):   #含有数字
     if ee_d.split('|')[0].strip() == '异地招聘' or ee_d.split('|')[0].find('-') == -1:  #地址为异地招聘或不含区
      l.append(self.open_URL(ee_link, 0))        #bs4打开网页查找公司地址
     else:
      l.append(ee_d.split('|')[0].strip())       #公司所在地
     l.append(ee_d.split('|')[1].strip())   #经验
     l.append("")
    else:    #不含数字
     if ee_d.find('在校') == -1 and ee_d.find('应届') == -1 and ee_d.find('无需经验') == -1:   #不包含经验要求
      if ee_d.split('|')[0].strip() == '异地招聘' or ee_d.split('|')[0].find('-') == -1:  #地址为异地招聘或不含区
       l.append(self.open_URL(ee_link, 0))        #bs4打开网页查找公司地址
      else:
       l.append(ee_d.split('|')[0].strip())       #公司所在地
      l.append("")
      l.append(ee_d.split('|')[1].strip())   #学历
     else:                                                  #包含经验要求
      if ee_d.split('|')[0].strip() == '异地招聘' or ee_d.split('|')[0].find('-') == -1:  #地址为异地招聘或不含区
       l.append(self.open_URL(ee_link, 0))        #bs4打开网页查找公司地址
      else:
       l.append(ee_d.split('|')[0].strip())       #公司所在地
      l.append(ee_d.split('|')[1].strip())   #经验
      l.append("")
   elif len(ee_d.split('|')) == 1:
    l.append(ee_d.split('|')[0].strip())
    l.append("")
    l.append("")
        
  self.pause()         #提供暂停
  l.append(ee_fl.strip())               #福利
  l.append(ee_c.strip())                #公司名称
  l.append(ee_link.strip())             #链接地址
  if ee_t_s.find('|') == -1:
   ee_t_tmp = self.open_URL(ee_link, 1)   #bs4打开网页查找公司类型
   if ee_t_tmp == "" or ee_t_tmp == None:
    l.append('null')
   else:
    l.append(ee_t_tmp)   #公司类型
   time.sleep(1)
   ee_s_tmp = self.open_URL(ee_link, 2)   #bs4打开网页查找公司大小
   if ee_s_tmp == "" or ee_s_tmp == None:
    l.append('null')
   else:
    l.append(ee_s_tmp)   #公司大小
  else:
   l.append(ee_t_s.split('|')[0].strip())    #公司类型
   l.append(ee_t_s.split('|')[1].strip())    #公司大小

  self.pause()         #提供暂停
  if ee_b == "":
   ee_b_tmp = self.open_URL(ee_link, 3)   #bs4打开网页查找公司业务方向
   if ee_b_tmp == "" or ee_b_tmp == None:
    l.append('null')
   else:
    l.append(ee_b_tmp)   #公司业务方向
  else:
   l.append(ee_b.strip())         #公司业务方向

  if ee_time.find('发布') != -1 and ee_time.find('-') != -1:    #发布日期转整数
   tmp_time = ee_time.split('发布')[0]
   l.append(int(tmp_time.replace('-', '')))
  else:
   l.append(ee_time.strip())
  l.append(ee_msg.replace('\n','').replace('\t','').replace('【',' ').replace('】',' ').replace('"','').replace('  ','').replace('急聘','').replace('.','').replace('*','').strip())   #职位描述和要求
  self.ui.log('---' + ee + '---')   #打印职位
  return l

 def get_job(self):
  global sw
  global smkw
  global driver
  global file_name
  global sleepTime
  global stopFlag
  global startFlag
  global runFlag
  global wb
  global ws
  self.current_url = driver.current_url
  while (self.isFirst or driver.current_url != self.current_url) and driver.current_url.find('51job.com') != -1:  #当前页面与上一页面不同，并且URL包含51job.com
   self.pause()            #提供暂停
   if self.mStop():       #提供停止
    return
   self.isFirst = False     #第一次循环设否
   self.ui.log("爬取当前页面中---------------------")
   self.current_url = driver.current_url
   while 1:
    self.pause()     #提供暂停
    if self.mStop():       #提供停止
     return
    try:
     job_list = driver.find_element_by_class_name('j_joblist').find_elements_by_class_name('e')   #获取列表
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("无法获取职位列表  休眠15秒后重试")
     self.pause()     #提供暂停
     if self.mStop():       #提供停止
      return
     time.sleep(15)
    else:
     if len(job_list) > 0:                 #获取列表成功
      break
     else:                                 #列表长度有误
      self.ui.log("错误：职位列表长度为0!  休眠10秒后重试")
      self.pause()     #提供暂停
      if self.mStop():       #提供停止
       return
      time.sleep(10)
    
   for i in range(len(job_list)):
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    e = job_list[i].find_element_by_class_name("jname").text      #获取工作名称
    tmp = 0
    while tmp < len(sw):
     self.pause()         #提供暂停
     if self.mStop():       #提供停止
      return
     if e.find(sw[tmp]) != -1:
      break             #包含停用词，跳到下一工作
     else:
      tmp += 1
    if tmp < len(sw):
     continue           #包含停用词，跳到下一工作
    tmp = 0
    while tmp < len(smkw):
     self.pause()         #提供暂停
     if self.mStop():       #提供停止
      return
     if e.find(smkw[tmp]) != -1:
      break             #包含关键词
     else:
      tmp += 1
    if tmp == len(smkw):
     continue           #都不包含关键词，跳到下一工作
    e_time = ""
    try:
     e_time = job_list[i].find_element_by_class_name("time").text      #获取发布时间
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有发布时间可获取")
    except:
     self.ui.log("获取发布时间出现其它错误!")
    e_l = job_list[i].find_element_by_class_name("sal").text     #薪酬
    e_d = job_list[i].find_element_by_class_name("d").text       #公司所在地 经验 学历
    e_c = job_list[i].find_element_by_class_name("cname").text   #公司名称
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    e_fl = ""
    try:
     e_fl = job_list[i].find_element_by_class_name("tags").get_attribute('title')   #福利
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有福利信息可获取")
    except:
     self.ui.log("获取福利信息出现其它错误!")
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    e_link = job_list[i].find_element_by_class_name("el").get_attribute('href')    #详情链接地址
    if self.open_URL(e_link, 5) == True:         #已经暂停招聘
     continue
    e_type_size = ""
    try:
     e_type_size = job_list[i].find_element_by_class_name("dc").text              #公司类型和大小
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有公司类型和大小信息可获取")
    except:
     self.ui.log("获取公司类型和大小信息出现其它错误!")
    e_business = ""
    try:
     e_business = job_list[i].find_element_by_class_name("int").text                  #业务方向
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有公司业务方向信息可获取")
    except:
     self.ui.log("获取公司业务方向信息出现其它错误!")
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    e_msg = self.open_URL(e_link, 4)        #职位描述和要求
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    ws.append(self.append_list_job(self.de_illegal(e), self.de_illegal(e_l), self.de_illegal(e_d), self.de_illegal(e_fl),
                                   self.de_illegal(e_c), self.de_illegal(e_link), self.de_illegal(e_type_size),
                                   self.de_illegal(e_business), self.de_illegal(e_time), self.de_illegal(e_msg)))

   while 1:
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    try:
     wb.save(file_name)       #保存数据到data文件
    except PermissionError:
     self.ui.log('表格被其它程序占用中，无法写入数据! 请关闭占用程序')
     tk.messagebox.showwarning("警告", "表格被其它程序占用中，请关闭占用程序")  #提示框
     self.ui.log('10秒后重试')
     self.pause()         #提供暂停
     if self.mStop():       #提供停止
      return
     time.sleep(10)
    else:
     break
   self.ui.log('保存成功,下一页')
   time.sleep(1)
   while 1:
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    try:
     driver.find_element_by_class_name("j_page").find_element_by_class_name("next").click()    #下一页
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("找不到点击下一页的按钮  休眠15秒后重试!")
     self.pause()         #提供暂停
     if self.mStop():       #提供停止
      return
     time.sleep(15)
    else:
     break
   self.ui.log('准备爬取当前页面  休眠中zzz...')
   self.pause()         #提供暂停
   if self.mStop():       #提供停止
    return
   time.sleep(sleepTime)        #根据设定间隔休眠

  self.ui.log('结束爬取循环!')
  self.ui.log('保存data文件中...')
  wb.save(file_name)
  self.isFirst = True  #重置第一次循环标志
  startFlag = False
  stopFlag = False
  runFlag = False
  self.ui.openBossButton.config(state=NORMAL)
  self.ui.openJobButton.config(state=NORMAL)
  self.ui.startButton.config(text = '开始爬取')
  self.ui.spinbox.config(state=NORMAL)
  self.ui.notGDButton.config(state=NORMAL)
  self.ui.checkButton.config(state=NORMAL)
  self.ui.startStopVirtualBtn.config(state=NORMAL)
  self.ui.log("爬虫已停止")
#-------------------------------------------------以上为爬取前程无忧网页-------------------------------------------------------------------------------------------------------------

 def append_list_boss(self, ee, ee_l, ee_a, ee_d, ee_fl, ee_c, ee_link, ee_t_s, ee_b, ee_time, ee_msg):  #职位名称  薪资  地区  经验|学历  福利  公司名称  链接  公司类型|大小  业务方向 发布时间 职位描述和要求
  l = []
  sal_low = 0
  sal_hight = 0
  sal_avg = 0
  self.pause()         #提供暂停
  if ee_d.find('个月') != -1:       #实习职位,在职位名称后添加实习时间
   if ee.find('实习') != -1:
    ee = ee + '(' + ee_d[:ee_d.find('个月') + 2] + ')'
   else:
    ee = ee + '(实习' + ee_d[:ee_d.find('个月') + 2] + ')'
   ee_d = '在校生/应届生'
  l.append(ee.replace(' ',''))
  ee_l = ee_l.replace(' ','')
  if ee_l.find('·') != -1:        #薪水含有13薪
   ee_fl = ee_fl + ' ' + ee_l[ee_l.find('·')+1:]    #将13薪放在福利后面
  if ee_l.find('-') == -1:          #薪酬字符串中不含-
   if ee_l.find('元/天') != -1:
    sal_low = int(ee_l.split('元/天')[0].strip()) * 22              #统一标准单位元/月
    sal_hight = int(ee_l.split('元/天')[0].strip()) * 26
    sal_avg = (sal_low + sal_hight) // 2
    l.append(sal_low)
    l.append(sal_hight)
    l.append(sal_avg)
   elif ee_l.find('K') != -1:
    sal_low = int(float(ee_l.split('K')[0].strip()) * 1000)             #统一标准单位元/月
    sal_hight = sal_low
    sal_avg = sal_low
    l.append(sal_low)
    l.append(sal_hight)
    l.append(sal_avg)
   elif ee_l.find('元/时') != -1:
    sal_low = int(float(ee_l.split('元/时')[0].strip()) * 8 * 26)             #统一标准单位元/月
    sal_hight = sal_low
    sal_avg = sal_low
    l.append(sal_low)
    l.append(sal_hight)
    l.append(sal_avg)
   else:
    l.append(ee_l.strip())
    l.append("null")
    l.append("null")
  else:
   if ee_l.find('K') != -1:
    try:
     sal_low = int(float(ee_l.split('K')[0].split('-')[0]) * 1000)            #统一标准单位元/月
     sal_hight = int(float(ee_l.split('K')[0].split('-')[1]) * 1000)
     sal_avg = (sal_low + sal_hight) // 2
     l.append(sal_low)
     l.append(sal_hight)
     l.append(sal_avg)
    except:
     l.append(ee_l.strip())
     l.append("null")
     l.append("null")
   elif ee_l.find('元/天') != -1:
    try:
     sal_low = int(float(ee_l.split('元/天')[0].split('-')[0]) * 26)            #统一标准单位元/月
     sal_hight = int(float(ee_l.split('元/天')[0].split('-')[1]) * 26)
     sal_avg = (sal_low + sal_hight) // 2
     l.append(sal_low)
     l.append(sal_hight)
     l.append(sal_avg)
    except:
     l.append(ee_l.strip())
     l.append("null")
     l.append("null")
   elif ee_l.find('元/时') != -1:
    try:
     sal_low = int(float(ee_l.split('元/时')[0].split('-')[0]) * 8 * 26)            #统一标准单位元/月
     sal_hight = int(float(ee_l.split('元/时')[0].split('-')[1]) * 8 * 26)
     sal_avg = (sal_low + sal_hight) // 2
     l.append(sal_low)
     l.append(sal_hight)
     l.append(sal_avg)
    except:
     l.append(ee_l.strip())
     l.append("null")
     l.append("null")
   else:
    l.append(ee_l.strip())
    l.append("null")
    l.append("null")
        
  self.pause()         #提供暂停
  if ee_a.find('·') != -1:
   l.append(ee_a.split('·')[0] + '-' + ee_a.split('·')[1])      #公司位置
  else:
   l.append(ee_a)

  if ee_d.find('学历不限') != -1:
   l.append(ee_d.split('学历不限')[0])        #经验要求
   l.append("学历不限")                       #学历要求
  elif ee_d.find('大专') != -1:
   l.append(ee_d.split('大专')[0])
   l.append("大专")
  elif ee_d.find('中专') != -1:
   l.append(ee_d.split('中专')[0])
   l.append("中专/中技")
  elif ee_d.find('高中') != -1:
   l.append(ee_d.split('高中')[0])
   l.append("高中")
  elif ee_d.find('初中及以下') != -1:
   l.append(ee_d.split('初中及以下')[0])
   l.append("初中及以下")
  elif ee_d.find('本科') != -1:
   l.append(ee_d.split('本科')[0])
   l.append("本科")
  elif ee_d.find('硕士') != -1:
   l.append(ee_d.split('硕士')[0])
   l.append("硕士")
  elif ee_d.find('博士') != -1:
   l.append(ee_d.split('博士')[0])
   l.append("博士")
  else:
   l.append(ee_d.strip())
   l.append("null")
        
  self.pause()         #提供暂停
  l.append(ee_fl.strip())               #福利
  l.append(ee_c.strip())                #公司名称
  l.append('https://www.zhipin.com' + ee_link.strip())    #链接地址
  tmp_s = ee_t_s.split(ee_b.strip())[1]
  if tmp_s.find('已上市') != -1:          #判断是否已上市
   l.append('已上市')              #公司类型
  else:
   l.append('未上市')
  index = 0
  while index < len(tmp_s):
   if tmp_s[index].isdigit():      #判断数字字符最初出现的位置
    break
   else:
    index += 1
  if index < len(tmp_s):
   l.append(tmp_s[index:len(tmp_s)])   #通过第一个数字字符位置分割字符串   公司大小
  else:
   l.append("")
  l.append(ee_b.strip())                            #公司定位业务方向

  if ee_time.find('发布于') != -1:      #发布日期
   if ee_time.find('月') != -1 and ee_time.find('日') != -1:
    l.append(int(ee_time.replace('发布于','').replace('月','').replace('日','')))  #日期转整数
   elif ee_time.find(':') != -1:
    l.append(int(time.strftime('%m%d', time.localtime())))    #获取当天日期转整数
   elif ee_time.find('昨天') != -1:
    l.append(int(time.strftime('%m%d', time.localtime(time.time() - 60*60*24))))    #获取昨天日期转整数
   else:
    l.append(ee_time.strip())
  else:
   l.append(ee_time.strip())
  l.append(ee_msg.replace('\n',' ').replace('\t',' ').replace('【',' ').replace('】',' ').replace('ü','').replace('●','').replace('★','').replace('◆','').replace('*','').replace('■','').replace('  ','')) #职位描述和要求
  self.ui.log('---' + ee + '---')   #打印职位
  return l

 def get_boss(self):
  global sw
  global smkw
  global tsw
  global driver
  global file_name
  global sleepTime
  global stopFlag
  global startFlag
  global runFlag
  global wb
  global ws
  self.current_url = driver.current_url
  while (self.isFirst or driver.current_url != self.current_url) and driver.current_url.find('zhipin.com') != -1:  #当前页面与上一页面不同，并且URL包含zhipin.com
   self.pause()            #提供暂停
   if self.mStop():       #提供停止
    return
   self.isFirst = False     #第一次循环设否
   self.ui.log("爬取当前页面中---------------------")
   self.current_url = driver.current_url
   while True:
    self.pause()     #提供暂停
    if self.mStop():       #提供停止
     return
    try:
     job_list = driver.find_element_by_class_name('job-list').find_elements_by_class_name('job-primary')   #获取列表
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("无法获取职位列表 休眠15秒后重试!")
     self.pause()     #提供暂停
     if self.mStop():       #提供停止
      return
     time.sleep(15)
    else:
     if len(job_list) > 0:                 #获取列表成功
      break
     else:                                 #列表长度有误
      self.ui.log("错误：职位列表长度为0!  休眠10秒后重试")
      self.pause()     #提供暂停
      if self.mStop():       #提供停止
       return
      time.sleep(10)
    
   for i in range(len(job_list)):
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    e = job_list[i].find_element_by_class_name("job-name").text      #获取工作名称
    e_a = job_list[i].find_element_by_class_name("job-area").text    #获取公司地区
    tmp = 0
    while tmp < len(sw):
     self.pause()         #提供暂停
     if self.mStop():       #提供停止
      return
     if e.find(sw[tmp]) != -1:
      break             #包含停用词，跳到下一工作
     else:
      tmp += 1
    if tmp < len(sw):
     continue           #包含停用词，跳到下一工作
    e_tag = ""
    try:
     e_tag = job_list[i].find_element_by_class_name("info-append").find_element_by_class_name("tags").text   #职位标签
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有职位标签可获取")
    except:
     self.ui.log("获取职位标签出现其它错误!")
    else:
     if e_tag != "":
      tmp = 0
      while tmp < len(tsw):
       self.pause()         #提供暂停
       if self.mStop():       #提供停止
        return
       if e_tag.find(tsw[tmp]) != -1:
        break             #包含标签停用词，跳到下一工作
       else:
        tmp += 1
      if tmp < len(tsw):
       continue           #包含标签停用词，跳到下一工作
    tmp = 0
    while tmp < len(smkw):
     self.pause()         #提供暂停
     if self.mStop():       #提供停止
      return
     if e.find(smkw[tmp]) != -1:
      break             #都不包含关键词，跳到下一工作
     else:
      tmp += 1
    if tmp == len(smkw):
     continue           #都不包含关键词，跳到下一工作
    e_l = job_list[i].find_element_by_class_name("job-limit").find_element_by_class_name("red").text  #薪酬
    e_d = ""
    try:
     e_d = job_list[i].find_element_by_class_name("job-limit").find_element_by_tag_name('p').text   #经验与学历
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有经验与学历要求可获取")
    except:
     self.ui.log("获取经验与学历要求出现其它错误!")
    e_c = job_list[i].find_element_by_class_name("company-text").find_element_by_class_name("name").text   #公司名称
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    e_fl = ""
    try:
     e_fl = job_list[i].find_element_by_class_name("info-append").find_element_by_class_name("info-desc").text   #福利
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有福利信息可获取")
    except:
     self.ui.log("获取福利信息出现其它错误!")
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    e_link = job_list[i].find_element_by_class_name("primary-box").get_attribute('href')    #详情链接地址
    e_type_size = ""
    try:
     e_type_size = job_list[i].find_element_by_class_name("company-text").find_element_by_tag_name("p").text  #公司类型和大小
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有公司类型和大小信息可获取")
    except:
     self.ui.log("获取公司类型和大小信息出现其它错误!")
    e_business = ""
    try:
     e_business = job_list[i].find_element_by_class_name("company-text").find_element_by_tag_name("p").find_element_by_tag_name("a").text    #公司业务定位
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有公司业务定位信息可获取")
    except:
     self.ui.log("获取公司业务定位信息出现其它错误!")
    e_time = ""
    try:
     e_time = job_list[i].find_element_by_class_name("job-pub-time").text    #发布时间
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有发布时间可获取")
    except:
     self.ui.log("获取发布时间出现其它错误!")
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    tmp_index = 5           #获取职位描述和要求最多有5次尝试机会
    tmp_sleep_time = 0
    e_msg = ""
    while tmp_index > 0:
     primary_box = job_list[i].find_element_by_class_name("primary-box")                      #获取primary-box元素
     ActionChains(driver).move_to_element(primary_box).perform()                              #移动鼠标到primary-box元素
     time.sleep(1)
     try:
      e_msg = job_list[i].find_element_by_class_name("info-detail").find_element_by_class_name("detail-bottom-text").text      #获取职位描述和要求
     except se.common.exceptions.NoSuchElementException:
      tmp_sleep_time += 5
      self.ui.log('无法获取职位描述和要求!  ' + str(tmp_sleep_time) + '秒后重试')
      driver.execute_script('window.scrollTo(0,document.body.scrollHeight)')      #跳到网页底部
      time.sleep(tmp_sleep_time)    #休眠
      tmp_index -= 1              #次数减一
     else:
      break
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    ws.append(self.append_list_boss(self.de_illegal(e), self.de_illegal(e_l), self.de_illegal(e_a), self.de_illegal(e_d),
                                    self.de_illegal(e_fl), self.de_illegal(e_c), self.de_illegal(e_link), self.de_illegal(e_type_size),
                                    self.de_illegal(e_business), self.de_illegal(e_time), self.de_illegal(e_msg)))

   while 1:
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    try:
     wb.save(file_name)       #保存数据到data文件
    except PermissionError:
     self.ui.log('表格被其它程序占用中，无法写入数据! 请关闭占用程序')
     tk.messagebox.showwarning("警告", "表格被其它程序占用中，请关闭占用程序")  #提示框
     self.ui.log('10秒后重试')
     self.pause()         #提供暂停
     if self.mStop():       #提供停止
      return
     time.sleep(10)
    else:
     break
   self.ui.log('保存成功,下一页')
   time.sleep(1)
   while 1:
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    try:
     driver.find_element_by_class_name("page").find_element_by_class_name("next").click()    #下一页
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("找不到点击下一页的按钮  休眠15秒后重试!")
     self.pause()         #提供暂停
     if self.mStop():       #提供停止
      return
     time.sleep(15)
    except se.common.exceptions.ElementClickInterceptedException:
     self.ui.log('下一页按钮无法按下!')
     break
    else:
     self.ui.log('准备爬取当前页面  休眠中zzz...')
     time.sleep(sleepTime)        #根据设定间隔休眠
     break
   
   self.pause()         #提供暂停
   if self.mStop():       #提供停止
    return

  self.ui.log('结束爬取循环!')
  self.ui.log('保存data文件中...')
  wb.save(file_name)
  self.isFirst = True  #重置第一次循环标志
  startFlag = False
  stopFlag = False
  runFlag = False
  self.ui.openBossButton.config(state=NORMAL)
  self.ui.openJobButton.config(state=NORMAL)
  self.ui.startButton.config(text = '开始爬取')
  self.ui.spinbox.config(state=NORMAL)
  self.ui.notGDButton.config(state=NORMAL)
  self.ui.checkButton.config(state=NORMAL)
  self.ui.startStopVirtualBtn.config(state=NORMAL)
  self.ui.log("爬虫已停止")


#---------------------------------------------------------------------以上为爬取Boss直聘------------------------------------------------------------------------------------------------------------
def p_self(ws_, range_min, range_max, queue_, now_time, isCheckGD):  #自查重子进程
 for i in range(range_min, range_max):
  if delete_not_gd(ws_['E'+str(i)].value, isCheckGD) == False:
   queue_.put(i)
   continue
  for j in range(i+1, range_max):
   if ws_['A'+str(i)].value != 'delete' and ws_['A'+str(i)].value.replace(' ','') == ws_['A'+str(j)].value.replace(' ','') and ws_['I'+str(i)].value == ws_['I'+str(j)].value and ws_['E'+str(i)].value == ws_['E'+str(j)].value and ws_['F'+str(i)].value == ws_['F'+str(j)].value: #职位名称 公司名称 地区 经验要求均相同
    if ws_['N'+str(i)].value == ws_['N'+str(j)].value:   #发布日期相同，包括今日
     tmp_rate = None
     if ws_['O'+str(i)].value != None and ws_['O'+str(j)].value != None:
      tmp_rate = match_rate(ws_['O'+str(i)].value, ws_['O'+str(j)].value)  #获取匹配率
     elif ws_['O'+str(i)].value == None and ws_['O'+str(j)].value == None:
      tmp_rate = match_rate("", "")  #获取匹配率
     elif ws_['O'+str(i)].value == None and ws_['O'+str(j)].value != None:
      tmp_rate = match_rate("", ws_['O'+str(j)].value)  #获取匹配率
     elif ws_['O'+str(i)].value != None and ws_['O'+str(j)].value == None:
      tmp_rate = match_rate(ws_['O'+str(i)].value, "")  #获取匹配率
     if tmp_rate == 1.0:          #匹配率100%相同
      queue_.put(i)    #职位描述及要求相同
      #print('删除重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
     elif tmp_rate >= 0.70:  #匹配率大于70%
      if (ws_['H'+str(i)].value == None and ws_['H'+str(j)].value == None) or (ws_['H'+str(i)].value != None and ws_['H'+str(j)].value != None):
       if len(ws_['O'+str(i)].value) > len(ws_['O'+str(j)].value):
        queue_.put(j)   #删除字数小的行
       elif len(ws_['O'+str(i)].value) <= len(ws_['O'+str(j)].value):
        queue_.put(i)  #删除字数小的行
      elif ws_['H'+str(i)].value == None:
       queue_.put(i)    #删除福利为空的行
      elif ws_['H'+str(j)].value == None:
       queue_.put(j)    #删除福利为空的行
     elif tmp_rate >= 0.60:   #匹配率小于70%大于等于60%，需自行比对
      P_tag_list = []
      P_tag_list.append(i)
      P_tag_list.append(j)
      P_tag_list.append(' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
      queue_.put(P_tag_list)
     break
    elif ws_['N'+str(i)].value <= now_time and ws_['N'+str(j)].value <= now_time:     #小于等于今日日期 今年1月1日到今日且发布日不相同
     if ws_['N'+str(i)].value > ws_['N'+str(j)].value:
      queue_.put(j)
     else:
      queue_.put(i)
    elif ws_['N'+str(i)].value > now_time and ws_['N'+str(j)].value > now_time:     #大于今日日期 去年明日到去年12月31日且发布日不相同
     if ws_['N'+str(i)].value > ws_['N'+str(j)].value:
      queue_.put(j)
     else:
      queue_.put(i)
    elif ws_['N'+str(i)].value <= now_time and ws_['N'+str(j)].value > now_time:   #一个今年，另一个去年  发布日不相同
     queue_.put(j)
    elif ws_['N'+str(i)].value > now_time and ws_['N'+str(j)].value <= now_time:   #一个去年，另一个今年  发布日不相同
     queue_.put(i)
    break

def p_each(ws_, i_range_min, i_range_max, j_range_min, j_range_max, q, now_time):  #对查重子进程
 for i in range(i_range_min, i_range_max):
  for j in range(j_range_min, j_range_max):
   if ws_['A'+str(i)].value != 'delete' and ws_['A'+str(i)].value.replace(' ','') == ws_['A'+str(j)].value.replace(' ','') and ws_['I'+str(i)].value == ws_['I'+str(j)].value and ws_['E'+str(i)].value == ws_['E'+str(j)].value and ws_['F'+str(i)].value == ws_['F'+str(j)].value: #职位名称 公司名称 地区 经验要求均相同
    if ws_['N'+str(i)].value == ws_['N'+str(j)].value:   #发布日期相同，包括今日
     tmp_rate = None
     if ws_['O'+str(i)].value != None and ws_['O'+str(j)].value != None:
      tmp_rate = match_rate(ws_['O'+str(i)].value, ws_['O'+str(j)].value)  #获取匹配率
     elif ws_['O'+str(i)].value == None and ws_['O'+str(j)].value == None:
      tmp_rate = match_rate("", "")  #获取匹配率
     elif ws_['O'+str(i)].value == None and ws_['O'+str(j)].value != None:
      tmp_rate = match_rate("", ws_['O'+str(j)].value)  #获取匹配率
     elif ws_['O'+str(i)].value != None and ws_['O'+str(j)].value == None:
      tmp_rate = match_rate(ws_['O'+str(i)].value, "")  #获取匹配率
     if tmp_rate == 1.0:          #匹配率100%相同
      q.put(i)    #职位描述及要求相同
      #print('删除重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
     elif tmp_rate >= 0.70:  #匹配率大于70%
      if (ws_['H'+str(i)].value == None and ws_['H'+str(j)].value == None) or (ws_['H'+str(i)].value != None and ws_['H'+str(j)].value != None):
       if len(ws_['O'+str(i)].value) > len(ws_['O'+str(j)].value):
        q.put(j)   #删除字数小的行
       elif len(ws_['O'+str(i)].value) <= len(ws_['O'+str(j)].value):
        q.put(i)  #删除字数小的行
      elif ws_['H'+str(i)].value == None:
       q.put(i)    #删除福利为空的行
      elif ws_['H'+str(j)].value == None:
       q.put(j)    #删除福利为空的行
     elif tmp_rate >= 0.60:   #匹配率小于70%大于等于60%，需自行比对
      P_tag_list = []
      P_tag_list.append(i)
      P_tag_list.append(j)
      P_tag_list.append(' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
      q.put(P_tag_list)
     break
    elif ws_['N'+str(i)].value <= now_time and ws_['N'+str(j)].value <= now_time:     #小于等于今日日期 今年1月1日到今日且发布日不相同
     if ws_['N'+str(i)].value > ws_['N'+str(j)].value:
      q.put(j)
     else:
      q.put(i)
    elif ws_['N'+str(i)].value > now_time and ws_['N'+str(j)].value > now_time:     #大于今日日期 去年明日到去年12月31日且发布日不相同
     if ws_['N'+str(i)].value > ws_['N'+str(j)].value:
      q.put(j)
     else:
      q.put(i)
    elif ws_['N'+str(i)].value <= now_time and ws_['N'+str(j)].value > now_time:   #一个今年，另一个去年  发布日不相同
     q.put(j)
    elif ws_['N'+str(i)].value > now_time and ws_['N'+str(j)].value <= now_time:   #一个去年，另一个今年  发布日不相同
     q.put(i)
    break

def match_rate(str1, str2):  #获取相似率
 s_1 = copy.deepcopy(str1)
 s_2 = copy.deepcopy(str2)
 rate = 0 #吻合度
 del_char = [' ','、','“','”','；','，','（','）','。','《','》','【','】','：','！','￥','……','…','？',
             '——','—','·','(',')','{','}','|','[',']',':',';','"','\'','\\','<','>','?',',','.','/',
             '!','`','~','*','+','$','&','%','#','@','^','-','\n']  #要去除的字符列表
 for each_del_char in del_char:
  s_1 = s_1.replace(each_del_char, '')  #去除以上字符
  s_2 = s_2.replace(each_del_char, '')
 min_num = 0   #小端
 l_1 = jieba.lcut(s_1) #分词
 l_2 = jieba.lcut(s_2)
 set_1 = set(l_1) #集合去重复
 set_2 = set(l_2)
 unit = set_1 & set_2   #取交集
 if len(set_1) < len(set_2):
  min_num = len(set_1)
 else:
  min_num = len(set_2)
 if min_num == 0:
  rate = 1.0
 else:
  rate = len(unit)/min_num
 del s_1, s_2, set_1, set_2, unit, min_num
 return rate

def delete_not_gd(str_area, isCheckGD):   #检查地区是否为非广东
 if isCheckGD == True:   #去除非广东数据
  if str_area == None:
   return True
  gd_city_name = ['广州', '深圳', '东莞', '佛山', '珠海', '惠州', '中山', '江门', '汕头', '湛江', '肇庆', '揭阳', '清远', '潮州', '梅州', '茂名', '韶关', '阳江', '河源', '汕尾', '云浮', '广东']
  index_tmp_city = 0
  while index_tmp_city < 22:
   if str_area.find(gd_city_name[index_tmp_city]) == -1:
    index_tmp_city += 1
   else:
    break
  if index_tmp_city < 22:   #广东城市
   return True
  else:                    #非广东城市
   return False
 else:
  return True

class Checking():
 def __init__(self, ui):
  self.ui = ui
  self.schedule = 0

 def update_program_bar(self, now_num, max_num):
  self.ui.canvas.coords(self.ui.fill_rec, (4,4, int(now_num / max_num * 311) + 4, 22))  #设置进度条宽度
  self.ui.update()

 def start_update_program_bar(self, now_n, max_n):
  if self.schedule != int(now_n / max_n * 311):  #进度是否变化
   self.schedule = int(now_n / max_n * 311)
   threading.Thread(target=self.update_program_bar, args=(now_n, max_n)).start()     #创建一个新线程更新进度条的显示

 def stop_check(self):
  global isStopChecking
  global isChecking
  global force_exit
  if force_exit:
   self.ui.log('查重停止!')
   return True
  if isStopChecking and isChecking:  #查重运行中
   if tk.messagebox.askyesno(title = '提示', message = '是否要停止查重？'):   #确定要停止
    self.ui.log('查重停止!')
    return True
   else:
    isStopChecking = False
    self.ui.log('取消停止，继续查重')
    return False
  else:
   return False

 def finish_check(self):
  global isChecking
  global isStopChecking
  global force_exit

  isChecking = False
  isStopChecking = False
  if not force_exit:
   self.ui.checkButton.config(text = '开始查重')  #UI界面恢复正常
   self.ui.stopButton.config(state=NORMAL)
   self.ui.startButton.config(state=NORMAL)
   self.ui.openGoogleButton.config(state=NORMAL)
   self.ui.openBossButton.config(state=NORMAL)
   self.ui.openJobButton.config(state=NORMAL)
   self.ui.spinbox.config(state=NORMAL)
   self.ui.notGDButton.config(state=NORMAL)
   self.ui.startStopVirtualBtn.config(state=NORMAL)
   self.ui.log('-------------查重结束!---------------')
   self.ui.log('若要再次查重请重启软件，以提高准确率!!!\n')
   self.schedule = 0
   self.ui.canvas.coords(self.ui.fill_rec, (4,4,4,22))  #恢复进度条宽度
   self.ui.canvas.itemconfig(self.ui.fill_rec, fill = '#fafafa')  #恢复进度条颜色
   self.ui.update()  #更新显示


 def do_check(self):
  global wb
  global ws
  global file_name
  global checkGD
  free_memory = psutil.virtual_memory().free/1024/1024    #获取系统空闲内存(MB)
  start_time = time.time()
  max_r = ws.max_row    #获取最大行数
  now_time = int(time.strftime('%m%d', time.localtime()))   #获取当前日期
  tmp = 0
  P_tag = []                #标记列表
  P_index = -1              #标记列表下标
  delete_row = []             #需要删除的行号
  self.ui.log('查重中...')
  pool = None
  queue = None
  p_n = 0
  if max_r >= 768:  #多进程
   row_per_MB = max_r/5960+45
   mem = max_r/row_per_MB              #计算所有招聘数据行所占内存(MB)
   p_n = multiprocessing.cpu_count()   #获取cpu逻辑核心个数
   while p_n > 0:
    if p_n*mem*1.12+72.5 < free_memory:          #分析最多可以使用cpu多少核心
     break
    p_n -= 1
  if p_n > 1:      #多进程
   pool = multiprocessing.Pool(p_n-1)
   queue = multiprocessing.Manager().Queue()
   self.ui.log('多进程生成中，请勿进行操作!')
   for x in range(2, p_n+1):
    pool.apply_async(p_self, (ws, (x-1)*max_r//p_n+1, x*max_r//p_n+1, queue, now_time, checkGD))   #多进程自查重
   if self.stop_check():   #提供停止
    self.finish_check()    #结束查重
    pool.close()
    pool.terminate()    #结束子进程
    pool.join()
    return
   time.sleep(int(p_n*mem*1.1/170))        #休眠等待子进程创建完毕
   if self.stop_check():   #提供停止
    self.finish_check()    #结束查重
    pool.close()
    pool.terminate()    #结束子进程
    pool.join()
    return
   self.ui.log('多进程生成成功!')
   for i in range(2, max_r//p_n + 1):
    self.start_update_program_bar(i//2, max_r//p_n)   #更新进度条，第一部分
    for j in range(i+1, max_r//p_n + 1):
     if self.stop_check():   #提供停止
      self.finish_check()    #结束查重
      pool.close()
      pool.terminate()    #结束子进程
      pool.join()
      return
     if ws['A'+str(i)].value != None and ws['A'+str(j)].value != None and ws['A'+str(i)].value.replace(' ','') == ws['A'+str(j)].value.replace(' ','') and ws['I'+str(i)].value == ws['I'+str(j)].value and ws['E'+str(i)].value == ws['E'+str(j)].value and ws['F'+str(i)].value == ws['F'+str(j)].value: #职位名称 公司名称 地区 经验要求均相同
      if ws['N'+str(i)].value == ws['N'+str(j)].value:   #发布日期相同，包括今日
       tmp_rate = None
       if ws['O'+str(i)].value != None and ws['O'+str(j)].value != None:
        tmp_rate = match_rate(ws['O'+str(i)].value, ws['O'+str(j)].value)  #获取匹配率
       elif ws['O'+str(i)].value == None and ws['O'+str(j)].value == None:
        tmp_rate = match_rate("", "")  #获取匹配率
       elif ws['O'+str(i)].value == None and ws['O'+str(j)].value != None:
        tmp_rate = match_rate("", ws['O'+str(j)].value)  #获取匹配率
       elif ws['O'+str(i)].value != None and ws['O'+str(j)].value == None:
        tmp_rate = match_rate(ws['O'+str(i)].value, "")  #获取匹配率
       if tmp_rate == 1.0:          #匹配率100%相同
        delete_row.append(i)    #职位描述及要求相同
        self.ui.log('删除重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
        tmp += 1
       elif tmp_rate >= 0.70:  #匹配率大于70%
        if (ws['H'+str(i)].value == None and ws['H'+str(j)].value == None) or (ws['H'+str(i)].value != None and ws['H'+str(j)].value != None):
         if len(ws['O'+str(i)].value) > len(ws['O'+str(j)].value):
          delete_row.append(j)   #删除字数小的行
          self.ui.log('重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
          tmp += 1
         elif len(ws['O'+str(i)].value) <= len(ws['O'+str(j)].value):
          delete_row.append(i)  #删除字数小的行
          self.ui.log('重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
          tmp += 1
        elif ws['H'+str(i)].value == None:
         delete_row.append(i)    #删除福利为空的行
         self.ui.log('重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
         tmp += 1
        elif ws['H'+str(j)].value == None:
         delete_row.append(j)    #删除福利为空的行
         self.ui.log('重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
         tmp += 1
       elif tmp_rate >= 0.60:   #匹配率小于70%大于等于60%，需自行比对
        P_tag.append(' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')  #先存储入列表保存 差别字符为 + d_str
        P_index += 1
        ws['P'+str(i)] = P_index   #作P标记
        ws['P'+str(j)] = P_index
        #self.ui.log('-------------->\n需自行删除的重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '% 差别字符为' + d_str)
       break
      elif ws['N'+str(i)].value <= now_time and ws['N'+str(j)].value <= now_time:     #小于等于今日日期 今年1月1日到今日且发布日不相同
       if ws['N'+str(i)].value > ws['N'+str(j)].value:
        delete_row.append(j)
       else:
        delete_row.append(i)
      elif ws['N'+str(i)].value > now_time and ws['N'+str(j)].value > now_time:     #大于今日日期 去年明日到去年12月31日且发布日不相同
       if ws['N'+str(i)].value > ws['N'+str(j)].value:
        delete_row.append(j)
       else:
        delete_row.append(i)
      elif ws['N'+str(i)].value <= now_time and ws['N'+str(j)].value > now_time:   #一个今年，另一个去年  发布日不相同
       delete_row.append(j)
      elif ws['N'+str(i)].value > now_time and ws['N'+str(j)].value <= now_time:   #一个去年，另一个今年  发布日不相同
       delete_row.append(i)
      self.ui.log('过时行检查: ' + str(i) + ' 和 ' + str(j))
      tmp += 1
      break
     if delete_not_gd(ws['E'+str(i)].value, checkGD) == False:
      delete_row.append(i)
   pool.close()
   pool.join()
   pool = None
   pool = multiprocessing.Pool(p_n-1)
   for x in range(1, p_n+1):
    for y in range(x+1, p_n+1):
     #self.ui.log(str(x)+','+str(y))
     if x == 1:
      if y != 2:
       pool.apply_async(p_each, (ws, 2, x*max_r//p_n+1, (y-1)*max_r//p_n+1, y*max_r//p_n+1, queue, now_time))  #多进程对查重
     else:
      pool.apply_async(p_each, (ws, (x-1)*max_r//p_n+1, x*max_r//p_n+1, (y-1)*max_r//p_n+1, y*max_r//p_n+1, queue, now_time))   #多进程对查重
   if self.stop_check():   #提供停止
    self.finish_check()    #结束查重
    pool.close()
    pool.terminate()    #结束子进程
    pool.join()
    return
   time.sleep(int(p_n*mem*1.1/170))        #休眠等待子进程创建完毕
   if self.stop_check():   #提供停止
    self.finish_check()    #结束查重
    pool.close()
    pool.terminate()    #结束子进程
    pool.join()
    return
   need_sleep_time = 0
   if (p_n*(p_n-1)-2)/2/(p_n-1) == (p_n*(p_n-1)-2)//2//(p_n-1):
    need_sleep_time = (p_n*(p_n-1)-2)//2//(p_n-1)/320000*max_r/p_n
   else:
    need_sleep_time = ((p_n*(p_n-1)-2)//2//(p_n-1)+1)/320000*max_r/p_n
#   self.ui.log(str(need_sleep_time))
   for i in range(2, max_r//p_n+1):
    self.start_update_program_bar(i//2+max_r//p_n//2, max_r//p_n)   #更新进度条，第一部分
    if p_n > 2:
     time.sleep(need_sleep_time)   #休眠以匹配子进程进度
    for j in range(max_r//p_n+1, 2*max_r//p_n + 1):
     if self.stop_check():   #提供停止
      self.finish_check()    #结束查重
      pool.close()
      pool.terminate()   #结束子进程
      pool.join()
      return
     if ws['A'+str(i)].value != None and ws['A'+str(j)].value != None and ws['A'+str(i)].value.replace(' ','') == ws['A'+str(j)].value.replace(' ','') and ws['I'+str(i)].value == ws['I'+str(j)].value and ws['E'+str(i)].value == ws['E'+str(j)].value and ws['F'+str(i)].value == ws['F'+str(j)].value: #职位名称 公司名称 地区 经验要求均相同
      if ws['N'+str(i)].value == ws['N'+str(j)].value:   #发布日期相同，包括今日
       tmp_rate = None
       if ws['O'+str(i)].value != None and ws['O'+str(j)].value != None:
        tmp_rate = match_rate(ws['O'+str(i)].value, ws['O'+str(j)].value)  #获取匹配率
       elif ws['O'+str(i)].value == None and ws['O'+str(j)].value == None:
        tmp_rate = match_rate("", "")  #获取匹配率
       elif ws['O'+str(i)].value == None and ws['O'+str(j)].value != None:
        tmp_rate = match_rate("", ws['O'+str(j)].value)  #获取匹配率
       elif ws['O'+str(i)].value != None and ws['O'+str(j)].value == None:
        tmp_rate = match_rate(ws['O'+str(i)].value, "")  #获取匹配率
       if tmp_rate == 1.0:          #匹配率100%相同
        delete_row.append(i)    #职位描述及要求相同
        self.ui.log('删除重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
        tmp += 1
       elif tmp_rate >= 0.70:  #匹配率大于70%
        if (ws['H'+str(i)].value == None and ws['H'+str(j)].value == None) or (ws['H'+str(i)].value != None and ws['H'+str(j)].value != None):
         if len(ws['O'+str(i)].value) > len(ws['O'+str(j)].value):
          delete_row.append(j)   #删除字数小的行
          self.ui.log('重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
          tmp += 1
         elif len(ws['O'+str(i)].value) <= len(ws['O'+str(j)].value):
          delete_row.append(i)  #删除字数小的行
          self.ui.log('重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
          tmp += 1
        elif ws['H'+str(i)].value == None:
         delete_row.append(i)    #删除福利为空的行
         self.ui.log('重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
         tmp += 1
        elif ws['H'+str(j)].value == None:
         delete_row.append(j)    #删除福利为空的行
         self.ui.log('重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
         tmp += 1
       elif tmp_rate >= 0.60:   #匹配率小于70%大于等于60%，需自行比对
        P_tag.append(' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')  #先存储入列表保存 差别字符为 + d_str
        P_index += 1
        ws['P'+str(i)] = P_index   #作P标记
        ws['P'+str(j)] = P_index
        #self.ui.log('-------------->\n需自行删除的重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '% 差别字符为' + d_str)
       break
      elif ws['N'+str(i)].value <= now_time and ws['N'+str(j)].value <= now_time:     #小于等于今日日期 今年1月1日到今日且发布日不相同
       if ws['N'+str(i)].value > ws['N'+str(j)].value:
        delete_row.append(j)
       else:
        delete_row.append(i)
      elif ws['N'+str(i)].value > now_time and ws['N'+str(j)].value > now_time:     #大于今日日期 去年明日到去年12月31日且发布日不相同
       if ws['N'+str(i)].value > ws['N'+str(j)].value:
        delete_row.append(j)
       else:
        delete_row.append(i)
      elif ws['N'+str(i)].value <= now_time and ws['N'+str(j)].value > now_time:   #一个今年，另一个去年  发布日不相同
       delete_row.append(j)
      elif ws['N'+str(i)].value > now_time and ws['N'+str(j)].value <= now_time:   #一个去年，另一个今年  发布日不相同
       delete_row.append(i)
      self.ui.log('过时行检查: ' + str(i) + ' 和 ' + str(j))
      tmp += 1
      break
     if delete_not_gd(ws['E'+str(i)].value, checkGD) == False:
      delete_row.append(i)
   pool.close()
   pool.join()
  else:  #单进程
   for i in range(2, max_r + 1):
    self.start_update_program_bar(i/3, max_r/3)   #更新进度条，第一部分
    for j in range(i+1, max_r + 1):
     if self.stop_check():   #提供停止
      self.finish_check()    #结束查重
      return
     if ws['A'+str(i)].value != None and ws['A'+str(j)].value != None and ws['A'+str(i)].value.replace(' ','') == ws['A'+str(j)].value.replace(' ','') and ws['I'+str(i)].value == ws['I'+str(j)].value and ws['E'+str(i)].value == ws['E'+str(j)].value and ws['F'+str(i)].value == ws['F'+str(j)].value: #职位名称 公司名称 地区 经验要求均相同
      if ws['N'+str(i)].value == ws['N'+str(j)].value:   #发布日期相同，包括今日
       tmp_rate = None
       if ws['O'+str(i)].value != None and ws['O'+str(j)].value != None:
        tmp_rate = match_rate(ws['O'+str(i)].value, ws['O'+str(j)].value)  #获取匹配率
       elif ws['O'+str(i)].value == None and ws['O'+str(j)].value == None:
        tmp_rate = match_rate("", "")  #获取匹配率
       elif ws['O'+str(i)].value == None and ws['O'+str(j)].value != None:
        tmp_rate = match_rate("", ws['O'+str(j)].value)  #获取匹配率
       elif ws['O'+str(i)].value != None and ws['O'+str(j)].value == None:
        tmp_rate = match_rate(ws['O'+str(i)].value, "")  #获取匹配率
       if tmp_rate == 1.0:          #匹配率100%相同
        delete_row.append(i)    #职位描述及要求相同
        self.ui.log('删除重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
        tmp += 1
       elif tmp_rate >= 0.70:  #匹配率大于70%
        if (ws['H'+str(i)].value == None and ws['H'+str(j)].value == None) or (ws['H'+str(i)].value != None and ws['H'+str(j)].value != None):
         if len(ws['O'+str(i)].value) > len(ws['O'+str(j)].value):
          delete_row.append(j)   #删除字数小的行
          self.ui.log('重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
          tmp += 1
         elif len(ws['O'+str(i)].value) <= len(ws['O'+str(j)].value):
          delete_row.append(i)  #删除字数小的行
          self.ui.log('重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
          tmp += 1
        elif ws['H'+str(i)].value == None:
         delete_row.append(i)    #删除福利为空的行
         self.ui.log('重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
         tmp += 1
        elif ws['H'+str(j)].value == None:
         delete_row.append(j)    #删除福利为空的行
         self.ui.log('重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
         tmp += 1
       elif tmp_rate >= 0.60:   #匹配率小于70%大于等于60%，需自行比对
        P_tag.append(' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')  #先存储入列表保存 差别字符为 + d_str
        P_index += 1
        ws['P'+str(i)] = P_index   #作P标记
        ws['P'+str(j)] = P_index
        #self.ui.log('-------------->\n需自行删除的重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '% 差别字符为' + d_str)
       break
      elif ws['N'+str(i)].value <= now_time and ws['N'+str(j)].value <= now_time:     #小于等于今日日期 今年1月1日到今日且发布日不相同
       if ws['N'+str(i)].value > ws['N'+str(j)].value:
        delete_row.append(j)
       else:
        delete_row.append(i)
      elif ws['N'+str(i)].value > now_time and ws['N'+str(j)].value > now_time:     #大于今日日期 去年明日到去年12月31日且发布日不相同
       if ws['N'+str(i)].value > ws['N'+str(j)].value:
        delete_row.append(j)
       else:
        delete_row.append(i)
      elif ws['N'+str(i)].value <= now_time and ws['N'+str(j)].value > now_time:   #一个今年，另一个去年  发布日不相同
       delete_row.append(j)
      elif ws['N'+str(i)].value > now_time and ws['N'+str(j)].value <= now_time:   #一个去年，另一个今年  发布日不相同
       delete_row.append(i)
      self.ui.log('过时行检查: ' + str(i) + ' 和 ' + str(j))
      tmp += 1
      break
     if delete_not_gd(ws['E'+str(i)].value, checkGD) == False:
      delete_row.append(i)

  if pool != None:
#   self.ui.log('queue大小'+str(queue.qsize()))
   while queue.qsize() != 0:
    if self.stop_check():   #提供停止
     self.finish_check()    #结束查重
     return
    tmp_ = queue.get()
    if type(tmp_) == list:
     P_index += 1
     P_tag.append(tmp_[2])
     ws['P'+str(tmp_[0])] = P_index
     ws['P'+str(tmp_[1])] = P_index
    else:
     tmp += 1
     delete_row.append(tmp_)
  self.ui.log('查找到重复过时或相似'+str(tmp)+'个')

  if tmp > 0:
   if self.stop_check():   #提供停止
    self.finish_check()    #结束查重
    return
   self.ui.log('删除中...')
   delete_row_set = set(delete_row)   #行号去重复
   delete_row = list(delete_row_set)
   del delete_row_set
   delete_row.sort()    #行号排序从小到大
   d_i = 0
   d_r_len = len(delete_row)
   for k in range(d_r_len):
    if self.stop_check():   #提供停止
     self.finish_check()    #结束查重
     return
    self.start_update_program_bar(k+1, d_r_len)   #更新进度条，第二部分
    ws.delete_rows(delete_row[k]-d_i)   #删除行
    d_i += 1
   self.ui.log('删除成功!')
  max_r = ws.max_row    #获取最大行数
  for i in range(2, max_r + 1):      #查找有没有P标记
   if self.stop_check():   #提供停止
    self.finish_check()    #结束查重
    return
   self.start_update_program_bar(i+1, max_r)   #更新进度条，第三部分
   if ws['P'+str(i)].value != None:
    P_index = ws['P'+str(i)].value
    for j in range(i+1, max_r + 1):
     if ws['P'+str(j)].value == P_index:  #查找到两行都有同一P标记
      self.ui.log('-------------->\n需自行删除的重复行: ' + str(i) + ' 和 ' + str(j) + P_tag[P_index])  #打印出需要自行对比删除行号
      ws['P'+str(i)] = None    #清除标记
      ws['P'+str(j)] = None
      break
     elif j == max_r and ws['P'+str(j)].value != P_index:
      ws['P'+str(i)] = None    #清除标记
  end_time = time.time()
  self.ui.log('查重完成! 用时' + str(int((end_time-start_time)/3600)) + '时' + str(int((end_time-start_time)%3600/60)) + '分' + str(int((end_time-start_time)%60)) + '秒')
  wb.save(file_name)
  tk.messagebox.showinfo("提示", "查重工作已完成!")
  self.finish_check()  #结束查重

#-------------------------------------------------------------------------------以上为查重------------------------------------------------------------------------------------------------------

def gz_Map_p(ws_, kw_list, max_num, min_num, html_name, absolute_kw, q):   #广州（各类）职位数量分布  多进程函数
 gz_path = os.path.abspath('..') + '\\广州'
 if not os.path.exists(gz_path):  #判断广州文件夹是否存在
  os.mkdir(gz_path)
 value = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
 area = ['白云区', '越秀区', '海珠区', '天河区', '番禺区', '黄埔区', '南沙区', '花都区', '从化区', '增城区', '荔湾区']
 max_r = ws_.max_row
 for i in range(2, max_r+1):
  if ws_['A'+str(i)].value != None:
   tmp = ws_['A'+str(i)].value
   j = 0
   while j < len(kw_list):
    if kw_list[j] in absolute_kw and kw_list[j] in jieba.lcut(tmp, cut_all = False):  #判断是否含有关键词
     if kw_list[j] == 'VR' and '摄影' not in kw_list and tmp.find('摄影') != -1: #VR摄影师不属于VR
      j+= 1
     else:
      break
    elif tmp.find(kw_list[j]) != -1:  #判断是否为相应职位
     break
    else:
     j += 1
   if j < len(kw_list):
    tmp_e = ws_['E'+str(i)].value
    if tmp_e != None:
     for area_index in range(len(area)):
      if tmp_e.find(area[area_index][:2]) != -1:
       value[area_index] += 1
       break

 gz_map = Map(init_opts = opts.InitOpts(width = '900px', height = '700px'))
 gz_map.add("", [list(z) for z in zip(area, value)], '广州', zoom = 1.2)
 gz_map.set_global_opts(title_opts = opts.TitleOpts(title = html_name, pos_left = 'center', pos_top = '2'),
                    visualmap_opts = opts.VisualMapOpts(max_= (lambda v: max(value) if v == None else max_num)(max_num), min_= (lambda v: min(value) if v == None else min_num)(min_num), is_piecewise = False),
                    toolbox_opts = opts.ToolboxOpts(is_show = True, feature = {'saveAsImage': {}}))
 gz_map.set_series_opts(label_opts = opts.LabelOpts(is_show = True, formatter = '{b} {c}'))
 gz_map.render(gz_path + '\\' + html_name + '.html')
 if not os.path.exists(gz_path + '\\guang3_dong1_guang3_zhou1.js'):  #广州文件夹下不存在guang3_dong1_guang3_zhou1.js文件
  try:
   js_f = open(os.getcwd() + '\\guang3_dong1_guang3_zhou1.js', 'r', encoding = 'utf-8')
   js_f.seek(0)
   js_str = js_f.read()
   js_f.close()
   js_f = open(gz_path + '\\guang3_dong1_guang3_zhou1.js', 'w', encoding = 'utf-8')  #拷贝guang3_dong1_guang3_zhou1.js文件到广州文件夹
   js_f.write(js_str)
   js_f.close()
   del(js_f, js_str)
  except Exception as e:
   q.put('拷贝文件guang3_dong1_guang3_zhou1.js出错:'+str(e))
   return
 time.sleep(2)
 html_f = open(gz_path + '\\' + html_name + '.html', 'r', encoding = 'utf-8')
 html_f.seek(0)      #指针指到开头
 html_str = html_f.read()
 html_str = html_str.replace('https://assets.pyecharts.org/assets/maps/guang3_dong1_guang3_zhou1.js', 'guang3_dong1_guang3_zhou1.js')  #地图轮廓文件改到本地获取
 html_f.close()
 html_f = open(gz_path + '\\' + html_name + '.html', 'w+', encoding = 'utf-8')
 html_f.seek(0)      #指针指到开头
 html_f.truncate()   #清空文件
 html_f.write(html_str)
 html_f.close()
 q.put(html_name + '图生成成功!!!')

def sz_Pie_p(ws_, kw_list, html_name, absolute_kw, q):   #深圳（各类）职位数量分布  多进程函数
 sz_path = os.path.abspath('..') + '\\深圳'
 if not os.path.exists(sz_path):  #判断深圳文件夹是否存在
  os.mkdir(sz_path)
 value = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
 area = ['南山区', '福田区', '罗湖区', '龙岗区', '龙华区', '宝安区', '光明新区', '坪山新区', '盐田区', '大鹏新区']
 max_r = ws_.max_row
 for i in range(2, max_r+1):
  if ws_['A'+str(i)].value != None:
   tmp = ws_['A'+str(i)].value
   j = 0
   while j < len(kw_list):
    if kw_list[j] in absolute_kw and kw_list[j] in jieba.lcut(tmp, cut_all = False):  #判断是否含有关键词
     if kw_list[j] == 'VR' and '摄影' not in kw_list and tmp.find('摄影') != -1: #VR摄影师不属于VR
      j+= 1
     else:
      break
    elif tmp.find(kw_list[j]) != -1:  #判断是否为相应职位
     break
    else:
     j += 1
   if j < len(kw_list):
    tmp_e = ws_['E'+str(i)].value
    if tmp_e != None:
     for area_index in range(len(area)):
      if tmp_e.find(area[area_index][:2]) != -1:
       value[area_index] += 1
       break

 sz_pie = Pie(init_opts = opts.InitOpts(width = '700px', height = '700px'))
 sz_pie.add('', [list(z) for z in zip(area, value)], radius = ['40%', '70%'])
 sz_pie.set_global_opts(title_opts = opts.TitleOpts(title = html_name, pos_left = 'center', pos_top = '2'),
                    legend_opts = opts.LegendOpts(is_show = False),
                    toolbox_opts = opts.ToolboxOpts(is_show = True, feature = {'saveAsImage': {}}))
 sz_pie.set_series_opts(label_opts = opts.LabelOpts(is_show = True, formatter = '{b} {c}'))
 sz_pie.render(sz_path + '\\' + html_name + '.html')
 q.put(html_name + '图生成成功!!!')

def gd_Money_Bar_p(ws_, kw_list, q):    #多进程函数
 count = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]  #各城市薪水个数计数器
 sum_avg = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] #D 平均薪水总和
 sum_max = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] #C 最高薪水总和
 sum_min = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] #B 最低薪水总和
 city = ['广州', '深圳', '东莞', '佛山', '珠海', '惠州', '中山', '江门', '汕头', '湛江', '肇庆', '揭阳', '清远', '潮州', '梅州', '茂名', '韶关', '阳江', '河源', '汕尾', '云浮']
 max_r = ws_.max_row
 for i in range(2, max_r+1):
  if ws_['A'+str(i)].value != None:
   tmp = ws_['A'+str(i)].value
   j = 0
   while j < len(kw_list):
    if tmp.find(kw_list[j]) != -1:  #判断是否为数媒行业
     break
    else:
     j += 1
   if j < len(kw_list):   #是数媒行业
    tmp_e = ws_['E'+str(i)].value
    for each_city in range(len(city)):
     if tmp_e != None and tmp_e.find(city[each_city]) != -1:   #是否为广东省城市
      if ws_['B'+str(i)].value != None and ws_['C'+str(i)].value != 'null' and ws_['D'+str(i)].value != 'null' and type(ws_['B'+str(i)].value) != str:
       sum_min[each_city] += ws_['B'+str(i)].value
       sum_max[each_city] += ws_['C'+str(i)].value
       sum_avg[each_city] += ws_['D'+str(i)].value
       count[each_city] += 1
      break
 avg_avg = []
 avg_max = []
 avg_min = []
 for z in range(len(count)):
  if count[z] != 0:
   avg_avg.append(sum_avg[z]//count[z])
   avg_max.append(sum_max[z]//count[z])
   avg_min.append(sum_min[z]//count[z])
  else:
   avg_avg.append(0)
   avg_max.append(0)
   avg_min.append(0)
 del(sum_avg, sum_max, sum_min, count)
 line = Line(init_opts = opts.InitOpts(width = '1400px', height = '700px'))
 line.add_xaxis(city)
 line.add_yaxis('平均最高薪酬', avg_max)
 line.add_yaxis('平均薪酬', avg_avg)
 line.add_yaxis('平均最低薪酬', avg_min)
 line.set_global_opts(title_opts = opts.TitleOpts(title = '广东城市数媒行业职位薪水对比', pos_left = '70', pos_top = '2'),
                    toolbox_opts = opts.ToolboxOpts(is_show = True, feature = {'saveAsImage': {}}),
                    yaxis_opts = opts.AxisOpts(min_ = min(min(avg_avg, avg_max, avg_min))-500, axislabel_opts=opts.LabelOpts(formatter = '{value}元/月')),
                     xaxis_opts = opts.AxisOpts(splitline_opts = opts.SplitLineOpts(is_show = True)))
 line.set_series_opts(label_opts = opts.LabelOpts(is_show = True, position = 'top', color = 'black'))
 line.render(os.path.abspath('..') + '\\广东城市数媒行业职位薪水对比.html')
 q.put('广东城市数媒行业职位薪水对比图生成成功!!!')

def eachJob_Money_Bar_p(ws_, kw_list_list, absolute_kw, q):   #  多进程函数
 count = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]  #各职位薪水个数计数器
 sum_avg = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] #D 平均薪水总和
 sum_max = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] #C 最高薪水总和
 sum_min = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] #B 最低薪水总和
 max_r = ws_.max_row
 for i in range(2, max_r+1):
  if ws_['A'+str(i)].value != None:
   tmp = ws_['A'+str(i)].value
   for each_index in range(16):  #判断是那个职位
    j = 0
    while j < len(kw_list_list[each_index]):
     if kw_list_list[each_index][j] in absolute_kw and kw_list_list[each_index][j] in jieba.lcut(tmp, cut_all = False):  #判断是否含有关键词
      if kw_list_list[each_index][j] == 'VR' and '摄影' not in kw_list_list[each_index] and tmp.find('摄影') != -1: #VR摄影师不属于VR
       j+= 1
      else:
       break
     elif tmp.find(kw_list_list[each_index][j]) != -1:  #判断是否为相应职位
      break
     else:
      j += 1
    if j < len(kw_list_list[each_index]):  #是数媒行业的哪一职位
     if ws_['B'+str(i)].value != None and ws_['C'+str(i)].value != 'null' and ws_['D'+str(i)].value != 'null' and type(ws_['B'+str(i)].value) != str:
      sum_min[each_index] += ws_['B'+str(i)].value
      sum_max[each_index] += ws_['C'+str(i)].value
      sum_avg[each_index] += ws_['D'+str(i)].value
      count[each_index] += 1
     break
    
 avg_avg = []
 avg_max = []
 avg_min = []
 for z in range(16):
  if count[z] != 0:
   avg_avg.append(sum_avg[z]//count[z])
   avg_max.append(sum_max[z]//count[z])
   avg_min.append(sum_min[z]//count[z])
  else:
   avg_avg.append(0)
   avg_max.append(0)
   avg_min.append(0)
 del(sum_avg, sum_max, sum_min, count)
 job = ['Android', '动画', 'C#', 'C++', '游戏', '平面', 'Java', '媒体', '建模', '摄影摄像', '影视后期', '数据库', '数媒技术', 'Unity', '虚拟现实', 'Web网页前端']
 job_bar = Bar(init_opts = opts.InitOpts(width = '1200px', height = '800px'))
 job_bar.add_xaxis(job)
 job_bar.add_yaxis('平均最高薪酬', avg_max)
 job_bar.add_yaxis('平均薪酬', avg_avg)
 job_bar.add_yaxis('平均最低薪酬', avg_min)
 job_bar.reversal_axis()     #反转坐标轴
 job_bar.set_global_opts(title_opts = opts.TitleOpts(title = '数媒行业各职位薪水对比', pos_left = '100', pos_top = '1'),
                    toolbox_opts = opts.ToolboxOpts(is_show = True, feature = {'saveAsImage': {}}),
                    yaxis_opts = opts.AxisOpts(splitline_opts = opts.SplitLineOpts(is_show = True)),
                    xaxis_opts = opts.AxisOpts(min_ = min(min(avg_avg, avg_max, avg_min))-1000, axislabel_opts=opts.LabelOpts(formatter = '{value}元/月')))
 job_bar.set_series_opts(label_opts = opts.LabelOpts(is_show = True, position = 'right', color = 'black'))
 job_bar.render(os.path.abspath('..') + '\\数媒行业各职位薪水对比.html')
 q.put('数媒行业各职位薪水对比图生成成功!!!')

def eachJob_Exp_Pie_p(ws_, kw_list, html_name, absolute_kw, q):   #  多进程函数
 jy_path = os.path.abspath('..') + '\\经验要求'
 if not os.path.exists(jy_path):  #判断经验要求文件夹是否存在
  os.mkdir(jy_path)
 exp = [0, 0, 0, 0, 0, 0, 0, 0]  #各经验要求
 max_r = ws_.max_row
 for i in range(2, max_r+1):
  if ws_['A'+str(i)].value != None:
   tmp = ws_['A'+str(i)].value
   j = 0
   while j < len(kw_list):
    if kw_list[j] in absolute_kw and kw_list[j] in jieba.lcut(tmp, cut_all = False):  #判断是否含有关键词
     if kw_list[j] == 'VR' and '摄影' not in kw_list and tmp.find('摄影') != -1:      #VR摄影师不属于VR
      j+= 1
     else:
      break
    elif tmp.find(kw_list[j]) != -1:  #判断是否为相应职位
     break
    else:
     j += 1
   if j < len(kw_list):  #是否为相应职位
    if ws_['F'+str(i)].value != None and ws_['F'+str(i)].value.find('生') != -1:
     exp[0] += 1
    elif ws_['F'+str(i)].value != None and (ws_['F'+str(i)].value.find('无需') != -1 or ws_['F'+str(i)].value.find('不限') != -1):
     exp[1] += 1
    elif ws_['F'+str(i)].value != None and ws_['F'+str(i)].value.find('1年') != -1:
     exp[2] += 1
    elif ws_['F'+str(i)].value != None and ws_['F'+str(i)].value.find('2年') != -1:
     exp[3] += 1
    elif ws_['F'+str(i)].value != None and (ws_['F'+str(i)].value.find('3-4年') != -1 or ws_['F'+str(i)].value.find('3-5年') != -1):
     exp[4] += 1
    elif ws_['F'+str(i)].value != None and ws_['F'+str(i)].value.find('5-7年') != -1:
     exp[5] += 1
    elif ws_['F'+str(i)].value != None and ws_['F'+str(i)].value.find('8-9年') != -1:
     exp[6] += 1
    elif ws_['F'+str(i)].value != None and ws_['F'+str(i)].value.find('10年') != -1:
     exp[7] += 1
    elif ws_['F'+str(i)].value != None and ws_['F'+str(i)].value.find('1-3年') != -1:
     exp[2] += 1
     exp[3] += 1
    elif ws_['F'+str(i)].value != None and ws_['F'+str(i)].value.find('5-10年') != -1:
     exp[5] += 1
     exp[6] += 1
      
 exp_list = ['在校生/应届生', '经验不限', '1年经验', '2年经验', '3-4年经验', '5-7年经验', '8-9年经验', '10年以上经验']
 exp_pie = Pie(init_opts = opts.InitOpts(width = '800px', height = '600px'))
 exp_pie.add('', [list(z) for z in zip(exp_list, exp)], radius = ['0%', '70%'])
 exp_pie.set_global_opts(title_opts = opts.TitleOpts(title = html_name, pos_left = 'center', pos_top = '2'),
                    legend_opts = opts.LegendOpts(is_show = False),
                    toolbox_opts = opts.ToolboxOpts(is_show = True, feature = {'saveAsImage': {}}))
 exp_pie.set_series_opts(label_opts = opts.LabelOpts(is_show = True, formatter = '{b} {c}'))
 exp_pie.render(jy_path + '\\' + html_name + '.html')
 q.put(html_name + '图生成成功!!!')

def eachJob_Edu_Pie_p(ws_, kw_list, html_name, absolute_kw, q):   #  多进程函数
 xl_path = os.path.abspath('..') + '\\学历要求'
 if not os.path.exists(xl_path):  #判断学历要求文件夹是否存在
  os.mkdir(xl_path)
 edu = [0, 0, 0, 0, 0, 0, 0, 0, 0]  #各学历要求
 max_r = ws_.max_row
 for i in range(2, max_r+1):
  if ws_['A'+str(i)].value != None:
   tmp = ws_['A'+str(i)].value
   j = 0
   while j < len(kw_list):
    if kw_list[j] in absolute_kw and kw_list[j] in jieba.lcut(tmp, cut_all = False):  #判断是否含有关键词
     if kw_list[j] == 'VR' and '摄影' not in kw_list and tmp.find('摄影') != -1: #VR摄影师不属于VR
      j+= 1
     else:
      break
    elif tmp.find(kw_list[j]) != -1:  #判断是否为相应职位
     break
    else:
     j += 1
   if j < len(kw_list):  #是否为相应职位
    if ws_['G'+str(i)].value != None and ws_['G'+str(i)].value.find('初中') != -1:
     edu[0] += 1
    elif ws_['G'+str(i)].value != None and ws_['G'+str(i)].value.find('高中') != -1:
     edu[1] += 1
    elif ws_['G'+str(i)].value != None and ws_['G'+str(i)].value.find('中专/中技') != -1:
     edu[2] += 1
     edu[3] += 1
    elif ws_['G'+str(i)].value != None and ws_['G'+str(i)].value.find('中专') != -1:
     edu[2] += 1
    elif ws_['G'+str(i)].value != None and ws_['G'+str(i)].value.find('中技') != -1:
     edu[3] += 1
    elif ws_['G'+str(i)].value != None and ws_['G'+str(i)].value.find('大专') != -1:
     edu[4] += 1
    elif ws_['G'+str(i)].value != None and ws_['G'+str(i)].value.find('本科') != -1:
     edu[5] += 1
    elif ws_['G'+str(i)].value != None and ws_['G'+str(i)].value.find('硕士') != -1:
     edu[6] += 1
    elif ws_['G'+str(i)].value != None and ws_['G'+str(i)].value.find('博士') != -1:
     edu[7] += 1
    elif ws_['G'+str(i)].value == None or ws_['G'+str(i)].value.find('不限') != -1:
     edu[8] += 1

 edu_list = ['初中及以下', '高中', '中专', '中技', '大专', '本科', '硕士', '博士', '学历不限']
 edu_pie = Pie(init_opts = opts.InitOpts(width = '800px', height = '600px'))
 edu_pie.add('', [list(z) for z in zip(edu_list, edu)], radius = ['0%', '70%'])
 edu_pie.set_global_opts(title_opts = opts.TitleOpts(title = html_name, pos_left = 'center', pos_top = '2'),
                    legend_opts = opts.LegendOpts(is_show = False),
                    toolbox_opts = opts.ToolboxOpts(is_show = True, feature = {'saveAsImage': {}}))
 edu_pie.set_series_opts(label_opts = opts.LabelOpts(is_show = True, formatter = '{b} {c}'))
 edu_pie.render(xl_path + '\\' + html_name + '.html')
 q.put(html_name + '图生成成功!!!')

def eachJob_CSize_Pie_p(ws_, kw_list, html_name, absolute_kw, q):   #  多进程函数
 dx_path = os.path.abspath('..') + '\\公司大小'
 if not os.path.exists(dx_path):  #判断公司大小文件夹是否存在
  os.mkdir(dx_path)
 size = [0, 0, 0, 0, 0, 0, 0]  #各公司大小类型
 max_r = ws_.max_row
 for i in range(2, max_r+1):
  if ws_['A'+str(i)].value != None:
   tmp = ws_['A'+str(i)].value
   j = 0
   while j < len(kw_list):
    if kw_list[j] in absolute_kw and kw_list[j] in jieba.lcut(tmp, cut_all = False):  #判断是否含有关键词
     if kw_list[j] == 'VR' and '摄影' not in kw_list and tmp.find('摄影') != -1: #VR摄影师不属于VR
      j+= 1
     else:
      break
    elif tmp.find(kw_list[j]) != -1:  #判断是否为相应职位
     break
    else:
     j += 1
   if j < len(kw_list):  #是否为相应职位
    if ws_['L'+str(i)].value != None and (ws_['L'+str(i)].value.find('少于') != -1 or ws_['L'+str(i)].value.find('0-20') != -1):
     size[0] += 1
    elif ws_['L'+str(i)].value != None and (ws_['L'+str(i)].value.find('50-150') != -1 or ws_['L'+str(i)].value.find('20-99') != -1):
     size[1] += 1
    elif ws_['L'+str(i)].value != None and (ws_['L'+str(i)].value.find('150-500') != -1 or ws_['L'+str(i)].value.find('100-499') != -1):
     size[2] += 1
    elif ws_['L'+str(i)].value != None and (ws_['L'+str(i)].value.find('500-1000') != -1 or ws_['L'+str(i)].value.find('500-999') != -1):
     size[3] += 1
    elif ws_['L'+str(i)].value != None and ws_['L'+str(i)].value.find('1000-5000') != -1:
     size[4] += 1
    elif ws_['L'+str(i)].value != None and ws_['L'+str(i)].value.find('5000-10000') != -1:
     size[5] += 1
    elif ws_['L'+str(i)].value != None and ws_['L'+str(i)].value.find('以上') != -1:
     size[6] += 1
    elif ws_['L'+str(i)].value != None and ws_['L'+str(i)].value.find('1000-9999') != -1:
     size[4] += 1
     size[5] += 1

 size_list = ['少于50人', '50-150人', '150-500人', '500-1000人', '1000-5000人', '5000-10000人', '10000人以上']
 size_pie = Pie(init_opts = opts.InitOpts(width = '800px', height = '600px'))
 size_pie.add('', [list(z) for z in zip(size_list, size)], radius = ['0%', '65%'])
 size_pie.set_global_opts(title_opts = opts.TitleOpts(title = html_name, pos_left = 'center', pos_top = '2'),
                    legend_opts = opts.LegendOpts(is_show = False),
                    toolbox_opts = opts.ToolboxOpts(is_show = True, feature = {'saveAsImage': {}}))
 size_pie.set_series_opts(label_opts = opts.LabelOpts(is_show = True, formatter = '{b} {c}'))
 size_pie.render(dx_path + '\\' + html_name + '.html')
 q.put(html_name + '图生成成功!!!')

def WordCloud_p(ws_, kw_list, absolute_kw, d, index):  #多进程词云分词获取函数
 cy_path = os.path.abspath('..') + '\\职位描述和要求'
 if not os.path.exists(cy_path):  #判断职位描述和要求文件夹是否存在
  os.mkdir(cy_path)
 wordCloudStopWord = """职能 类别 良好 的 年假 以上 优先 工作 经验 任职 要求 熟练 使用 相关 周末 养老 希望 元 生育 五天 聚餐 调薪 两 每周 礼金 冰箱 午餐 猫 之间 跟进 地址 六薪 
专业 有 较 及 以上 等 关键字 一定 负责 公司 优先 考虑 独立 任职 资格 具有 具备 作者 物料 医疗 一种 职责 生活 天 月 小时 欢迎 宝贝 礼品 医院 饮料 男女 一位 公交 灾 上班 四大 双薪 
职能 好 要求 五险 一金 类别 带薪 年度 年 法定 节假日 善于 毕业 过程 中 常用 强烈 自己 团建 年终 会 岗位 带领 地铁 站 每天 期待 礼物 广场 结婚 零食 六险 尤佳 一家 失业 春节 底薪 
至少 根据 员工 解决 福利 办公 环境 不 限于 限 于 交办 常用 描述 全日制 地点 婚假 投递 简历 奖金 下午茶 薪资 羽毛球 一款 午休 全勤奖 同事 栋 丧假 男女 率低 号线 印刷 病假 邮箱 
产假 承受 参与 旅游 并 进行 晋升 遇到 加入 我们 你 我 入职 优先 住房 公积金 假期 薪酬 大小 强具 生日 1 一起 5 附上 从业 TV 工伤 定期 下车 孪生 以下 附带 分项 园区 面试 话费 
深厚 深刻 其他 其它 部门 规定 购买 社保 保险 聘 请 急 ü ● ★ ◆ ■ 丰富 富 周末 双休 者 能够 人员 具 分红 发放 广阔 一门 一次 一个 。 ° 一次 年终奖 周一 座 大厦 免费 或"""   #停用词表
 wc_sw = wordCloudStopWord.replace('\n', '').split(' ')
 max_r = ws_.max_row
 fc_tmp = ''
 for i in range(2, max_r+1):
  if ws_['A'+str(i)].value != None:
   tmp_A = ws_['A'+str(i)].value
   j = 0
   while j < len(kw_list):
    if kw_list[j] in absolute_kw and kw_list[j] in jieba.lcut(tmp_A, cut_all = False):  #判断是否含有关键词
     if kw_list[j] == 'VR' and '摄影' not in kw_list and tmp_A.find('摄影') != -1: #VR摄影师不属于VR
      j+= 1
     else:
      break
    elif tmp_A.find(kw_list[j]) != -1:  #判断是否为相应职位
     break
    else:
     j += 1
   if j < len(kw_list):  #是否为相应职位
    if ws_['O'+str(i)].value != None:   #职位描述和要求不为空
     tmp_O = ws_['O'+str(i)].value
     for z in wc_sw:
      tmp_O = tmp_O.replace(z, '')
     tmp_O = tmp_O.replace('C#', '/Csharp/').replace('c#', '/Csharp/').replace('C++', '/Cpp/').replace('c++', '/Cpp/').replace('Java', '/Java/')
     tmp_O = tmp_O.replace('Android', '/Android/').replace('Unity', '/Unity/').replace('Unity3D', '/Unity3D/').replace('U3D', '/U3D/')
     fc_tmp += ' '.join(jieba.cut(tmp_O, cut_all = False))
 d.update({index : fc_tmp})


def gz_fun(ws_, sm_kw, kw_list_list, absolute_kw, q_):
 gz_Map_p(ws_, sm_kw, 4000, None, '广州市数媒相关行业职位数量分布', absolute_kw, q_)
 gz_Map_p(ws_, kw_list_list[0], 120, 0, '广州市Android安卓相关职位数量分布', absolute_kw, q_)
 gz_Map_p(ws_, kw_list_list[1], 120, 0, '广州市动画相关职位数量分布', absolute_kw, q_)
 gz_Map_p(ws_, kw_list_list[2], 120, 0, '广州市C#相关职位数量分布', absolute_kw, q_)
 gz_Map_p(ws_, kw_list_list[3], 120, 0, '广州市C++相关职位数量分布', absolute_kw, q_)
 gz_Map_p(ws_, kw_list_list[4], 320, 0, '广州市游戏相关职位数量分布', absolute_kw, q_)
 gz_Map_p(ws_, kw_list_list[5], 1200, 0, '广州市平面相关职位数量分布', absolute_kw, q_)
 gz_Map_p(ws_, kw_list_list[6], 600, 0, '广州市Java相关职位数量分布', absolute_kw, q_)
 gz_Map_p(ws_, kw_list_list[7], 1200, 0, '广州市媒体相关职位数量分布', absolute_kw, q_)
 gz_Map_p(ws_, kw_list_list[8], 75, 0, '广州市建模相关职位数量分布', absolute_kw, q_)
 gz_Map_p(ws_, kw_list_list[9], None, 0, '广州市摄影摄像相关职位数量分布', absolute_kw, q_)
 gz_Map_p(ws_, kw_list_list[10], 600, 0, '广州市影视后期相关职位数量分布', absolute_kw, q_)
 gz_Map_p(ws_, kw_list_list[11], 30, 0, '广州市数据库相关职位数量分布', absolute_kw, q_)
 gz_Map_p(ws_, kw_list_list[12], 200, 0, '广州市数媒技术相关职位数量分布', absolute_kw, q_)
 gz_Map_p(ws_, kw_list_list[13], 55, 0, '广州市Unity相关职位数量分布', absolute_kw, q_)
 gz_Map_p(ws_, kw_list_list[14], None, 0, '广州市虚拟现实相关职位数量分布', absolute_kw, q_)
 gz_Map_p(ws_, kw_list_list[15], 500, 0, '广州市Web网页前端相关职位数量分布', absolute_kw, q_)

def sz_fun(ws_, sm_kw, kw_list_list, absolute_kw, q_, file_name_list):
 sz_Pie_p(ws_, sm_kw ,'深圳市数媒相关行业职位数量分布', absolute_kw, q_)
 for x in range(16):
  sz_Pie_p(ws_, kw_list_list[x], '深圳市'+file_name_list[x]+'相关职位数量分布', absolute_kw, q_)

def eachJob_Exp_fun(ws_, sm_kw, kw_list_list, absolute_kw, q_, file_name_list):
 eachJob_Exp_Pie_p(ws_, sm_kw ,'数媒行业职位经验要求分布', absolute_kw, q_)
 for x in range(16):
  eachJob_Exp_Pie_p(ws_, kw_list_list[x], file_name_list[x]+'相关职位经验要求分布', absolute_kw, q_)

def eachJob_Edu_fun(ws_, sm_kw, kw_list_list, absolute_kw, q_, file_name_list):
 eachJob_Edu_Pie_p(ws_, sm_kw ,'数媒行业职位学历要求分布', absolute_kw, q_)
 for x in range(16):
  eachJob_Edu_Pie_p(ws_, kw_list_list[x], file_name_list[x]+'相关职位学历要求分布', absolute_kw, q_)

def eachJob_CSize_fun(ws_, sm_kw, kw_list_list, absolute_kw, q_, file_name_list):
 eachJob_CSize_Pie_p(ws_, sm_kw ,'数媒行业职位公司大小分布', absolute_kw, q_)
 for x in range(16):
  eachJob_CSize_Pie_p(ws_, kw_list_list[x], file_name_list[x]+'相关职位公司大小分布', absolute_kw, q_)


class Visualization():  #可视化类
 def __init__(self, ui):
  self.ui = ui

 def gd_Map(self, kw_list, max_num, min_num, html_name):  #广东（各类）职位数量分布  单进程
  global ws
  global absolute_kw
  self.ui.log(html_name + '图生成中...')
  gd_path = os.path.abspath('..') + '\\广东'
  if not os.path.exists(gd_path):  #判断广东文件夹是否存在
   os.mkdir(gd_path)
  value = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
  city = ['广州市', '深圳市', '东莞市', '佛山市', '珠海市', '惠州市', '中山市', '江门市', '汕头市', '湛江市', '肇庆市', '揭阳市', '清远市', '潮州市', '梅州市', '茂名市', '韶关市', '阳江市', '河源市', '汕尾市', '云浮市']
  max_r = ws.max_row
  #absolute_kw = ['flash', 'Flash', 'FLASH', 'PS', 'Ps', 'ps', 'CG', 'cg', 'Cg', 'VR', 'Vr', 'vr', 'AR', 'Ar', 'ar']
  for i in range(2, max_r+1):
   if self.stop():  #提供停止
    return
   if ws['A'+str(i)].value != None:
    tmp = ws['A'+str(i)].value
    j = 0
    while j < len(kw_list):
     if kw_list[j] in absolute_kw and kw_list[j] in jieba.lcut(tmp, cut_all = False):  #判断是否含有关键词
      if kw_list[j] == 'VR' and '摄影' not in kw_list and tmp.find('摄影') != -1: #VR摄影师不属于VR
       j+= 1
      else:
       break
     elif tmp.find(kw_list[j]) != -1:  #判断是否为相应职位
      break
     else:
      j += 1
    if j < len(kw_list):
     tmp_e = ws['E'+str(i)].value
     if tmp_e != None:
      for city_index in range(len(city)):
       if tmp_e.find(city[city_index][:2]) != -1:
        value[city_index] += 1
        break
  
  if self.stop():  #提供停止
   return
  gd_map = Map(init_opts = opts.InitOpts(width = '900px', height = '700px')) #Map
  gd_map.add("", [list(z) for z in zip(city, value)], '广东', zoom = 1.2)
  gd_map.set_global_opts(title_opts = opts.TitleOpts(title = html_name, pos_left = 'center', pos_top = '2'),
                    visualmap_opts = opts.VisualMapOpts(max_=(lambda v: max(value) if v == None else max_num)(max_num), min_=(lambda v: min(value) if v == None else min_num)(min_num), is_piecewise = False),
                    toolbox_opts = opts.ToolboxOpts(is_show = True, feature = {'saveAsImage': {}}))
  gd_map.set_series_opts(label_opts = opts.LabelOpts(is_show = True, formatter = '{b} {c}'))
  gd_map.render(gd_path + '\\' + html_name + '.html')
  if not os.path.exists(gd_path + '\\guangdong.js'):   #广东文件夹下不存在guangdong.js文件
   try:
    js_f = open(os.getcwd() + '\\guangdong.js', 'r', encoding = 'utf-8')
    js_f.seek(0)
    js_str = js_f.read()
    js_f.close()
    js_f = open(gd_path + '\\guangdong.js', 'w', encoding = 'utf-8')  #拷贝guangdong.js文件到广东文件夹
    js_f.write(js_str)
    js_f.close()
    del(js_f, js_str)
   except Exception as e:
    self.ui.log('拷贝文件guangdong.js出错:'+str(e))
    return
  time.sleep(2)
  html_f = open(gd_path + '\\' + html_name + '.html', 'r', encoding = 'utf-8')
  html_f.seek(0)      #指针指到开头
  html_str = html_f.read()
  html_str = html_str.replace('https://assets.pyecharts.org/assets/maps/guangdong.js', 'guangdong.js')  #地图轮廓文件改到本地获取
  html_f.close()
  html_f = open(gd_path + '\\' + html_name + '.html', 'w+', encoding = 'utf-8')
  html_f.seek(0)      #指针指到开头
  html_f.truncate()   #清空文件
  html_f.write(html_str)
  html_f.close()
  self.ui.log(html_name + '图生成成功!!!')
  self.ui.log('----------------------------------------------------')

 def gz_Map(self, kw_list, max_num, min_num, html_name):   #广州（各类）职位数量分布  单进程
  global ws
  global absolute_kw
  self.ui.log(html_name + '图生成中...')
  gz_path = os.path.abspath('..') + '\\广州'
  if not os.path.exists(gz_path):  #判断广州文件夹是否存在
   os.mkdir(gz_path)
  value = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
  area = ['白云区', '越秀区', '海珠区', '天河区', '番禺区', '黄埔区', '南沙区', '花都区', '从化区', '增城区', '荔湾区']
  #absolute_kw = ['flash', 'Flash', 'FLASH', 'PS', 'Ps', 'ps', 'CG', 'cg', 'Cg', 'VR', 'Vr', 'vr', 'AR', 'Ar', 'ar']
  max_r = ws.max_row
  for i in range(2, max_r+1):
   if self.stop():  #提供停止
    return
   if ws['A'+str(i)].value != None:
    tmp = ws['A'+str(i)].value
    j = 0
    while j < len(kw_list):
     if kw_list[j] in absolute_kw and kw_list[j] in jieba.lcut(tmp, cut_all = False):  #判断是否含有关键词
      if kw_list[j] == 'VR' and '摄影' not in kw_list and tmp.find('摄影') != -1: #VR摄影师不属于VR
       j+= 1
      else:
       break
     elif tmp.find(kw_list[j]) != -1:  #判断是否为相应职位
      break
     else:
      j += 1
    if j < len(kw_list):
     tmp_e = ws['E'+str(i)].value
     if tmp_e != None:
      for area_index in range(len(area)):
       if tmp_e.find(area[area_index][:2]) != -1:
        value[area_index] += 1
        break

  if self.stop():  #提供停止
   return
  gz_map = Map(init_opts = opts.InitOpts(width = '900px', height = '700px'))
  gz_map.add("", [list(z) for z in zip(area, value)], '广州', zoom = 1.2)
  gz_map.set_global_opts(title_opts = opts.TitleOpts(title = html_name, pos_left = 'center', pos_top = '2'),
                    visualmap_opts = opts.VisualMapOpts(max_= (lambda v: max(value) if v == None else max_num)(max_num), min_= (lambda v: min(value) if v == None else min_num)(min_num), is_piecewise = False),
                    toolbox_opts = opts.ToolboxOpts(is_show = True, feature = {'saveAsImage': {}}))
  gz_map.set_series_opts(label_opts = opts.LabelOpts(is_show = True, formatter = '{b} {c}'))
  gz_map.render(gz_path + '\\' + html_name + '.html')
  if not os.path.exists(gz_path + '\\guang3_dong1_guang3_zhou1.js'):   #广州文件夹下不存在guang3_dong1_guang3_zhou1.js文件
   try:
    js_f = open(os.getcwd() + '\\guang3_dong1_guang3_zhou1.js', 'r', encoding = 'utf-8')
    js_f.seek(0)
    js_str = js_f.read()
    js_f.close()
    js_f = open(gz_path + '\\guang3_dong1_guang3_zhou1.js', 'w', encoding = 'utf-8')  #拷贝guang3_dong1_guang3_zhou1.js文件到广州文件夹
    js_f.write(js_str)
    js_f.close()
    del(js_f, js_str)
   except Exception as e:
    self.ui.log('拷贝文件guang3_dong1_guang3_zhou1.js出错:'+str(e))
    return
  time.sleep(2)
  html_f = open(gz_path + '\\' + html_name + '.html', 'r', encoding = 'utf-8')
  html_f.seek(0)      #指针指到开头
  html_str = html_f.read()
  html_str = html_str.replace('https://assets.pyecharts.org/assets/maps/guang3_dong1_guang3_zhou1.js', 'guang3_dong1_guang3_zhou1.js')  #地图轮廓文件改到本地获取
  html_f.close()
  html_f = open(gz_path + '\\' + html_name + '.html', 'w+', encoding = 'utf-8')
  html_f.seek(0)      #指针指到开头
  html_f.truncate()   #清空文件
  html_f.write(html_str)
  html_f.close()
  self.ui.log(html_name + '图生成成功!!!')
  self.ui.log('----------------------------------------------------')

 def sz_Pie(self, kw_list, html_name):   #深圳（各类）职位数量分布  单进程
  global ws
  global absolute_kw
  self.ui.log(html_name + '图生成中...')
  sz_path = os.path.abspath('..') + '\\深圳'
  if not os.path.exists(sz_path):  #判断深圳文件夹是否存在
   os.mkdir(sz_path)
  value = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
  area = ['南山区', '福田区', '罗湖区', '龙岗区', '龙华区', '宝安区', '光明新区', '坪山新区', '盐田区', '大鹏新区']
  #absolute_kw = ['flash', 'Flash', 'FLASH', 'PS', 'Ps', 'ps', 'CG', 'cg', 'Cg', 'VR', 'Vr', 'vr', 'AR', 'Ar', 'ar']
  max_r = ws.max_row
  for i in range(2, max_r+1):
   if self.stop():  #提供停止
    return
   if ws['A'+str(i)].value != None:
    tmp = ws['A'+str(i)].value
    j = 0
    while j < len(kw_list):
     if kw_list[j] in absolute_kw and kw_list[j] in jieba.lcut(tmp, cut_all = False):  #判断是否含有关键词
      if kw_list[j] == 'VR' and '摄影' not in kw_list and tmp.find('摄影') != -1: #VR摄影师不属于VR
       j+= 1
      else:
       break
     elif tmp.find(kw_list[j]) != -1:  #判断是否为相应职位
      break
     else:
      j += 1
    if j < len(kw_list):
     tmp_e = ws['E'+str(i)].value
     if tmp_e != None:
      for area_index in range(len(area)):
       if tmp_e.find(area[area_index][:2]) != -1:
        value[area_index] += 1
        break

  if self.stop():  #提供停止
   return
  sz_pie = Pie(init_opts = opts.InitOpts(width = '700px', height = '700px'))
  sz_pie.add('', [list(z) for z in zip(area, value)], radius = ['40%', '70%'])
  sz_pie.set_global_opts(title_opts = opts.TitleOpts(title = html_name, pos_left = 'center', pos_top = '2'),
                    legend_opts = opts.LegendOpts(is_show = False),
                    toolbox_opts = opts.ToolboxOpts(is_show = True, feature = {'saveAsImage': {}}))
  sz_pie.set_series_opts(label_opts = opts.LabelOpts(is_show = True, formatter = '{b} {c}'))
  sz_pie.render(sz_path + '\\' + html_name + '.html')
  self.ui.log(html_name + '图生成成功!!!')
  self.ui.log('----------------------------------------------------')

 def gd_Money_Bar(self, kw_list):
  global ws
  self.ui.log('广东城市数媒行业职位薪水对比图生成中...')
  count = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]  #各城市薪水个数计数器
  sum_avg = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] #D 平均薪水总和
  sum_max = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] #C 最高薪水总和
  sum_min = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] #B 最低薪水总和
  city = ['广州', '深圳', '东莞', '佛山', '珠海', '惠州', '中山', '江门', '汕头', '湛江', '肇庆', '揭阳', '清远', '潮州', '梅州', '茂名', '韶关', '阳江', '河源', '汕尾', '云浮']
  max_r = ws.max_row
  for i in range(2, max_r+1):
   if self.stop():  #提供停止
    return
   if ws['A'+str(i)].value != None:
    tmp = ws['A'+str(i)].value
    j = 0
    while j < len(kw_list):
     if tmp.find(kw_list[j]) != -1:  #判断是否为数媒行业
      break
     else:
      j += 1
    if j < len(kw_list):   #是数媒行业
     tmp_e = ws['E'+str(i)].value
     for each_city in range(len(city)):
      if tmp_e != None and tmp_e.find(city[each_city]) != -1:   #是否为广东省城市
       if ws['B'+str(i)].value != None and ws['C'+str(i)].value != 'null' and ws['D'+str(i)].value != 'null' and type(ws['B'+str(i)].value) != str:
        sum_min[each_city] += ws['B'+str(i)].value
        sum_max[each_city] += ws['C'+str(i)].value
        sum_avg[each_city] += ws['D'+str(i)].value
        count[each_city] += 1
       break
  avg_avg = []
  avg_max = []
  avg_min = []
  for z in range(len(count)):
   if self.stop():  #提供停止
    return
   if count[z] != 0:
    avg_avg.append(sum_avg[z]//count[z])
    avg_max.append(sum_max[z]//count[z])
    avg_min.append(sum_min[z]//count[z])
   else:
    avg_avg.append(0)
    avg_max.append(0)
    avg_min.append(0)
  del(sum_avg, sum_max, sum_min, count)
  line = Line(init_opts = opts.InitOpts(width = '1400px', height = '700px'))
  line.add_xaxis(city)
  line.add_yaxis('平均最高薪酬', avg_max)
  line.add_yaxis('平均薪酬', avg_avg)
  line.add_yaxis('平均最低薪酬', avg_min)
  if self.stop():  #提供停止
   return
  line.set_global_opts(title_opts = opts.TitleOpts(title = '广东城市数媒行业职位薪水对比', pos_left = '70', pos_top = '2'),
                    toolbox_opts = opts.ToolboxOpts(is_show = True, feature = {'saveAsImage': {}}),
                    yaxis_opts = opts.AxisOpts(min_ = min(min(avg_avg, avg_max, avg_min))-500, axislabel_opts=opts.LabelOpts(formatter = '{value}元/月')),
                     xaxis_opts = opts.AxisOpts(splitline_opts = opts.SplitLineOpts(is_show = True)))
  line.set_series_opts(label_opts = opts.LabelOpts(is_show = True, position = 'top', color = 'black'))
  line.render(os.path.abspath('..') + '\\广东城市数媒行业职位薪水对比.html')
  self.ui.log('广东城市数媒行业职位薪水对比图生成成功!!!')
  self.ui.log('----------------------------------------------------')

 def eachJob_Money_Bar(self, kw_list_list):
  global ws
  global absolute_kw
  self.ui.log('数媒行业各职位薪水对比图生成中...')
  count = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]  #各职位薪水个数计数器
  sum_avg = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] #D 平均薪水总和
  sum_max = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] #C 最高薪水总和
  sum_min = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0] #B 最低薪水总和
  #absolute_kw = ['flash', 'Flash', 'FLASH', 'PS', 'Ps', 'ps', 'CG', 'cg', 'Cg', 'VR', 'Vr', 'vr', 'AR', 'Ar', 'ar']
  max_r = ws.max_row
  for i in range(2, max_r+1):
   if self.stop():  #提供停止
    return
   if ws['A'+str(i)].value != None:
    tmp = ws['A'+str(i)].value
    for each_index in range(16):  #判断是那个职位
     j = 0
     while j < len(kw_list_list[each_index]):
      if kw_list_list[each_index][j] in absolute_kw and kw_list_list[each_index][j] in jieba.lcut(tmp, cut_all = False):  #判断是否含有关键词
       if kw_list_list[each_index][j] == 'VR' and '摄影' not in kw_list_list[each_index] and tmp.find('摄影') != -1: #VR摄影师不属于VR
        j+= 1
       else:
        break
      elif tmp.find(kw_list_list[each_index][j]) != -1:  #判断是否为相应职位
       break
      else:
       j += 1
     if j < len(kw_list_list[each_index]):  #是数媒行业的哪一职位
      if ws['B'+str(i)].value != None and ws['C'+str(i)].value != 'null' and ws['D'+str(i)].value != 'null' and type(ws['B'+str(i)].value) != str:
       sum_min[each_index] += ws['B'+str(i)].value
       sum_max[each_index] += ws['C'+str(i)].value
       sum_avg[each_index] += ws['D'+str(i)].value
       count[each_index] += 1
      break
    
  avg_avg = []
  avg_max = []
  avg_min = []
  for z in range(16):
   #self.ui.log(str(sum_avg[z])+' '+str(sum_max[z])+' '+str(sum_min[z])+' '+str(count[z]))
   if self.stop():  #提供停止
    return
   if count[z] != 0:
    avg_avg.append(sum_avg[z]//count[z])
    avg_max.append(sum_max[z]//count[z])
    avg_min.append(sum_min[z]//count[z])
   else:
    avg_avg.append(0)
    avg_max.append(0)
    avg_min.append(0)
  del(sum_avg, sum_max, sum_min, count)
  job = ['Android', '动画', 'C#', 'C++', '游戏', '平面', 'Java', '媒体', '建模', '摄影摄像', '影视后期', '数据库', '数媒技术', 'Unity', '虚拟现实', 'Web网页前端']
  job_bar = Bar(init_opts = opts.InitOpts(width = '1200px', height = '800px'))
  job_bar.add_xaxis(job)
  job_bar.add_yaxis('平均最高薪酬', avg_max)
  job_bar.add_yaxis('平均薪酬', avg_avg)
  job_bar.add_yaxis('平均最低薪酬', avg_min)
  job_bar.reversal_axis()     #反转坐标轴
  if self.stop():  #提供停止
   return
  job_bar.set_global_opts(title_opts = opts.TitleOpts(title = '数媒行业各职位薪水对比', pos_left = '100', pos_top = '1'),
                    toolbox_opts = opts.ToolboxOpts(is_show = True, feature = {'saveAsImage': {}}),
                    yaxis_opts = opts.AxisOpts(splitline_opts = opts.SplitLineOpts(is_show = True)),
                    xaxis_opts = opts.AxisOpts(min_ = min(min(avg_avg, avg_max, avg_min))-1000, axislabel_opts=opts.LabelOpts(formatter = '{value}元/月')))
  job_bar.set_series_opts(label_opts = opts.LabelOpts(is_show = True, position = 'right', color = 'black'))
  job_bar.render(os.path.abspath('..') + '\\数媒行业各职位薪水对比.html')
  self.ui.log('数媒行业各职位薪水对比图生成成功!!!')
  self.ui.log('----------------------------------------------------')

 def eachJob_Exp_Pie(self, kw_list, html_name):
  global ws
  global absolute_kw
  self.ui.log(html_name + '图生成中...')
  jy_path = os.path.abspath('..') + '\\经验要求'
  if not os.path.exists(jy_path):  #判断经验要求文件夹是否存在
   os.mkdir(jy_path)
  exp = [0, 0, 0, 0, 0, 0, 0, 0]  #各经验要求
  #absolute_kw = ['flash', 'Flash', 'FLASH', 'PS', 'Ps', 'ps', 'CG', 'cg', 'Cg', 'VR', 'Vr', 'vr', 'AR', 'Ar', 'ar']
  max_r = ws.max_row
  for i in range(2, max_r+1):
   if self.stop():  #提供停止
    return
   if ws['A'+str(i)].value != None:
    tmp = ws['A'+str(i)].value
    j = 0
    while j < len(kw_list):
     if kw_list[j] in absolute_kw and kw_list[j] in jieba.lcut(tmp, cut_all = False):  #判断是否含有关键词
      if kw_list[j] == 'VR' and '摄影' not in kw_list and tmp.find('摄影') != -1:      #VR摄影师不属于VR
       j+= 1
      else:
       break
     elif tmp.find(kw_list[j]) != -1:  #判断是否为相应职位
      break
     else:
      j += 1
    if j < len(kw_list):  #是否为相应职位
     if ws['F'+str(i)].value != None and ws['F'+str(i)].value.find('生') != -1:
      exp[0] += 1
     elif ws['F'+str(i)].value != None and (ws['F'+str(i)].value.find('无需') != -1 or ws['F'+str(i)].value.find('不限') != -1):
      exp[1] += 1
     elif ws['F'+str(i)].value != None and ws['F'+str(i)].value.find('1年') != -1:
      exp[2] += 1
     elif ws['F'+str(i)].value != None and ws['F'+str(i)].value.find('2年') != -1:
      exp[3] += 1
     elif ws['F'+str(i)].value != None and (ws['F'+str(i)].value.find('3-4年') != -1 or ws['F'+str(i)].value.find('3-5年') != -1):
      exp[4] += 1
     elif ws['F'+str(i)].value != None and ws['F'+str(i)].value.find('5-7年') != -1:
      exp[5] += 1
     elif ws['F'+str(i)].value != None and ws['F'+str(i)].value.find('8-9年') != -1:
      exp[6] += 1
     elif ws['F'+str(i)].value != None and ws['F'+str(i)].value.find('10年') != -1:
      exp[7] += 1
     elif ws['F'+str(i)].value != None and ws['F'+str(i)].value.find('1-3年') != -1:
      exp[2] += 1
      exp[3] += 1
     elif ws['F'+str(i)].value != None and ws['F'+str(i)].value.find('5-10年') != -1:
      exp[5] += 1
      exp[6] += 1
      
  exp_list = ['在校生/应届生', '经验不限', '1年经验', '2年经验', '3-4年经验', '5-7年经验', '8-9年经验', '10年以上经验']
  if self.stop():  #提供停止
   return
  exp_pie = Pie(init_opts = opts.InitOpts(width = '800px', height = '600px'))
  exp_pie.add('', [list(z) for z in zip(exp_list, exp)], radius = ['0%', '70%'])
  exp_pie.set_global_opts(title_opts = opts.TitleOpts(title = html_name, pos_left = 'center', pos_top = '2'),
                    legend_opts = opts.LegendOpts(is_show = False),
                    toolbox_opts = opts.ToolboxOpts(is_show = True, feature = {'saveAsImage': {}}))
  exp_pie.set_series_opts(label_opts = opts.LabelOpts(is_show = True, formatter = '{b} {c}'))
  exp_pie.render(jy_path + '\\' + html_name + '.html')
  self.ui.log(html_name + '图生成成功!!!')
  self.ui.log('----------------------------------------------------')

 def eachJob_Edu_Pie(self, kw_list, html_name):
  global ws
  global absolute_kw
  self.ui.log(html_name + '图生成中...')
  xl_path = os.path.abspath('..') + '\\学历要求'
  if self.stop():  #提供停止
   return
  if not os.path.exists(xl_path):  #判断学历要求文件夹是否存在
   os.mkdir(xl_path)
  edu = [0, 0, 0, 0, 0, 0, 0, 0, 0]  #各学历要求
  #absolute_kw = ['flash', 'Flash', 'FLASH', 'PS', 'Ps', 'ps', 'CG', 'cg', 'Cg', 'VR', 'Vr', 'vr', 'AR', 'Ar', 'ar']
  max_r = ws.max_row
  for i in range(2, max_r+1):
   if self.stop():  #提供停止
    return
   if ws['A'+str(i)].value != None:
    tmp = ws['A'+str(i)].value
    j = 0
    while j < len(kw_list):
     if kw_list[j] in absolute_kw and kw_list[j] in jieba.lcut(tmp, cut_all = False):  #判断是否含有关键词
      if kw_list[j] == 'VR' and '摄影' not in kw_list and tmp.find('摄影') != -1: #VR摄影师不属于VR
       j+= 1
      else:
       break
     elif tmp.find(kw_list[j]) != -1:  #判断是否为相应职位
      break
     else:
      j += 1
    if j < len(kw_list):  #是否为相应职位
     if ws['G'+str(i)].value != None and ws['G'+str(i)].value.find('初中') != -1:
      edu[0] += 1
     elif ws['G'+str(i)].value != None and ws['G'+str(i)].value.find('高中') != -1:
      edu[1] += 1
     elif ws['G'+str(i)].value != None and ws['G'+str(i)].value.find('中专') != -1:
      edu[2] += 1
     elif ws['G'+str(i)].value != None and ws['G'+str(i)].value.find('中技') != -1:
      edu[3] += 1
     elif ws['G'+str(i)].value != None and ws['G'+str(i)].value.find('大专') != -1:
      edu[4] += 1
     elif ws['G'+str(i)].value != None and ws['G'+str(i)].value.find('本科') != -1:
      edu[5] += 1
     elif ws['G'+str(i)].value != None and ws['G'+str(i)].value.find('硕士') != -1:
      edu[6] += 1
     elif ws['G'+str(i)].value != None and ws['G'+str(i)].value.find('博士') != -1:
      edu[7] += 1
     elif ws['G'+str(i)].value == None or ws['G'+str(i)].value.find('不限') != -1:
      edu[8] += 1
  edu_list = ['初中及以下', '高中', '中专', '中技', '大专', '本科', '硕士', '博士', '学历不限']
  if self.stop():  #提供停止
   return
  edu_pie = Pie(init_opts = opts.InitOpts(width = '800px', height = '600px'))
  edu_pie.add('', [list(z) for z in zip(edu_list, edu)], radius = ['0%', '70%'])
  edu_pie.set_global_opts(title_opts = opts.TitleOpts(title = html_name, pos_left = 'center', pos_top = '2'),
                    legend_opts = opts.LegendOpts(is_show = False),
                    toolbox_opts = opts.ToolboxOpts(is_show = True, feature = {'saveAsImage': {}}))
  edu_pie.set_series_opts(label_opts = opts.LabelOpts(is_show = True, formatter = '{b} {c}'))
  edu_pie.render(xl_path + '\\' + html_name + '.html')
  self.ui.log(html_name + '图生成成功!!!')
  self.ui.log('----------------------------------------------------')

 def eachJob_CSize_Pie(self, kw_list, html_name):
  global ws
  global absolute_kw
  self.ui.log(html_name + '图生成中...')
  dx_path = os.path.abspath('..') + '\\公司大小'
  if self.stop():  #提供停止
   return
  if not os.path.exists(dx_path):  #判断公司大小文件夹是否存在
   os.mkdir(dx_path)
  size = [0, 0, 0, 0, 0, 0, 0]  #各公司大小类型
  #absolute_kw = ['flash', 'Flash', 'FLASH', 'PS', 'Ps', 'ps', 'CG', 'cg', 'Cg', 'VR', 'Vr', 'vr', 'AR', 'Ar', 'ar']
  max_r = ws.max_row
  for i in range(2, max_r+1):
   if self.stop():  #提供停止
    return
   if ws['A'+str(i)].value != None:
    tmp = ws['A'+str(i)].value
    j = 0
    while j < len(kw_list):
     if kw_list[j] in absolute_kw and kw_list[j] in jieba.lcut(tmp, cut_all = False):  #判断是否含有关键词
      if kw_list[j] == 'VR' and '摄影' not in kw_list and tmp.find('摄影') != -1: #VR摄影师不属于VR
       j+= 1
      else:
       break
     elif tmp.find(kw_list[j]) != -1:  #判断是否为相应职位
      break
     else:
      j += 1
    if j < len(kw_list):  #是否为相应职位
     if ws['L'+str(i)].value != None and (ws['L'+str(i)].value.find('少于') != -1 or ws['L'+str(i)].value.find('0-20') != -1):
      size[0] += 1
     elif ws['L'+str(i)].value != None and (ws['L'+str(i)].value.find('50-150') != -1 or ws['L'+str(i)].value.find('20-99') != -1):
      size[1] += 1
     elif ws['L'+str(i)].value != None and (ws['L'+str(i)].value.find('150-500') != -1 or ws['L'+str(i)].value.find('100-499') != -1):
      size[2] += 1
     elif ws['L'+str(i)].value != None and (ws['L'+str(i)].value.find('500-1000') != -1 or ws['L'+str(i)].value.find('500-999') != -1):
      size[3] += 1
     elif ws['L'+str(i)].value != None and ws['L'+str(i)].value.find('1000-5000') != -1:
      size[4] += 1
     elif ws['L'+str(i)].value != None and ws['L'+str(i)].value.find('5000-10000') != -1:
      size[5] += 1
     elif ws['L'+str(i)].value != None and ws['L'+str(i)].value.find('以上') != -1:
      size[6] += 1
     elif ws['L'+str(i)].value != None and ws['L'+str(i)].value.find('1000-9999') != -1:
      size[4] += 1
      size[5] += 1

  size_list = ['少于50人', '50-150人', '150-500人', '500-1000人', '1000-5000人', '5000-10000人', '10000人以上']
  if self.stop():  #提供停止
   return
  size_pie = Pie(init_opts = opts.InitOpts(width = '800px', height = '600px'))
  size_pie.add('', [list(z) for z in zip(size_list, size)], radius = ['0%', '65%'])
  size_pie.set_global_opts(title_opts = opts.TitleOpts(title = html_name, pos_left = 'center', pos_top = '2'),
                    legend_opts = opts.LegendOpts(is_show = False),
                    toolbox_opts = opts.ToolboxOpts(is_show = True, feature = {'saveAsImage': {}}))
  size_pie.set_series_opts(label_opts = opts.LabelOpts(is_show = True, formatter = '{b} {c}'))
  size_pie.render(dx_path + '\\' + html_name + '.html')
  self.ui.log(html_name + '图生成成功!!!')
  self.ui.log('----------------------------------------------------')

 def eachJob_Req_WordCloud(self, KW_list_list, pic_name_list):  #词云图生成  单进程
  global ws
  global absolute_kw
  cy_path = os.path.abspath('..') + '\\职位描述和要求'
  if not os.path.exists(cy_path):  #判断职位描述和要求文件夹是否存在
   os.mkdir(cy_path)
  wordCloudStopWord = """职能 类别 良好 的 年假 以上 优先 工作 经验 任职 要求 熟练 使用 相关 周末 养老 希望 元 生育 五天 聚餐 调薪 两 每周 礼金 冰箱 午餐 猫 之间 跟进 地址 六薪 
专业 有 较 及 以上 等 关键字 一定 负责 公司 优先 考虑 独立 任职 资格 具有 具备 作者 物料 医疗 一种 职责 生活 天 月 小时 欢迎 宝贝 礼品 医院 饮料 男女 一位 公交 灾 上班 四大 双薪 
职能 好 要求 五险 一金 类别 带薪 年度 年 法定 节假日 善于 毕业 过程 中 常用 强烈 自己 团建 年终 会 岗位 带领 地铁 站 每天 期待 礼物 广场 结婚 零食 六险 尤佳 一家 失业 春节 底薪 
至少 根据 员工 解决 福利 办公 环境 不 限于 限 于 交办 常用 描述 全日制 地点 婚假 投递 简历 奖金 下午茶 薪资 羽毛球 一款 午休 全勤奖 同事 栋 丧假 男女 率低 号线 印刷 病假 邮箱 
产假 承受 参与 旅游 并 进行 晋升 遇到 加入 我们 你 我 入职 优先 住房 公积金 假期 薪酬 大小 强具 生日 1 一起 5 附上 从业 TV 工伤 定期 下车 孪生 以下 附带 分项 园区 面试 话费 
深厚 深刻 其他 其它 部门 规定 购买 社保 保险 聘 请 急 ü ● ★ ◆ ■ 丰富 富 周末 双休 者 能够 人员 具 分红 发放 广阔 一门 一次 一个 。 ° 一次 年终奖 周一 座 大厦 免费 或"""   #停用词表
  if self.stop():  #提供停止
   return False
  wc_sw = wordCloudStopWord.replace('\n', '').split(' ')
  #pic_name_list = ['Android安卓', '动画', 'C#', 'C++', '游戏', '平面', 'Java', '媒体', '建模', '摄影摄像', '影视后期', '数据库', '数媒技术', 'Unity', '虚拟现实', 'Web网页前端']
  schedule = -1
  max_r = ws.max_row
  fc_list =['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']  #各职位描述和要求分词列表
  self.ui.log('获取职位描述和要求分词中...')
  for x in range(16):
   if schedule != int((x+1)/16*100-1):
    schedule = int((x+1)/16*100-1)
    self.ui.log(str(schedule)+'%')
   for i in range(2, max_r+1):
    if self.stop():  #提供停止
     return False
    if ws['A'+str(i)].value != None:
     tmp_A = ws['A'+str(i)].value
     j = 0
     while j < len(KW_list_list[x]):
      if KW_list_list[x][j] in absolute_kw and KW_list_list[x][j] in jieba.lcut(tmp_A, cut_all = False):  #判断是否含有关键词
       if KW_list_list[x][j] == 'VR' and '摄影' not in KW_list_list[x] and tmp_A.find('摄影') != -1: #VR摄影师不属于VR
        j+= 1
       else:
        break
      elif tmp_A.find(KW_list_list[x][j]) != -1:  #判断是否为相应职位
       break
      else:
       j += 1
     if j < len(KW_list_list[x]):  #是否为相应职位
      if ws['O'+str(i)].value != None:   #职位描述和要求不为空
       tmp_O = ws['O'+str(i)].value
       for z in wc_sw:
        tmp_O = tmp_O.replace(z, '')
       tmp_O = tmp_O.replace('C#', '/Csharp/').replace('c#', '/Csharp/').replace('C++', '/Cpp/').replace('c++', '/Cpp/').replace('Java', '/Java/')
       tmp_O = tmp_O.replace('Android', '/Android/').replace('Unity', '/Unity/').replace('Unity3D', '/Unity3D/').replace('U3D', '/U3D/')
       fc_list[x] += ' '.join(jieba.cut(tmp_O, cut_all = False))
   
  if self.stop():  #提供停止
   return False
  try:
   sw_file = open(os.getcwd() + '\\my_stopwords.txt','r', encoding = 'utf-8')
  except FileNotFoundError:
   self.ui.log('当前文件目录'+os.getcwd()+'下找不到停用词文件my_stopwords.txt')
   self.ui.log('词云图生成失败!')
   return False
  if self.stop():  #提供停止
   return False
  cn_stopwords = set()
  content = [line.strip() for line in sw_file.readlines()]
  cn_stopwords.update(content)
  sw_file.close()
  if os.path.exists(os.getcwd() + '\\msyh.ttc') == False:
   self.ui.log('当前文件目录'+os.getcwd()+'下找不到微软雅黑字体文件msyh.ttc')
   self.ui.log('词云图生成失败!')
   return False
  self.ui.log('正在生成词云图...')
  if self.stop():  #提供停止
   return False
  wc = WordCloud(font_path = 'msyh.ttc', max_words = 1600, max_font_size = 120, min_font_size = 8, width = 1280, height = 800,
                prefer_horizontal = 1, random_state = 50, margin = 6, stopwords = cn_stopwords, background_color='white', collocations = False)
  for x in range(16):
   if self.stop():  #提供停止
    return False
   self.ui.log(pic_name_list[x]+'相关职位描述和要求词云图生成中...')
   try:
    wc.generate_from_text(fc_list[x].strip())
    wc.to_file(cy_path + '\\' + pic_name_list[x]+'相关职位描述和要求词云.png')
   except Exception as err:
    self.ui.log('词云生成失败: '+str(err))
    return False
   self.ui.log(pic_name_list[x]+'相关职位描述和要求词云图生成成功!')
   self.ui.log('----------------------------------------------------')
  self.ui.log('数媒行业职位描述和要求词云图生成中...')
  if self.stop():  #提供停止
   return False
  try:
   wc.generate_from_text(' '.join(fc_list))  #数媒行业的总分词
   wc.to_file(cy_path + '\\' +'数媒行业职位描述和要求词云.png')
  except Exception as err:
   self.ui.log('词云生成失败: '+str(err))
   return False
  self.ui.log('数媒行业职位描述和要求词云图生成成功!')
  self.ui.log('----------------------------------------------------')
  del(wc, cn_stopwords, content, sw_file, fc_list)
  return True

 def do(self):
  global ws
  global android_kw
  global animate_kw
  global cs_kw
  global cpp_kw
  global game_kw
  global graphic_kw
  global java_kw
  global media_kw
  global model_kw
  global photography_kw
  global post_production_kw
  global sql_kw
  global technology_kw
  global unity_kw
  global virtual_kw
  global web_kw
  global smkw
  global absolute_kw
  tmp_kw_list_list = []
  tmp_kw_list_list.append(android_kw)
  tmp_kw_list_list.append(animate_kw)
  tmp_kw_list_list.append(cs_kw)
  tmp_kw_list_list.append(cpp_kw)
  tmp_kw_list_list.append(game_kw)
  tmp_kw_list_list.append(graphic_kw)
  tmp_kw_list_list.append(java_kw)
  tmp_kw_list_list.append(media_kw)
  tmp_kw_list_list.append(model_kw)
  tmp_kw_list_list.append(photography_kw)
  tmp_kw_list_list.append(post_production_kw)
  tmp_kw_list_list.append(sql_kw)
  tmp_kw_list_list.append(technology_kw)
  tmp_kw_list_list.append(unity_kw)
  tmp_kw_list_list.append(virtual_kw)
  tmp_kw_list_list.append(web_kw)
  file_name_list = ['Android安卓', '动画', 'C#', 'C++', '游戏', '平面', 'Java', '媒体', '建模', '摄影摄像', '影视后期', '数据库', '数媒技术', 'Unity', '虚拟现实', 'Web网页前端']
  if self.stop():  #提供停止
   self.finish()
   return

  start_time = time.time()
  single_process = False
  num_of_job = self.ui.gzEachJobvar.get()+self.ui.szEachJobvar.get()+self.ui.eachCityMoneyvar.get()+self.ui.eachJobMoneyvar.get()+self.ui.eachJobExpvar.get()+self.ui.eachJobEduvar.get()+self.ui.eachJobCSizevar.get()
  p_n = 0
  pool = None
  queue = None
  free_memory = psutil.virtual_memory().free/1024/1024    #获取系统空闲内存(MB)
  max_r = ws.max_row
  row_per_MB = max_r/5960+45
  mem = max_r/row_per_MB              #计算所有招聘数据行所占内存(MB)
  p_n = multiprocessing.cpu_count()   #获取cpu逻辑核心个数
  while p_n > 0:
   if p_n*mem*1.12+72.5 < free_memory:          #分析最多可以使用cpu多少核心
    break
   p_n -= 1
  if p_n > 1 and num_of_job > 0:
   pool = multiprocessing.Pool(min(p_n-1, num_of_job))  #多进程
   queue = multiprocessing.Manager().Queue()
  elif p_n == 1 and num_of_job > 0:
   single_process = True  #单进程
  elif p_n == 0:
   self.ui.log('内存不足，无法执行可视化操作!')
   tk.messagebox.showwarning('警告', '内存不足!')
   self.finish()
   return

  generate_num = 0
  if pool != None:
   self.ui.log('多进程生成中...')
  if self.ui.gzEachJobvar.get() == 1 and pool != None:
   generate_num += 17
   pool.apply_async(gz_fun, (ws, smkw, tmp_kw_list_list, absolute_kw, queue)) #广州
  if self.ui.szEachJobvar.get() == 1 and pool != None:
   generate_num += 17
   pool.apply_async(sz_fun, (ws, smkw, tmp_kw_list_list, absolute_kw, queue, file_name_list))  #深圳
  if self.ui.eachCityMoneyvar.get() == 1 and pool != None:
   generate_num += 1
   pool.apply_async(gd_Money_Bar_p, (ws, smkw, queue))  #广东城市薪酬
  if self.ui.eachJobMoneyvar.get() == 1 and pool != None:
   generate_num += 1
   pool.apply_async(eachJob_Money_Bar_p, (ws, tmp_kw_list_list, absolute_kw, queue))  #数媒各薪酬
  if self.ui.eachJobExpvar.get() == 1 and pool != None:
   generate_num += 17
   pool.apply_async(eachJob_Exp_fun, (ws, smkw, tmp_kw_list_list, absolute_kw, queue, file_name_list))  #经验要求
  if self.ui.eachJobEduvar.get() == 1 and pool != None:
   generate_num += 17
   pool.apply_async(eachJob_Edu_fun, (ws, smkw, tmp_kw_list_list, absolute_kw, queue, file_name_list))  #学历要求
  if self.ui.eachJobCSizevar.get() == 1 and pool != None:
   generate_num += 17
   pool.apply_async(eachJob_CSize_fun, (ws, smkw, tmp_kw_list_list, absolute_kw, queue, file_name_list))  #公司大小

  if self.ui.gdEachJobvar.get() == 1:                 #广东省数媒相关行业各职位数量分布  主进程
   if self.stop():  #提供停止
    self.finish()
    if pool != None:
     pool.close()
     pool.terminate()
     pool.join()
    return
   self.gd_Map(smkw, 3000, None, '广东省数媒相关行业职位数量分布')
   while queue != None and queue.qsize() != 0:
    self.ui.log(queue.get())
    self.ui.log('----------------------------------------------------')
   if self.stop():  #提供停止
    self.finish()
    if pool != None:
     pool.close()
     pool.terminate()
     pool.join()
    return
   self.gd_Map(android_kw, 80, 0, '广东省Android安卓相关职位数量分布')
   while queue != None and queue.qsize() != 0:
    self.ui.log(queue.get())
    self.ui.log('----------------------------------------------------')
   if self.stop():  #提供停止
    self.finish()
    if pool != None:
     pool.close()
     pool.terminate()
     pool.join()
    return
   self.gd_Map(animate_kw, 40, 0, '广东省动画相关职位数量分布')
   while queue != None and queue.qsize() != 0:
    self.ui.log(queue.get())
    self.ui.log('----------------------------------------------------')
   if self.stop():  #提供停止
    self.finish()
    if pool != None:
     pool.close()
     pool.terminate()
     pool.join()
    return
   self.gd_Map(cs_kw, 150, 0, '广东省C#相关职位数量分布')
   while queue != None and queue.qsize() != 0:
    self.ui.log(queue.get())
    self.ui.log('----------------------------------------------------')
   if self.stop():  #提供停止
    self.finish()
    if pool != None:
     pool.close()
     pool.terminate()
     pool.join()
    return
   self.gd_Map(cpp_kw, 70, 0, '广东省C++相关职位数量分布')
   while queue != None and queue.qsize() != 0:
    self.ui.log(queue.get())
    self.ui.log('----------------------------------------------------')
   if self.stop():  #提供停止
    self.finish()
    if pool != None:
     pool.close()
     pool.terminate()
     pool.join()
    return
   self.gd_Map(game_kw, 80, 0, '广东省游戏相关职位数量分布')
   while queue != None and queue.qsize() != 0:
    self.ui.log(queue.get())
    self.ui.log('----------------------------------------------------')
   if self.stop():  #提供停止
    self.finish()
    if pool != None:
     pool.close()
     pool.terminate()
     pool.join()
    return
   self.gd_Map(graphic_kw, 900, 0, '广东省平面相关职位数量分布')
   while queue != None and queue.qsize() != 0:
    self.ui.log(queue.get())
    self.ui.log('----------------------------------------------------')
   if self.stop():  #提供停止
    self.finish()
    if pool != None:
     pool.close()
     pool.terminate()
     pool.join()
    return
   self.gd_Map(java_kw, 400, 0, '广东省Java相关职位数量分布')
   while queue != None and queue.qsize() != 0:
    self.ui.log(queue.get())
    self.ui.log('----------------------------------------------------')
   if self.stop():  #提供停止
    self.finish()
    if pool != None:
     pool.close()
     pool.terminate()
     pool.join()
    return
   self.gd_Map(media_kw, 650, 0, '广东省媒体相关职位数量分布')
   while queue != None and queue.qsize() != 0:
    self.ui.log(queue.get())
    self.ui.log('----------------------------------------------------')
   if self.stop():  #提供停止
    self.finish()
    if pool != None:
     pool.close()
     pool.terminate()
     pool.join()
    return
   self.gd_Map(model_kw, 50, 0, '广东省建模相关职位数量分布')
   while queue != None and queue.qsize() != 0:
    self.ui.log(queue.get())
    self.ui.log('----------------------------------------------------')
   if self.stop():  #提供停止
    self.finish()
    if pool != None:
     pool.close()
     pool.terminate()
     pool.join()
    return
   self.gd_Map(photography_kw, 200, 0, '广东省摄影摄像相关职位数量分布')
   while queue != None and queue.qsize() != 0:
    self.ui.log(queue.get())
    self.ui.log('----------------------------------------------------')
   if self.stop():  #提供停止
    self.finish()
    if pool != None:
     pool.close()
     pool.terminate()
     pool.join()
    return
   self.gd_Map(post_production_kw, 250, 0, '广东省影视后期相关职位数量分布')
   while queue != None and queue.qsize() != 0:
    self.ui.log(queue.get())
    self.ui.log('----------------------------------------------------')
   if self.stop():  #提供停止
    self.finish()
    if pool != None:
     pool.close()
     pool.terminate()
     pool.join()
    return
   self.gd_Map(sql_kw, 25, 0, '广东省数据库相关职位数量分布')
   while queue != None and queue.qsize() != 0:
    self.ui.log(queue.get())
    self.ui.log('----------------------------------------------------')
   if self.stop():  #提供停止
    self.finish()
    if pool != None:
     pool.close()
     pool.terminate()
     pool.join()
    return
   self.gd_Map(technology_kw, 150, 0, '广东省数媒技术相关职位数量分布')
   while queue != None and queue.qsize() != 0:
    self.ui.log(queue.get())
    self.ui.log('----------------------------------------------------')
   if self.stop():  #提供停止
    self.finish()
    if pool != None:
     pool.close()
     pool.terminate()
     pool.join()
    return
   self.gd_Map(unity_kw, 16, 0, '广东省Unity相关职位数量分布')
   while queue != None and queue.qsize() != 0:
    self.ui.log(queue.get())
    self.ui.log('----------------------------------------------------')
   if self.stop():  #提供停止
    self.finish()
    if pool != None:
     pool.close()
     pool.terminate()
     pool.join()
    return
   self.gd_Map(virtual_kw, 7, 0, '广东省虚拟现实相关职位数量分布')
   while queue != None and queue.qsize() != 0:
    self.ui.log(queue.get())
    self.ui.log('----------------------------------------------------')
   if self.stop():  #提供停止
    self.finish()
    if pool != None:
     pool.close()
     pool.terminate()
     pool.join()
    return
   self.gd_Map(web_kw, 300, 0, '广东省Web网页前端相关职位数量分布')
   while queue != None and queue.qsize() != 0:
    self.ui.log(queue.get())
    self.ui.log('----------------------------------------------------')
   if self.stop():  #提供停止
    self.finish()
    if pool != None:
     pool.close()
     pool.terminate()
     pool.join()
    return
  elif generate_num > 0:
   generate_index = 0
   while generate_index < generate_num:
    while queue.qsize() != 0:
     self.ui.log(queue.get())
     self.ui.log('----------------------------------------------------')
     generate_index += 1
    if self.stop():  #提供停止
     self.finish()
     if pool != None:
      pool.close()
      pool.terminate()
      pool.join()
     return
    time.sleep(0.5)
#-----------------------------------------------------------------------------------------------
  if self.ui.gzEachJobvar.get() == 1 and single_process == True:                    #广州市数媒相关行业各职位数量分布  单进程
   if self.stop():  #提供停止
    self.finish()
    return
   self.gz_Map(smkw, 4000, None, '广州市数媒相关行业职位数量分布')
   if self.stop():  #提供停止
    self.finish()
    return
   self.gz_Map(android_kw, 120, 0, '广州市Android安卓相关职位数量分布')
   if self.stop():  #提供停止
    self.finish()
    return
   self.gz_Map(animate_kw, 220, 0, '广州市动画相关职位数量分布')
   if self.stop():  #提供停止
    self.finish()
    return
   self.gz_Map(cs_kw, 50, 0, '广州市C#相关职位数量分布')
   if self.stop():  #提供停止
    self.finish()
    return
   self.gz_Map(cpp_kw, 120, 0, '广州市C++相关职位数量分布')
   if self.stop():  #提供停止
    self.finish()
    return
   self.gz_Map(game_kw, 320, 0, '广州市游戏相关职位数量分布')
   if self.stop():  #提供停止
    self.finish()
    return
   self.gz_Map(graphic_kw, 1200, 0, '广州市平面相关职位数量分布')
   if self.stop():  #提供停止
    self.finish()
    return
   self.gz_Map(java_kw, 600, 0, '广州市Java相关职位数量分布')
   if self.stop():  #提供停止
    self.finish()
    return
   self.gz_Map(media_kw, 1200, 0, '广州市媒体相关职位数量分布')
   if self.stop():  #提供停止
    self.finish()
    return
   self.gz_Map(model_kw, 75, 0, '广州市建模相关职位数量分布')
   if self.stop():  #提供停止
    self.finish()
    return
   self.gz_Map(photography_kw, None, 0, '广州市摄影摄像相关职位数量分布')
   if self.stop():  #提供停止
    self.finish()
    return
   self.gz_Map(post_production_kw, 600, 0, '广州市影视后期相关职位数量分布')
   if self.stop():  #提供停止
    self.finish()
    return
   self.gz_Map(sql_kw, 30, 0, '广州市数据库相关职位数量分布')
   if self.stop():  #提供停止
    self.finish()
    return
   self.gz_Map(technology_kw, 160, 0, '广州市数媒技术相关职位数量分布')
   if self.stop():  #提供停止
    self.finish()
    return
   self.gz_Map(unity_kw, 55, 0, '广州市Unity相关职位数量分布')
   if self.stop():  #提供停止
    self.finish()
    return
   self.gz_Map(virtual_kw, None, 0, '广州市虚拟现实相关职位数量分布')
   if self.stop():  #提供停止
    self.finish()
    return
   self.gz_Map(web_kw, 500, 0, '广州市Web网页前端相关职位数量分布')
   if self.stop():  #提供停止
    self.finish()
    return
#-----------------------------------------------------------------------------------------------
  if self.ui.szEachJobvar.get() == 1 and single_process == True:                    #深圳市数媒相关行业各职位数量分布  单进程
   if self.stop():  #提供停止
    self.finish()
    return
   self.sz_Pie(smkw, '深圳市数媒相关行业职位数量分布')
   for x in range(16):
    if self.stop():  #提供停止
     self.finish()
     return
    self.sz_Pie(tmp_kw_list_list[x], '深圳市'+file_name_list[x]+'相关职位数量分布')
   if self.stop():  #提供停止
    self.finish()
    return
#-----------------------------------------------------------------------------------------------
  if self.ui.eachCityMoneyvar.get() == 1 and single_process == True:               #广东城市数媒行业职位薪水对比  单进程
   if self.stop():  #提供停止
    self.finish()
    return
   self.gd_Money_Bar(smkw)
   if self.stop():  #提供停止
    self.finish()
    return
#-----------------------------------------------------------------------------------------------
  if self.ui.eachJobMoneyvar.get() == 1 and single_process == True:                #数媒行业各职位薪水对比  单进程
   if self.stop():  #提供停止
    self.finish()
    return
   self.eachJob_Money_Bar(tmp_kw_list_list)
   if self.stop():  #提供停止
    self.finish()
    return
#-----------------------------------------------------------------------------------------------
  if self.ui.eachJobExpvar.get() == 1 and single_process == True:               #数媒行业各职位经验要求分布  单进程
   if self.stop():  #提供停止
    self.finish()
    return
   self.eachJob_Exp_Pie(smkw, '数媒行业职位经验要求分布')
   for x in range(16):
    if self.stop():  #提供停止
     self.finish()
     return
    self.eachJob_Exp_Pie(tmp_kw_list_list[x], file_name_list[x]+'相关职位经验要求分布')
   if self.stop():  #提供停止
    self.finish()
    return
#-----------------------------------------------------------------------------------------------
  if self.ui.eachJobEduvar.get() == 1 and single_process == True:               #数媒行业各职位学历要求分布  单进程
   if self.stop():  #提供停止
    self.finish()
    return
   self.eachJob_Edu_Pie(smkw, '数媒行业职位学历要求分布')
   for x in range(16):
    if self.stop():  #提供停止
     self.finish()
     return
    self.eachJob_Edu_Pie(tmp_kw_list_list[x], file_name_list[x]+'相关职位学历要求分布')
   if self.stop():  #提供停止
    self.finish()
    return
#------------------------------------------------------------------------------------------------
  if self.ui.eachJobCSizevar.get() == 1 and single_process == True:               #数媒行业各职位公司大小分布  单进程
   if self.stop():  #提供停止
    self.finish()
    return
   self.eachJob_CSize_Pie(smkw, '数媒行业职位公司大小分布')
   for x in range(16):
    if self.stop():  #提供停止
     self.finish()
     return
    self.eachJob_CSize_Pie(tmp_kw_list_list[x], file_name_list[x]+'相关职位公司大小分布')
   if self.stop():  #提供停止
    self.finish()
    return
#-----------------------------------------------------------------------------------------------
  if pool != None:
   pool.close()
   pool.join()
   pool = None
   queue = None
  fc_dict = None
  if not os.path.exists(os.getcwd() + '\\my_stopwords.txt') or not os.path.exists(os.getcwd() + '\\msyh.ttc'):
   if not os.path.exists(os.getcwd() + '\\my_stopwords.txt'):
    self.ui.log('当前文件目录'+os.getcwd()+'下找不到中文词云停用词文件my_stopwords.txt')
   if not os.path.exists(os.getcwd() + '\\msyh.ttc'):
    self.ui.log('当前文件目录'+os.getcwd()+'下找不到微软雅黑字体文件msyh.ttc')
   self.ui.log('词云图生成失败!')
   self.finish()  #可视化结束
   return
  if p_n > 1 and self.ui.eachJobReqvar.get() == 1:
   pool = multiprocessing.Pool(min(p_n, 16))
   fc_dict = multiprocessing.Manager().dict()
  if self.ui.eachJobReqvar.get() == 1:                                        #数媒行业各职位描述和要求词云图
   if pool == None:          #单进程
    finish_wc = self.eachJob_Req_WordCloud(tmp_kw_list_list, file_name_list)
    if not finish_wc:  #中途退出返回False
     self.finish()
     return
   else:                     #多进程
    cy_path = os.path.abspath('..') + '\\职位描述和要求'  #词云图存放路径
    if not os.path.exists(cy_path):  #判断职位描述和要求文件夹是否存在
     os.mkdir(cy_path)
    self.ui.log('获取职位描述和要求分词中...')
    self.ui.log('生成多进程中...')
    for x in range(16):
     pool.apply_async(WordCloud_p, (ws, tmp_kw_list_list[x], absolute_kw, fc_dict, x))
    schedule = 0
    while len(fc_dict) != 16:
     if schedule != int(len(fc_dict)/16*100):
      schedule = int(len(fc_dict)/16*100)
      self.ui.log(str(schedule)+'%')
     time.sleep(1)
     if self.stop():  #提供停止
      self.finish()
      pool.close()
      pool.terminate()
      pool.join()
      return
    self.ui.log('100%')
    pool.close()
    pool.join()
    fc_list =['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']
    for i in range(16):
     fc_list[i] = fc_dict.get(i)
    try:
     sw_file = open(os.getcwd() + '\\my_stopwords.txt','r', encoding = 'utf-8')
    except FileNotFoundError:
     self.ui.log('当前文件目录'+os.getcwd()+'下找不到停用词文件my_stopwords.txt')
     self.ui.log('词云图生成失败!')
     self.finish()  #可视化结束
     return
    cn_stopwords = set()
    content = [line.strip() for line in sw_file.readlines()]
    cn_stopwords.update(content)
    sw_file.close()
    self.ui.log('正在生成词云图...')
    if self.stop():  #提供停止
     self.finish()
     return
    wc = WordCloud(font_path = 'msyh.ttc', max_words = 1600, max_font_size = 120, min_font_size = 8, width = 1280, height = 800,
                prefer_horizontal = 1, random_state = 50, margin = 6, stopwords = cn_stopwords, background_color='white', collocations = False)
    for x in range(16):
     if self.stop():  #提供停止
      self.finish()
      return
     self.ui.log(file_name_list[x]+'相关职位描述和要求词云图生成中...')
     try:
      wc.generate_from_text(fc_list[x].strip())
      wc.to_file(cy_path + '\\' + file_name_list[x]+'相关职位描述和要求词云.png')
     except Exception as err:
      self.ui.log('词云生成失败: '+str(err))
      self.finish()
      return
     self.ui.log(file_name_list[x]+'相关职位描述和要求词云图生成成功!')
     self.ui.log('----------------------------------------------------')
    self.ui.log('数媒行业职位描述和要求词云图生成中...')
    if self.stop():  #提供停止
     self.finish()
     return
    try:
     wc.generate_from_text(' '.join(fc_list))  #数媒行业的总分词
     wc.to_file(cy_path + '\\' +'数媒行业职位描述和要求词云.png')
    except Exception as err:
     self.ui.log('词云生成失败: '+str(err))
     self.finish()
     return
    self.ui.log('数媒行业职位描述和要求词云图生成成功!')
    self.ui.log('----------------------------------------------------')
    del(wc, cn_stopwords, content, sw_file)
  if pool != None:
   del(pool, fc_dict, fc_list)
#-----------------------------------------------------------------------------------------------
  self.ui.log('可视化完成!')
  end_time = time.time()
  self.ui.log('用时' + str(int((end_time-start_time)/3600)) + '时' + str(int((end_time-start_time)%3600/60)) + '分' + str(int((end_time-start_time)%60)) + '秒')
  self.finish()  #可视化结束

 def stop(self):
  global isStopVisualization
  global isVisualization
  global force_exit
  if force_exit:
   self.ui.log('可视化停止!')
   return True
  if isStopVisualization and isVisualization:  #可视化运行中
   if tk.messagebox.askyesno(title = '提示', message = '是否要停止可视化？'):   #确定要停止
    self.ui.log('可视化停止!')
    return True
   else:
    isStopVisualization = False
    self.ui.log('取消停止，继续可视化')
    self.ui.startStopVirtualBtn.config(state=NORMAL)
    return False
  else:
   return False
   

 def finish(self):
  global isStopVisualization
  global isVisualization
  global force_exit
  isVisualization = False
  isStopVisualization = False
  if not force_exit:
   self.ui.startStopVirtualBtn.config(state=NORMAL)
   self.ui.startStopVirtualBtn.config(text = '开始')
   self.ui.gdEachJobBtn.config(state=NORMAL)
   self.ui.gzEachJobBtn.config(state=NORMAL)
   self.ui.szEachJobBtn.config(state=NORMAL)
   self.ui.eachCityMoneyBtn.config(state=NORMAL)
   self.ui.eachJobMoneyBtn.config(state=NORMAL)
   self.ui.eachJobExpBtn.config(state=NORMAL)
   self.ui.eachJobEduBtn.config(state=NORMAL)
   self.ui.eachJobReqBtn.config(state=NORMAL)
   self.ui.openGoogleButton.config(state=NORMAL)
   self.ui.openBossButton.config(state=NORMAL)
   self.ui.openJobButton.config(state=NORMAL)
   self.ui.spinbox.config(state=NORMAL)
   self.ui.startButton.config(state=NORMAL)
   self.ui.stopButton.config(state=NORMAL)
   self.ui.checkButton.config(state=NORMAL)
   self.ui.notGDButton.config(state=NORMAL)
   self.ui.eachJobCSizeBtn.config(state=NORMAL)
   self.ui.seleteAllBtn.config(state=NORMAL)
   self.ui.invertAllBtn.config(state=NORMAL)

  
#界面自定义类-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
class UiRoot(Tk):
 def __init__(self):
  Tk.__init__(self)

class UiFrame(Frame):
 def run(self, func):
  self.master.run(func)
  
class BackgroundTask(object):
 """Similar to Android's AsyncTask"""
 def __init__(self, ui):
  self.ui = ui       #tk的root

 def doBefore(self):
  """Runs on the main thread, returns arg"""
  pass
 def do(self, arg):
  """Runs on a separate thread, returns result"""
  pass
 def doAfter(self, result):
  """Runs on the main thread again"""
  pass

 def run(self):
  """Invoke this on the main thread only"""
  arg = self.doBefore()
  threading.Thread(target=self._onThread, args=[arg]).start()     #新建一个线程处理去按钮触发的事件

 def _onThread(self, arg):
  result = self.do(arg)
  self.doAfter(result)
  #self.ui.run(lambda: self.doAfter(result))

class ShowTask(BackgroundTask):  #继承BackgroundTask类       打开谷歌浏览器触发事件
 def doBefore(self):
  self.ui.openGoogleButton.config(state=DISABLED)
  self.ui.notGDButton.config(state=DISABLED)
  self.ui.checkButton.config(state=DISABLED)
  self.ui.startStopVirtualBtn.config(state=DISABLED)

 def do(self, arg):
  self.ui.log('打开谷歌浏览器')
  global driver
  global chromeOpen
  global startFlag
  global stopFlag
  try:
   driver = webdriver.Chrome()
  except se.common.exceptions.WebDriverException as err:
   self.ui.log('找不到chromedriver.exe文件，无法打开谷歌浏览器!')
   self.ui.log(str(err))
   return
  except Exception as err:
   self.ui.log('打开谷歌浏览器出错: '+str(err))
   return
  else:
   driver.set_page_load_timeout(10)      #设置超时限制时间
  while 1:
   try:
    driver.current_url   #通过获取当前URL，检查浏览器状态
    time.sleep(1)
   except se.common.exceptions.WebDriverException or se.common.exceptions.InvalidSessionIdException:   #出现异常说明浏览器已退出
    driver.quit()
    self.ui.log('谷歌浏览器已关闭')
    tk.messagebox.showinfo("提示", "Google浏览器已关闭")
    if startFlag:         #在运行或暂停中则停止爬虫
     stopFlag = True
     self.ui.log('与浏览器失去联系，爬虫被迫停止!')
     tk.messagebox.showerror("错误", "与浏览器失去联系，爬虫被迫停止!")
     tk.messagebox.showinfo("提示", "请重启爬虫程序")   #建议退出程序
     os._exit(0)         #退出程序
    break
   except Exception as err:
    self.ui.log('无法连接到浏览器，请关闭浏览器重试！')
    self.ui.log(str(err))
    tk.messagebox.showwarning("警告", "无法连接到浏览器，请关闭浏览器重试！")
    break
   else:
    if not chromeOpen:
     self.ui.log("浏览器打开成功")
     chromeOpen = True

 def doAfter(self, result):
  global chromeOpen
  chromeOpen = False
  self.ui.openGoogleButton.config(state=NORMAL)
  self.ui.notGDButton.config(state=NORMAL)
  self.ui.checkButton.config(state=NORMAL)
  self.ui.startStopVirtualBtn.config(state=NORMAL)


class OpenBoss(BackgroundTask):            #打开Boss直聘按钮触发事件
 def doBefore(self):
  self.ui.openBossButton.config(state=DISABLED)

 def do(self, arg):
  global chromeOpen
  global driver
  if chromeOpen == False:                #判断浏览器是否打开
   self.ui.log("Google浏览器未打开")
   tk.messagebox.showwarning("警告", "Google浏览器未打开！")
   return
  self.ui.log("打开Boss直聘网页")
  while 1:
   try:
    driver.get('https://www.zhipin.com/')   #打开Boss直聘
   except se.common.exceptions.WebDriverException:
    self.ui.log('连接超时，请检查网络')
    self.ui.log('5秒后重连')
    time.sleep(5)
   else:
    break
  self.ui.log('请在网页设置好要爬取的职业和地区!')

 def doAfter(self, result):
  self.ui.openBossButton.config(state=NORMAL)


class OpenJob(BackgroundTask):         #打开前程无忧按钮触发事件
 def doBefore(self):
  self.ui.openJobButton.config(state=DISABLED)

 def do(self, arg):
  global chromeOpen
  global driver
  if chromeOpen == False:                 #判断浏览器是否打开
   self.ui.log("Google浏览器未打开")
   tk.messagebox.showwarning("警告", "Google浏览器未打开！")
   return
  self.ui.log("打开前程无忧网页")
  while 1:
   try:
    driver.get('https://www.51job.com/')   #打开前程无忧
   except se.common.exceptions.WebDriverException:
    self.ui.log('连接超时，请检查网络')
    self.ui.log('5秒后重连')
    time.sleep(5)
   else:
    break
  self.ui.log('请在网页设置好要爬取的职业和地区!')

 def doAfter(self, result):
  self.ui.openJobButton.config(state=NORMAL)


class Start_Pause(BackgroundTask):      #开始/暂停爬虫按钮触发事件
 def doBefore(self):
  pass

 def do(self, arg):
  global runFlag
  global startFlag
  global driver
  global wb
  global ws

  if chromeOpen == False:                           #判断浏览器是否打开
   self.ui.log("Google浏览器未打开")
   tk.messagebox.showwarning("警告", "Google浏览器未打开！")
   return

  url = driver.current_url
  if url.find('zhipin.com') == -1 and url.find('51job.com') == -1:     #判断网页是否打开正确
   tk.messagebox.showerror('错误', '无法识别该网页！')
   self.ui.log("无法识别网页！ 请打开Boss直聘或前程无忧网页")
   return

  self.ui.openBossButton.config(state=DISABLED)      #爬取期间禁止再次打开或更换网页
  self.ui.openJobButton.config(state=DISABLED)
  if runFlag == False:     #检查是否在运行
   self.ui.log("开始爬取")
   global sleepTime
   sleepTime = int(self.ui.spinbox.get())  #设置休眠时间
   self.ui.startButton.config(text = '暂停爬取')
   self.ui.spinbox.config(state=DISABLED)
   self.ui.notGDButton.config(state=DISABLED)
   self.ui.checkButton.config(state=DISABLED)
   self.ui.startStopVirtualBtn.config(state=DISABLED)
   runFlag = True                 #运行中标志
   if startFlag == False:     #第一次开始
    startFlag = True               #已经按过开始键标志
    if wb == None and ws == None:
     initWB(self.ui).open_data_file(True)  #初始化wb和ws，打开data文件
    if url.find('51job.com') != -1:
     Crawling(self.ui).get_job()          #执行爬取前程无忧网页
    else:
     Crawling(self.ui).get_boss()          #执行爬取Boss直聘网页
  else:
   self.ui.log("暂停爬取中...")
   self.ui.startButton.config(text = '开始爬取')
   self.ui.spinbox.config(state=NORMAL)
   runFlag = False            #暂停中标志

 def doAfter(self, result):
  pass

class Stop(BackgroundTask):          #停止按钮触发事件
 def doBefore(self):
  self.ui.stopButton.config(state=DISABLED)
  self.ui.startButton.config(state=DISABLED)

 def do(self, arg):
  global runFlag
  global stopFlag
  global startFlag
  if not startFlag:                 #爬虫未开始却按下停止键
   self.ui.log("爬虫未处于运行或暂停中")
   tk.messagebox.showinfo("提示", "爬虫未处于运行或暂停中！")
   self.ui.stopButton.config(state=NORMAL)
   self.ui.startButton.config(state=NORMAL)
   return
  else:          #没停止
   if tk.messagebox.askyesno(title = '提示', message = '是否要停止爬取？'):   #确定要停止
    self.ui.log("停止爬取中...")
    self.ui.startButton.config(text = '开始爬取')
    self.ui.spinbox.config(state=NORMAL)
    stopFlag = True         #停止运行中的爬虫标志
   else:
    self.ui.stopButton.config(state=NORMAL)
    self.ui.startButton.config(state=NORMAL)
   

 def doAfter(self, result):
  pass

class Check(BackgroundTask):          #查重按钮触发事件
 def doBefore(self):
  pass

 def do(self, arg):
  global isChecking
  global isStopChecking
  global wb
  global ws
  if not isChecking and not isStopChecking:
   isChecking = True        #查重中标志
   if wb == None and ws == None:
    if not initWB(self.ui).open_data_file(False):    #初始化wb和ws，打开data文件
     self.ui.log(os.path.abspath('.') + ' 目录下找不到data.xlsx文件!')
     isChecking = False
     return
   self.ui.log("初始化中...")
   self.ui.canvas.itemconfig(self.ui.fill_rec, fill = '#00d000')   #初始化进度条颜色 绿色
   self.ui.update()   #更新显示
   self.ui.checkButton.config(text = '停止查重')
   self.ui.stopButton.config(state=DISABLED)        #UI按钮不可用
   self.ui.startButton.config(state=DISABLED)
   self.ui.openGoogleButton.config(state=DISABLED)
   self.ui.openBossButton.config(state=DISABLED)
   self.ui.openJobButton.config(state=DISABLED)
   self.ui.spinbox.config(state=DISABLED)
   self.ui.notGDButton.config(state=DISABLED)
   self.ui.startStopVirtualBtn.config(state=DISABLED)
   Checking(self.ui).do_check()        #执行查重
  else:
   self.ui.log("停止查重中...")
   isStopChecking = True    #结束标志

 def doAfter(self, result):
  pass


class Change_checkGD(BackgroundTask):          #勾选框发事件
 def doBefore(self):
  pass

 def do(self, arg):
  global checkGD
  checkGD = not checkGD   #是否去除非广东数据

 def doAfter(self, result):
  pass

class VisualizationButton(BackgroundTask):     #可视化开始事件
 def doBefore(self):
  pass

 def do(self, arg):
  global isVisualization
  global isStopVisualization
  global wb
  global ws
  if self.ui.gdEachJobvar.get() == 0 and self.ui.gzEachJobvar.get() == 0 and self.ui.szEachJobvar.get() == 0 and self.ui.eachCityMoneyvar.get() == 0 and self.ui.eachJobMoneyvar.get() == 0 and self.ui.eachJobExpvar.get() == 0 and self.ui.eachJobEduvar.get() == 0 and self.ui.eachJobReqvar.get() == 0 and self.ui.eachJobCSizevar.get() == 0:
   self.ui.log('你未选择任何一个可视化项目!')
   tk.messagebox.showinfo('提示', '你未选择任何一个可视化项目!')
   return
  self.ui.startStopVirtualBtn.config(state=DISABLED)
  self.ui.seleteAllBtn.config(state=DISABLED)
  self.ui.invertAllBtn.config(state=DISABLED)
  if not isVisualization and not isStopVisualization:
   isVisualization = True
   if wb == None and ws == None:
    if not initWB(self.ui).open_data_file(False):    #初始化wb和ws，打开data文件
     self.ui.log(os.path.abspath('.') + ' 目录下找不到data.xlsx文件!')
     isVisualization = False
     self.ui.startStopVirtualBtn.config(state=NORMAL)
     self.ui.seleteAllBtn.config(state=NORMAL)
     self.ui.invertAllBtn.config(state=NORMAL)
     return
   self.ui.log('初始化中...')
   self.ui.startStopVirtualBtn.config(state=NORMAL)
   self.ui.startStopVirtualBtn.config(text = '停止')
   self.ui.gdEachJobBtn.config(state=DISABLED)
   self.ui.gzEachJobBtn.config(state=DISABLED)
   self.ui.szEachJobBtn.config(state=DISABLED)
   self.ui.eachCityMoneyBtn.config(state=DISABLED)
   self.ui.eachJobMoneyBtn.config(state=DISABLED)
   self.ui.eachJobExpBtn.config(state=DISABLED)
   self.ui.eachJobEduBtn.config(state=DISABLED)
   self.ui.eachJobReqBtn.config(state=DISABLED)
   self.ui.openGoogleButton.config(state=DISABLED)
   self.ui.openBossButton.config(state=DISABLED)
   self.ui.openJobButton.config(state=DISABLED)
   self.ui.spinbox.config(state=DISABLED)
   self.ui.startButton.config(state=DISABLED)
   self.ui.stopButton.config(state=DISABLED)
   self.ui.checkButton.config(state=DISABLED)
   self.ui.notGDButton.config(state=DISABLED)
   self.ui.eachJobCSizeBtn.config(state=DISABLED)
#   self.ui.seleteAllBtn.config(state=DISABLED)
#   self.ui.invertAllBtn.config(state=DISABLED)
   Visualization(self.ui).do()         #执行可视化
  else:
   self.ui.log("停止可视化中...")
   isStopVisualization = True

 def doAfter(self, result):
  pass

class Selete_All(BackgroundTask):          #全选按钮触发事件
 def doBefore(self):
  pass

 def do(self, arg):
  self.ui.gdEachJobvar.set(1)
  self.ui.gzEachJobvar.set(1)
  self.ui.szEachJobvar.set(1)
  self.ui.eachCityMoneyvar.set(1)
  self.ui.eachJobMoneyvar.set(1)
  self.ui.eachJobExpvar.set(1)
  self.ui.eachJobEduvar.set(1)
  self.ui.eachJobReqvar.set(1)
  self.ui.eachJobCSizevar.set(1)

 def doAfter(self, result):
  pass

class Invert_All(BackgroundTask):          #反选按钮触发事件
 def doBefore(self):
  pass

 def do(self, arg):
  self.ui.gdEachJobvar.set(abs(self.ui.gdEachJobvar.get()-1))
  self.ui.gzEachJobvar.set(abs(self.ui.gzEachJobvar.get()-1))
  self.ui.szEachJobvar.set(abs(self.ui.szEachJobvar.get()-1))
  self.ui.eachCityMoneyvar.set(abs(self.ui.eachCityMoneyvar.get()-1))
  self.ui.eachJobMoneyvar.set(abs(self.ui.eachJobMoneyvar.get()-1))
  self.ui.eachJobExpvar.set(abs(self.ui.eachJobExpvar.get()-1))
  self.ui.eachJobEduvar.set(abs(self.ui.eachJobEduvar.get()-1))
  self.ui.eachJobReqvar.set(abs(self.ui.eachJobReqvar.get()-1))
  self.ui.eachJobCSizevar.set(abs(self.ui.eachJobCSizevar.get()-1))

 def doAfter(self, result):
  pass

class Search(UiFrame):     #推荐界面
 def __init__(self, parent, **kwargs):
  UiFrame.__init__(self, parent, **kwargs)
  #mFrame = Labelframe(self)
#self.openGoogleButton self.openBossButton self.openJobButton self.spinbox self.startButton self.stopButton self.checkButton self.notGDButton self.startStopVirtualBtn.config(state=DISABLED)

#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#运行界面
class MainUi(UiRoot):
 """主窗口"""   
 def __init__(self):
  global root
  UiRoot.__init__(self)
  root = self
  self.protocol('WM_DELETE_WINDOW', Do_Destroy)
  self.geometry('400x650')   #窗口大小
  self.maxsize(width = 400, height = 650)      #窗口大小不可变
  self.minsize(width = 400, height = 650)      #窗口大小不可变
  self.title('Selenium控制型爬虫')         #窗口标题
  img1 = PhotoImage(file = 'google.gif')
  img2 = PhotoImage(file = 'boss.gif')
  img3 = PhotoImage(file = '51job.gif')
  
  tabs = ttk.Notebook(self, width = 388, height = 230)
  crawlerFrame = Frame(tabs)
  self.openGoogleButton = Button(crawlerFrame, text = '打开浏览器', image = img1, width = 60, height = 60, command = ShowTask(self).run)  #按钮 
  self.openGoogleButton.grid(row = 0, column = 0, padx = 40, pady = 10)
  self.openBossButton = Button(crawlerFrame, text = '打开Boss直聘', image = img2, width = 60, height = 60, command = OpenBoss(self).run)
  self.openBossButton.grid(row = 0, column = 1, padx = 0, pady = 10)
  self.openJobButton = Button(crawlerFrame, text = '打开前程无忧', image = img3, width = 60, height = 60, command = OpenJob(self).run)
  self.openJobButton.grid(row = 0, column = 2, padx = 40, pady = 10)

  Label(crawlerFrame, text='打开浏览器', font = '12').grid(row = 1, column = 0, padx = 0, pady = 0)    #按钮提示文字 
  Label(crawlerFrame, text='打开Boss直聘', font = '12').grid(row = 1, column = 1, padx = 0, pady = 0)
  Label(crawlerFrame, text='打开前程无忧', font = '12').grid(row = 1, column = 2, padx = 0, pady = 0)

  Label(crawlerFrame, text='爬取每一页面间隔休眠时间(单位:秒)', font = '3').place(x = 30, y = 130)
  self.spinbox = Spinbox(crawlerFrame, values = (5,10,15,20,25,30,35,40,45,50,60,70,80,90,100), state = 'readonly', font = '3', width = 4)
  self.spinbox.place(x = 305, y = 132)

  self.startButton = Button(crawlerFrame, text = '开始爬取', width = 16, height = 2,   font = '13', command = Start_Pause(self).run)  #按钮  
  self.startButton.place(x = 40, y= 174)
  self.stopButton = Button(crawlerFrame, text = '结束爬取', width = 16, height = 2,   font = '13', command = Stop(self).run)
  self.stopButton.place(x = 220, y= 174)
  tabs.add(crawlerFrame, text = '爬虫')
  #--------------------------------------------------------------以上为爬虫界面--------------------------------------------------------------------------------------------------
  checkFrame = Frame(tabs)
  Label(checkFrame, text='查出并去掉重复职位数据', font = ('宋体', 20)).place(x = 40, y = 30)
  self.checkButton = Button(checkFrame, text = '开始查重', width = 16, height = 2,   font = '13', command = Check(self).run)  #按钮
  self.checkButton.place(x = 120, y= 120)
  self.canvas = Canvas(checkFrame, width = 315, height = 22, bg = '#fafafa')    #画布
  self.canvas.place(x = 30, y = 185)
  self.canvas.create_rectangle(2,2,316,23, outline = 'black', width = 1)     #进度条
  self.fill_rec = self.canvas.create_rectangle(4,4,4,22, outline = '', width = 0, fill = '#fafafa')
  self.notGDButton = Checkbutton(checkFrame, text = '去除非广东的招聘信息', font = '10', command = Change_checkGD(self).run)   #勾选框
  self.notGDButton.place(x = 100, y = 80)
  tabs.add(checkFrame, text = '查重')
  #--------------------------------------------------------------以上为查重界面--------------------------------------------------------------------------------------------------
  visualizationFrame = Frame(tabs)
  self.gdEachJobvar = IntVar()
  self.gzEachJobvar = IntVar()
  self.szEachJobvar = IntVar()
  self.eachCityMoneyvar = IntVar()
  self.eachJobMoneyvar = IntVar()
  self.eachJobExpvar = IntVar()
  self.eachJobEduvar = IntVar()
  self.eachJobReqvar = IntVar()
  self.eachJobCSizevar = IntVar()
  
  self.gdEachJobBtn = Checkbutton(visualizationFrame, text = '广东数媒行业各职位数量分布', font = '6', command = None, variable = self.gdEachJobvar)  #17Map
  self.gdEachJobBtn.place(x = 6, y = 5)
  self.gzEachJobBtn = Checkbutton(visualizationFrame, text = '广州数媒行业各职位数量分布', font = '6', command = None, variable = self.gzEachJobvar)  #17Map
  self.gzEachJobBtn.place(x = 6, y = 29)
  self.szEachJobBtn = Checkbutton(visualizationFrame, text = '深圳数媒行业各职位数量分布', font = '6', command = None, variable = self.szEachJobvar)  #17Pie
  self.szEachJobBtn.place(x = 6, y = 53)
  self.eachCityMoneyBtn = Checkbutton(visualizationFrame, text = '广东各城市数媒行业职位薪水对比', font = '6', command = None, variable = self.eachCityMoneyvar)  #1Line
  self.eachCityMoneyBtn.place(x = 6, y = 77)
  self.eachJobMoneyBtn = Checkbutton(visualizationFrame, text = '数媒行业各职位薪水对比', font = '6', command = None, variable = self.eachJobMoneyvar)  #1Bar
  self.eachJobMoneyBtn.place(x = 6, y = 101)
  self.eachJobExpBtn = Checkbutton(visualizationFrame, text = '数媒行业各职位经验要求对比', font = '6', command = None, variable = self.eachJobExpvar)  #17Pie
  self.eachJobExpBtn.place(x = 6, y = 125)
  self.eachJobEduBtn = Checkbutton(visualizationFrame, text = '数媒行业各职位学历要求对比', font = '6', command = None, variable = self.eachJobEduvar)  #17Pie
  self.eachJobEduBtn.place(x = 6, y = 149)
  self.eachJobCSizeBtn = Checkbutton(visualizationFrame, text = '数媒行业各职位公司大小分布', font = '6', command = None, variable = self.eachJobCSizevar)  #17Pie
  self.eachJobCSizeBtn.place(x = 6, y = 173)
  self.eachJobReqBtn = Checkbutton(visualizationFrame, text = '数媒行业各职位描述要求词云', font = '6', command = None, variable = self.eachJobReqvar)  #17WordCloud
  self.eachJobReqBtn.place(x = 6, y = 197)
  self.startStopVirtualBtn = Button(visualizationFrame, text = '开始', width = 7, height = 3,   font = '10', command = VisualizationButton(self).run)  #按钮
  self.startStopVirtualBtn.place(x = 304, y = 85)
  self.seleteAllBtn = Button(visualizationFrame, text = '全选', width = 4, height = 1,   font = '2', command = Selete_All(self).run)  #按钮
  self.seleteAllBtn.place(x = 315, y = 35)
  self.invertAllBtn = Button(visualizationFrame, text = '反选', width = 4, height = 1,   font = '2', command = Invert_All(self).run)  #按钮
  self.invertAllBtn.place(x = 315, y = 170)
  tabs.add(visualizationFrame, text = '可视化')

  #-----------------------------------------------------------以上为数据可视化界面-----------------------------------------------------------------------------------------------
  #tabs.add(Search(self), text = '职位推送')
  tabs.place(x = 5, y = 5)

  #-----------------------------------------------------------------------------------------------------------------------------------------------------------------
  self.text = Text(self, width = 55, height = 29)  #Debug显示框
  self.text.place(x = 5, y = 265)
  self.text.configure(state=DISABLED)
  
  self.mainloop()

 def log(self, msg):   #打印字符的函数
  print(msg)
  self.text.configure(state=NORMAL)
  self.text.insert(END, msg + '\n')
  self.text.configure(state=DISABLED)
  self.text.see(END)



if __name__ == '__main__':
 multiprocessing.freeze_support()
 MainUi()
