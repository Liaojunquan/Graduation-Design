"""import tkinter as tk
import tkinter.messagebox
from tkinter import *
import selenium as se
from selenium import webdriver
from selenium.webdriver.remote.command import Command
import urllib3
import time
import openpyxl
import requests
import bs4"""

"""driver = webdriver.Chrome()
driver.get('https://jobs.51job.com/guangzhou-hzq/128123507.html?s=01&t=0')
time.sleep(5)
#print(driver.current_url)
#print(type(driver.current_url))
msg = driver.find_element_by_class_name('tCompany_main').find_element_by_class_name('bmsg').text
print(msg.strip())
time.sleep(10)
driver.quit()"""

"""try:
 driver.execute(Command.STATUS)
except OSError:
 print("OsError!!!")
except ConnectionRefusedError:
 print("ConnectionRefusedError!!!")
except urllib3.exceptions.NewConnectionError:
 print("NewConnectionError!!!")
except urllib3.exceptions.MaxRetryError:
 print("MaxRetryError!!!")
except Exception as e:
 print(e)"""


"""
root = Tk()
s = Spinbox(root, values = (0,5,10,15,20))
s.pack()
print(type(int(s.get())))
tk.messagebox.showerror('错误', '无法识别该网页!')
if tk.messagebox.askyesno(title = '警告', message = '是否要停止爬取?'):
 print('是')
else:
 print('否')
root.mainloop()"""


"""while True:
 try:
  time.sleep(1)
 except KeyboardInterrupt:
  print('中断')
  break
 else:
  print('else')"""

"""driver = webdriver.Chrome()
driver.set_page_load_timeout(10)
try:
 driver.get('https://jobs.51job.com/guangzhou-hzq/128123507.html?s=01&t=0')
except se.common.exceptions.InvalidArgumentException:
 print('URL参数错误!')
except se.common.exceptions.WebDriverException:
 print('连接超时，请检查网络！')
else:
 print('正常')"""

#wb = openpyxl.open(r'C:\Users\Administrator\Desktop\datatest.xlsx')
#ws = wb['Sheet']

#max_r = ws.max_row
#tmp = 0
#print('查重中...')
#for i in range(2, max_r + 1):
# for j in range(i+1, max_r + 1):
#  if ws['A'+str(i)].value == ws['A'+str(j)].value and ws['I'+str(i)].value == ws['I'+str(j)].value and ws['E'+str(i)].value == ws['E'+str(j)].value: #职位名称 公司名称 地区均相同   and ws['E'+str(i)].value == ws['E'+str(j)].value
#   if ws['N'+str(i)].value < 1000 and ws['N'+str(j)].value < 1000:     #1月至9月发布
#    if ws['N'+str(i)].value > ws['N'+str(j)].value:
#     ws['A'+str(j)] = 'delete'
#    else:
#     ws['A'+str(i)] = 'delete'
#   elif ws['N'+str(i)].value >= 1000 and ws['N'+str(j)].value >= 1000:  #去年10月至12月发布
#    if ws['N'+str(i)].value > ws['N'+str(j)].value:
#     ws['A'+str(j)] = 'delete'
#    else:
#     ws['A'+str(i)] = 'delete'
#   elif ws['N'+str(i)].value < 1000 and ws['N'+str(j)].value >= 1000:  #一个1月至9月发布，另一个去年10月至12月发布
#    ws['A'+str(j)] = 'delete'
#   elif ws['N'+str(i)].value >= 1000 and ws['N'+str(j)].value < 1000:  #一个1月至9月发布，另一个去年10月至12月发布
#    ws['A'+str(i)] = 'delete'
#   print('重复i = ' + str(i) + '  j = ' + str(j))
#   tmp += 1
#   break
#print('查找到重复'+str(tmp)+'个')
#if tmp > 0:
# print('删除中...')
# ws_tmp = wb.create_sheet('tmp_sheet')
# l = []
# for k in range(1, max_r + 1):
#  if ws['A'+str(k)].value != 'delete':
#   l.clear()
#   r = ws[str(k)]
#   for z in r:
#    l.append(z.value)
#   ws_tmp.append(l)
# del wb['Sheet']
# ws_tmp.title = 'Sheet'
# ws_tmp.freeze_panes = 'A2'
# print('删除成功')
# wb.save(r'C:\Users\Administrator\Desktop\datatest.xlsx')
#print(ws.max_row)
#wb.close()

"""headers = {'User-Agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.89 Safari/537.36'}
while True:
 try:
  res = requests.get('https://jobs.51job.com/guangzhou-yxq/128155496.html?s=01&t=0', headers=headers ,timeout = 5)
 except requests.exceptions.ConnectTimeout:
  print('连接超时！  10秒后重连')
  time.sleep(10)
 except requests.exceptions.ConnectionError:
  print('无法连接该网页，请检查网络！  10秒后重连')
  time.sleep(10)
 else:
  break
print('连接成功')"""

"""def open_URL(url):                #使用requests模块获取前程无忧网站各工作的职位描述和要求
        tmp_time = 3           #最多重连3次
        bmsg = None
        while tmp_time > 0:
                headers = {'User-Agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.89 Safari/537.36'}
                while True:
                        try:
                                res = requests.get(url, headers=headers ,timeout = 5)   #5秒后超时
                        except requests.exceptions.ConnectTimeout:
                                print('连接超时！ 5秒后重连')
                                time.sleep(5)
                        except (requests.exceptions.MissingSchema, requests.exceptions.InvalidURL):
                                print('URL参数错误！')
                                return ""
                        except requests.exceptions.ConnectionError:
                                print('连接出错！')
                                return ""
                        else:
                                break
                res.encoding = 'gbk'      #字符编码为gbk
                soup = bs4.BeautifulSoup(res.text,'html.parser')
                bmsg = soup.find('div', class_='bmsg')
                if bmsg == None:
                        print('获取页面信息失败！  5秒后重试')  #bs4找不到该元素
                        time.sleep(5)
                        tmp_time -= 1
                else:
                        break
        if tmp_time == 0:
                return ""
        else:
                return bmsg.text.replace('\n',' ').replace('\xa0',' ').replace('微信分享','')

print(open_URL('http://s'))"""

"""class A():
 def __init__(self, x):
  self.x = x
 def power(self, y):
  return y * y
 def power_two(self):
  return self.power()*self.power()
print(A(2).power(3))"""

"""ee_time = '1231'
if ee_time.find('发布') != -1 and ee_time.find('-') != -1:
 tmp_time = ee_time.split('发布')[0]
 print(int(tmp_time.split('-')[0] + tmp_time.split('-')[1]))
else:
 print(ee_time.strip())"""

import jieba

s_1 = """岗位职责： 负责ICT、终端移动应用等领域图像与分类/检测算法、内容生成与推荐算法，以及游戏自动化驱动AI等领域算法研究，引入业界最新最前沿的技术，支持华为在ICT领域和终端领域生态系统能力达到业界领先水平。 岗位要求： 1、 有机器学习和深度学习、迁移学习基础，掌握Tensorflow和Caffe、Pytorch框架，熟悉常见的CNN、RNN等神经网络和算法； 2、 熟悉图像分类、物体检测、文本识别和分类等算法，有研究应用自动布局、VAE、GAN领域算法经验，CV、NLP、RL等领域算法应用经验更佳； 专业: 计算机、AI、电子信息相关专业 工作地点：东莞、北京、上海、武汉 职位招聘10-15人"""

s_2 = """岗位职责： 负责ICT、终端移动应用等领域图像与分类/检测算法、内容生成与推荐算法，以及游戏自动化驱动AI等领域算法研究，引入业界最新最前沿的技术，支持华为在ICT领域和终端领域生态系统能力达到业界领先水平。 岗位要求： 1、 有机器学习和深度学习、迁移学习基础，掌握Tensorflow和Caffe、Pytorch框架，熟悉常见的CNN、RNN等神经网络和算法； 2、 熟悉图像分类、物体检测、文本识别和分类等算法，有研究应用自动布局、VAE、GAN领域算法经验，CV、NLP、RL等领域算法应用经验更佳； 专业: 计算机、AI、电子信息相关专业 工作地点：东莞、北京、上海、武汉"""
rate = 0  #吻合度
del_char = [' ','、','“','”','；','，','（','）','。','《','》','【','】','：','！','￥','……','…',
            '——','—','·','(',')','{','}','|','[',']',':',';','"','\'','\\','<','>','?',',','.','/',
            '!','`','~',]
for i in del_char:
 s_1 = s_1.replace(i, '')  #去特殊字符
 s_2 = s_2.replace(i, '')

max_num = 0
l_1 = jieba.lcut(s_1) #分词
l_2 = jieba.lcut(s_2)

set_1 = set(l_1) #去重复
set_2 = set(l_2)

unit = set_1 & set_2   #取交集

if len(set_1) > len(set_2):
 max_num = len(set_1)
else:
 max_num = len(set_2)

print('吻合度: ' + str(len(unit)/max_num*100) + '%')
