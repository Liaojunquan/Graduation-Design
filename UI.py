import tkinter as tk
from tkinter import ttk   #使用Notebook模块
from tkinter import *
import tkinter.messagebox
import threading
import selenium as se
from selenium import webdriver
from selenium.webdriver.remote.command import Command              #浏览器状态命令
from selenium.webdriver.common.action_chains import ActionChains   #用于操控鼠标在网页上移动
import urllib3
import time
import os
import openpyxl
import requests     #网页请求模块
import bs4          #html解析模块
import jieba
import copy


#全局变量
driver = None            #浏览器驱动变量
sleepTime = 5            #用于间隔休眠的时间，单位秒
startFlag = False        #是否未停止，用于控制爬虫是否停止。一旦按下开始键，是否暂停中或运行中，一律当爬虫未停止。
runFlag = False          #是否运行中，用于开始或暂停爬虫
stopFlag = False         #传递到爬虫内部，用于彻底停止爬虫
chromeOpen = False       #谷歌浏览器是否已打开
file_name = os.getcwd() + r'\data.xlsx'           #数据保存文件名
wb = None                #Excel工作本Workbook
ws = None                #Excel活动表格
KeyWord = """引擎 客户端 开发 端游 手游 小游戏 Unity U3D u3d U3d UE4 ue4 Ue4 UE 虚幻 cocos Cocos COCOS 2d 平面设计 三维 建模 渲染 视觉 MAYA Maya maya C4D c4d C4d 3DMAX 3dmax cad CAD ZBrush 
2D 3d 3D 测试 策划 美工 美术 设计 UI ui Ui 特效 动画 动作 服务器 维护 维稳 脚本 数据 前端 WEB web Web 贴图 材质 影视 视频 拍摄 摄影 摄像 导演 编导 前期 后期 剪辑 分镜 AR VR 研发经理 全栈 
小程序 数据库 数据管理 数据库管理 数据分析 DBA sql SQL Sql JAVA Java java C++ c++ C c Windows windows Android android 安卓 Python python 爬虫 图像识别 音频 音视频 VR  Vr vr 虚拟现实 AR ar 
Ar 虚拟仿真 Unreal 交互开发 程序开发 计算机图形 Flash flash FLASH An AN MG mg Mg 二维 分镜 AE Ae ae PR pr Pr 达芬奇 多媒体 Davinci 调色 修图 调光 图片 数字媒体 AI ai Ai 架构 算法 大数据 
数据挖掘 .NET .net .Net C# 游戏"""     #关键词

StopWord = """销售 推广 地产 经理 主管 客服 人事 教育 讲师 分销 电商 投放 运营 翻译 英语 商务 教师 教研 老师 玩游戏 陪 质检 猎头 游戏机 装配 美主 gm gs GM GS Gm Gs 生产 * 激光 
扫描 机械 公关 理财 译员 发行 主播 直播 玩家 试玩 帮派 治疗 市场 营销 游戏店 服务员 体验 助教 经纪人 管培 投资 顾问 合伙人 投标 审核 制片人 演员 自动驾驶 售前 售后 院长 企业文化 体验馆 店长 
接待 财务 店员 营业员 内容生态 看房 看楼 讲解 ARM arm 收账 会计 半导体 工艺 交通 能源 射频 SAE 标注 漆 涂料 化工 甲油胶 建筑 建材 油墨 印刷 环保 化学 热转印 汽车 胶 粉体 灯光 CAE FAE 
硬件 嵌入式 自动化 芯片 AE工程师 供应链 AE应用工程师 PR工程师 仪器 AE技术 驱动 电路 免费 调色师傅 调色员 美甲 印 电子 鞋 调色工程师 调色技术 电镀 技工 粉 妆 涂 油 配色 普工 操作员 剂 分子 
调色员 墨 瓷砖 工厂 家居 SMT 外贸 带头 FPGA 审核 智能检测 咨询 设备 保险 家庭 邀约 招生 KA 社区 置业 司机 配送 孵化 机器人 BA 需求 微商 业务 寒假 货 蓝牙 幼 营运 电玩 店 打游戏 送餐 
辅导 单片机 保育 厨 网吧 秘书 总裁 家教 主持人 物流 收银 咖啡 物料 装修 灯 室内 仓 保健 干部 封装 采购 导玩 乐园 代驾 班主任 跟单 茶 篮球 引导员 达人 法务 解说 信用 婴 团长 爱好 电话 HR 
人力资源 催收 吃 住 人寿 前台 出纳 金融 电竞 骑手 保安 交易员 贷 代练 拓展 渠道 人力 地推 赚钱 风控 投诉 游戏学徒 公会 党 游戏专员 电销 银行 测评 管理培训 文员 政 大使 底层 政策 电话 网销 
导师 出口 操作 游戏助理 游戏管理员 快递 临时工 保姆 嫂 模特 SEO SEM seo sem Seo Sem 交易 证券 时工 上门 分拣 教练 美容 礼仪 清洁 维修 检员 打字 塑 房 预算 造价 社 融 赛事 金牌"""  #停用词

TagStopWord = "漆 涂 化工 甲油胶 建筑 建材 油 墨 印刷 环保 化学 汽车 胶 粉 妆 灯光 调研 营销 市场 客户 乙方 射频 销售 公关 电商 创意 推广 广告 运营 项目 执行 策划 创意 玩具 装饰 装修 物流"  #标签停用词
sw = StopWord.split(' ')       #停用词列表
kw = KeyWord.split(' ')        #关键词列表
tsw = TagStopWord.split(' ')   #标签停用词列表
isChecking = False             #是否查重中标志
isStopChecking = False         #是否停止查重
root = None                    #tkinter的root
force_exit = False             #用于强制退出程序的标志

def Do_Destroy():          #全局函数用于监听窗口是否关闭
 global runFlag
 global startFlag
 global stopFlag
 global wb
 global file_name
 global isChecking
 global isStopChecking
 global root
 global force_exit
 if startFlag or runFlag or isChecking:   #爬虫或查重运行中
  if tk.messagebox.askyesno(title = '警告', message = '程序正在运行，是否确定要退出？'):  #抛出提示
   force_exit = True
   stopFlag = True
   isStopChecking = True
   tk.messagebox.showinfo('提示', '点击确认3秒后退出程序')
   time.sleep(3)
   wb.close()
   root.destroy()
  else:
   return
 elif tk.messagebox.askyesno(title = '警告', message = '是否确定要退出？'):  #抛出提示
  if wb != None:
   wb.close()
  root.destroy()

#-------------------------------------------------------------------------------------------------------------------------------------------------------------------
#Excel工作簿的初始化
class initWB():       #初始化工作本和工作表，无则创建
 def __init__(self, ui):
  self.ui = ui
  
 def open_data_file(self):
  global wb
  global ws
  try:
   wb = openpyxl.open(file_name)      #打开Excel文件
  except FileNotFoundError:
   self.ui.log('找不到Excel文件data!')
   self.ui.log('创建一个新的Excel文件')
   wb = openpyxl.Workbook()       #创建一个新的工作本
   ws = wb.active
   ws.append(['职位名称', '最低薪酬(元/月)', '最高薪酬', '平均薪酬', '公司所在地', '经验要求', '学历要求', '公司福利', '公司名称', '链接地址', '公司类型', '公司大小', '业务定位方向', '发布时间', '职位要求和描述']) #首行标题
   ws.freeze_panes = 'A2'  #冻结首行
   wb.save(file_name)      #保存文件
   self.ui.log('Excel文件data创建完成!')
  else:
   self.ui.log('成功找到并打开Excel文件data')
   ws = wb.active

#--------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#爬虫实现
class Crawling():
 def __init__(self, ui):
  self.ui = ui                #UI root
  self.current_url = ''       #当前URL
  self.isFirst = True         #是否第一次进入爬取循环

 def pause(self):     #暂停函数
  global runFlag
  global stopFlag
  global startFlag
  first_in = True
  while startFlag == True and runFlag == False and stopFlag == False and first_in == True:
   first_in = False
   self.ui.log('暂停中')
   time.sleep(1)      #休眠实现暂停

 def mStop(self):
  global stopFlag
  global startFlag
  global runFlag
  global wb
  global file_name
  if startFlag == True and stopFlag == True:
   self.ui.log('结束爬取循环!')
   wb.save(file_name)
   wb.close()     #关闭工作本
   self.isFirst = True  #重置第一次循环标志
   stopFlag = False
   startFlag = False
   runFlag = False
   self.ui.log("爬虫已被停止")
   return True
  else:
   return False

 def de_illegal(self, o_str):    #去非法字符函数
  illegal_str = ['\000','\001','\002','\003','\004','\005','\006','\007','\010','\013','\014','\016','\017','\020','\021',
                 '\022','\023','\024','\025','\026','\027','\030','\031','\032','\033','\034','\035','\036','\037']  #excel无法识别的非法字符
  n_str = o_str    #字符串浅拷贝
  for each_i_s in illegal_str:
   n_str = n_str.replace(each_i_s, '')   #去除非法字符
  return n_str
  
 def open_URL(self, url):                #使用requests模块获取前程无忧网站各工作的职位描述和要求
  #print(url)
  tmp_time = 5           #最多重连3次
  bmsg = None
  tmp_sleep_time = 0     #休眠时间
  while tmp_time > 0:
   self.pause()         #提供暂停
   if self.mStop():     #提供停止
    return ""
   headers = {'User-Agent':'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.89 Safari/537.36'}
   while True:
    self.pause()         #提供暂停
    try:
     res = requests.get(url, headers=headers ,timeout = 8)   #8秒后超时
    except requests.exceptions.ConnectTimeout:
     tmp_sleep_time += 5
     self.ui.log('连接超时！ ' + str(tmp_sleep_time) + '秒后重连')
     if self.mStop():    #提供停止
      return ""
     time.sleep(tmp_sleep_time)
    except (requests.exceptions.MissingSchema, requests.exceptions.InvalidURL):
     self.ui.log('URL参数错误！')
     return ""
    except requests.exceptions.ConnectionError:
     self.ui.log('连接出错！')
     return ""
    else:
     break
   self.pause()         #提供暂停
   res.encoding = 'gbk'      #字符编码为gbk
   soup = bs4.BeautifulSoup(res.text,'html.parser')
   bmsg = soup.find('div', class_='bmsg')
   if bmsg == None:
    tmp_sleep_time += 5
    self.ui.log('bs4找不到职位描述元素！ ' + str(tmp_sleep_time) + '秒后重试')  #bs4找不到该元素
    if self.mStop():   #提供停止
     return ""
    time.sleep(tmp_sleep_time)
    tmp_time -= 1     #次数减一
   else:
    break
  self.pause()         #提供暂停
  if tmp_time == 0:
   return ""
  else:
   return bmsg.text.replace('\n',' ').replace('\xa0',' ').replace('微信分享','').replace('【',' ').replace('】',' ')

 def append_list_job(self, ee, ee_l, ee_d, ee_fl, ee_c, ee_link, ee_t_s, ee_b, ee_time, ee_msg):  #职位名称  薪资  地区|经验|学历  福利  公司名称  链接  公司类型|大小  业务方向 发布时间 职位描述和要求
  l = []
  sal_low = 0
  sal_hight = 0
  sal_avg = 0
  self.pause()         #提供暂停
  l.append(ee.replace(' ',''))
  ee_l = ee_l.replace(' ','')
  if ee_l.find('-') == -1:
   if ee_l.find('元/日') != -1:
    sal_low = int(ee_l.split('元/日')[0].strip()) * 22              #统一标准单位元/月
    sal_hight = int(ee_l.split('元/日')[0].strip()) * 26
    sal_avg = (sal_low + sal_hight) // 2
    l.append(sal_low)
    l.append(sal_hight)
    l.append(sal_avg)
   elif ee_l.find('元/天') != -1:
    sal_low = int(ee_l.split('元/天')[0].strip()) * 22              #统一标准单位元/月
    sal_hight = int(ee_l.split('元/天')[0].strip()) * 26
    sal_avg = (sal_low + sal_hight) // 2
    l.append(sal_low)
    l.append(sal_hight)
    l.append(sal_avg)
   elif ee_l.find('万/月') != -1:
    sal_low = int(float(ee_l.split('万/月')[0].strip()) * 10000)             #统一标准单位元/月
    sal_hight = sal_low
    sal_avg = sal_low
    l.append(sal_low)
    l.append(sal_hight)
    l.append(sal_avg)
   elif ee_l.find('千/月') != -1:
    sal_low = int(float(ee_l.split('千/月')[0].strip()) * 1000)             #统一标准单位元/月
    sal_hight = sal_low
    sal_avg = sal_low
    l.append(sal_low)
    l.append(sal_hight)
    l.append(sal_avg)
   elif ee_l.find('万/年') != -1:
    sal_low = int(float(ee_l.split('万/年')[0].strip()) * 10000 // 12)             #统一标准单位元/月
    sal_hight = sal_low
    sal_avg = sal_low
    l.append(sal_low)
    l.append(sal_hight)
    l.append(sal_avg)
   else:
    l.append(ee_l.strip())
    l.append("null")
    l.append("null")
  else:
   if ee_l.find('万/月') != -1:
    try:
     sal_low = int(float(ee_l.split('万/月')[0].split('-')[0]) * 10000)      #统一标准单位元/月
     sal_hight = int(float(ee_l.split('万/月')[0].split('-')[1]) * 10000)
     sal_avg = (sal_low + sal_hight) // 2
     l.append(sal_low)
     l.append(sal_hight)
     l.append(sal_avg)
    except:
     l.append(ee_l.strip())
     l.append("null")
     l.append("null")
   elif ee_l.find('千/月') != -1:
    try:
     sal_low = int(float(ee_l.split('千/月')[0].split('-')[0]) * 1000)            #统一标准单位元/月
     sal_hight = int(float(ee_l.split('千/月')[0].split('-')[1]) * 1000)
     sal_avg = (sal_low + sal_hight) // 2
     l.append(sal_low)
     l.append(sal_hight)
     l.append(sal_avg)
    except:
     l.append(ee_l.strip())
     l.append("null")
     l.append("null")
   elif ee_l.find('万/年') != -1:
    try:
     sal_low = int(float(ee_l.split('万/年')[0].split('-')[0]) * 10000 // 12)            #统一标准单位元/月
     sal_hight = int(float(ee_l.split('万/年')[0].split('-')[1]) * 10000 // 12)
     sal_avg = (sal_low + sal_hight) // 2
     l.append(sal_low)
     l.append(sal_hight)
     l.append(sal_avg)
    except:
     l.append(ee_l.strip())
     l.append("null")
     l.append("null")
   elif ee_l.find('元/日') != -1:
    try:
     sal_low = int(float(ee_l.split('元/日')[0].split('-')[0]) * 26)            #统一标准单位元/月
     sal_hight = int(float(ee_l.split('元/日')[0].split('-')[1]) * 26)
     sal_avg = (sal_low + sal_hight) // 2
     l.append(sal_low)
     l.append(sal_hight)
     l.append(sal_avg)
    except:
     l.append(ee_l.strip())
     l.append("null")
     l.append("null")
   elif ee_l.find('元/天') != -1:
    try:
     sal_low = int(float(ee_l.split('元/天')[0].split('-')[0]) * 26)            #统一标准单位元/月
     sal_hight = int(float(ee_l.split('元/天')[0].split('-')[1]) * 26)
     sal_avg = (sal_low + sal_hight) // 2
     l.append(sal_low)
     l.append(sal_hight)
     l.append(sal_avg)
    except:
     l.append(ee_l.strip())
     l.append("null")
     l.append("null")
   else:
    l.append(ee_l.strip())
    l.append("null")
    l.append("null")
        
  self.pause()         #提供暂停
  if ee_d.find('|') == -1:
   l.append(ee_d.strip())
   l.append("")
   l.append("")
  else:
   ee_d = ee_d.replace(' ','')       #去空格
   tmp_index = ee_d.rfind('招')      #从右往左找
   if tmp_index != -1:
    ee_d = ee_d[0:tmp_index-1]        #去掉招收人数
   if len(ee_d.split('|')) == 3:
    l.append(ee_d.split('|')[0].strip())       #公司所在地
    l.append(ee_d.split('|')[1].strip())       #经验
    l.append(ee_d.split('|')[2].strip())       #学历
   elif len(ee_d.split('|')) == 2:
    tmp_index = 0
    while tmp_index < len(ee_d):
     if ee_d[tmp_index].isdigit():      #判断字符串中是否含有数字
      break
     else:
      tmp_index += 1
    if tmp_index < len(ee_d):               #含有数字
     l.append(ee_d.split('|')[0].strip())   #公司所在地
     l.append(ee_d.split('|')[1].strip())   #经验
     l.append("")
    else:
     if ee_d.find('在校生') == -1 and ee_d.find('应届生') == -1 and ee_d.find('无需经验') == -1:   #不包含经验要求
      l.append(ee_d.split('|')[0].strip())   #公司所在地
      l.append("")
      l.append(ee_d.split('|')[1].strip())   #学历
     else:                                                              #包含经验要求
      l.append(ee_d.split('|')[0].strip())   #公司所在地
      l.append(ee_d.split('|')[1].strip())   #经验
      l.append("")
   elif len(ee_d.split('|')) == 1:
    l.append(ee_d.split('|')[0].strip())
    l.append("")
    l.append("")
        
  self.pause()         #提供暂停
  l.append(ee_fl.strip())               #福利
  l.append(ee_c.strip())                #公司名称
  l.append(ee_link.strip())             #链接地址
  if ee_t_s.find('|') == -1:
   if ee_t_s.find('-') == -1:
    l.append(ee_t_s.strip())
    l.append('null')
   else:
    l.append('null')
    l.append(ee_t_s.strip())
  else:
   l.append(ee_t_s.split('|')[0].strip())    #公司类型
   l.append(ee_t_s.split('|')[1].strip())    #公司大小
  l.append(ee_b.strip())                            #公司定位业务方向

  if ee_time.find('发布') != -1 and ee_time.find('-') != -1:    #发布时间转数值
   tmp_time = ee_time.split('发布')[0]
   l.append(int(tmp_time.split('-')[0] + tmp_time.split('-')[1]))
  else:
   l.append(ee_time.strip())
  l.append(ee_msg.strip())                          #职位描述和要求
  self.ui.log('---' + ee + '---')   #打印职位
  return l

 def get_job(self):
  global sw
  global kw
  global driver
  global file_name
  global sleepTime
  global stopFlag
  global startFlag
  global runFlag
  global wb
  global ws
  self.current_url = driver.current_url
  while (self.isFirst or driver.current_url != self.current_url) and driver.current_url.find('51job.com') != -1:  #当前页面与上一页面不同，并且URL包含51job.com
   self.pause()            #提供暂停
   if self.mStop():       #提供停止
    return
   self.isFirst = False     #第一次循环设否
   self.ui.log("爬取当前页面中---------------------")
   self.current_url = driver.current_url
   while True:
    self.pause()     #提供暂停
    if self.mStop():       #提供停止
     return
    try:
     job_list = driver.find_element_by_class_name('j_joblist').find_elements_by_class_name('e')   #获取列表
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("无法获取职位列表  休眠15秒后重试")
     self.pause()     #提供暂停
     if self.mStop():       #提供停止
      return
     time.sleep(15)
    else:
     if len(job_list) > 0:                 #获取列表成功
      break
     else:                                 #列表长度有误
      self.ui.log("错误：职位列表长度为0!  休眠10秒后重试")
      self.pause()     #提供暂停
      if self.mStop():       #提供停止
       return
      time.sleep(10)
    
   for i in range(len(job_list)):
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    e = job_list[i].find_element_by_class_name("jname").text      #获取工作名称
    tmp = 0
    while tmp < len(sw):
     self.pause()         #提供暂停
     if self.mStop():       #提供停止
      return
     if e.find(sw[tmp]) != -1:
      break             #包含停用词，跳到下一工作
     else:
      tmp += 1
    if tmp < len(sw):
     continue           #包含停用词，跳到下一工作
    tmp = 0
    while tmp < len(kw):
     self.pause()         #提供暂停
     if self.mStop():       #提供停止
      return
     if e.find(kw[tmp]) != -1:
      break             #都不包含关键词，跳到下一工作
     else:
      tmp += 1
    if tmp == len(kw):
     continue           #都不包含关键词，跳到下一工作
    e_time = ""
    try:
     e_time = job_list[i].find_element_by_class_name("time").text      #获取发布时间
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有发布时间可获取")
    except:
     self.ui.log("获取发布时间出现其它错误!")
    e_l = job_list[i].find_element_by_class_name("sal").text     #薪酬
    e_d = job_list[i].find_element_by_class_name("d").text       #公司所在地 经验 学历
    e_c = job_list[i].find_element_by_class_name("cname").text   #公司名称
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    e_fl = ""
    try:
     e_fl = job_list[i].find_element_by_class_name("tags").get_attribute('title')   #福利
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有福利信息可获取")
    except:
     self.ui.log("获取福利信息出现其它错误!")
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    e_link = job_list[i].find_element_by_class_name("el").get_attribute('href')    #详情链接地址
    e_type_size = ""
    try:
     e_type_size = job_list[i].find_element_by_class_name("dc").text              #公司类型和大小
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有公司类型和大小信息可获取")
    except:
     self.ui.log("获取公司类型和大小信息出现其它错误!")
    e_business = ""
    try:
     e_business = job_list[i].find_element_by_class_name("int").text                  #业务方向
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有公司业务方向信息可获取")
    except:
     self.ui.log("获取公司业务方向信息出现其它错误!")
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    e_msg = self.open_URL(e_link)        #职位描述和要求
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    ws.append(self.append_list_job(self.de_illegal(e), self.de_illegal(e_l), self.de_illegal(e_d), self.de_illegal(e_fl),
                                   self.de_illegal(e_c), self.de_illegal(e_link), self.de_illegal(e_type_size),
                                   self.de_illegal(e_business), self.de_illegal(e_time), self.de_illegal(e_msg)))

   while True:
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    try:
     wb.save(file_name)       #保存数据到data文件
    except PermissionError:
     self.ui.log('表格被其它程序占用中，无法写入数据! 请关闭占用程序')
     tk.messagebox.showwarning("警告", "表格被其它程序占用中，请关闭占用程序")  #提示框
     self.ui.log('10秒后重试')
     self.pause()         #提供暂停
     if self.mStop():       #提供停止
      return
     time.sleep(10)
    else:
     break
   self.ui.log('保存成功,下一页')
   time.sleep(1)
   while True:
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    try:
     driver.find_element_by_class_name("j_page").find_element_by_class_name("next").click()    #下一页
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("找不到点击下一页的按钮  休眠15秒后重试!")
     self.pause()         #提供暂停
     if self.mStop():       #提供停止
      return
     time.sleep(15)
    else:
     break
   self.ui.log('准备爬取当前页面  休眠中zzz...')
   self.pause()         #提供暂停
   if self.mStop():       #提供停止
    return
   time.sleep(sleepTime)        #根据设定间隔休眠

  self.ui.log('结束爬取循环!')
  wb.save(file_name)
  wb.close()     #关闭工作本
  self.isFirst = True  #重置第一次循环标志
  startFlag = False
  stopFlag = False
  runFlag = False
  self.ui.openBossButton.config(state=NORMAL)
  self.ui.openJobButton.config(state=NORMAL)
  self.ui.startButton.config(text = '开始爬取')
  self.ui.spinbox.config(state=NORMAL)
  self.ui.log("爬虫已被停止")
#-------------------------------------------------以上为爬取前程无忧网页-------------------------------------------------------------------------------------------------------------

 def append_list_boss(self, ee, ee_l, ee_a, ee_d, ee_fl, ee_c, ee_link, ee_t_s, ee_b, ee_time, ee_msg):  #职位名称  薪资  地区  经验|学历  福利  公司名称  链接  公司类型|大小  业务方向 发布时间 职位描述和要求
  l = []
  sal_low = 0
  sal_hight = 0
  sal_avg = 0
  self.pause()         #提供暂停
  if ee_d.find('个月') != -1:       #实习职位,在职位名称后添加实习时间
   if ee.find('实习') != -1:
    ee = ee + '(' + ee_d[:ee_d.find('个月') + 2] + ')'
   else:
    ee = ee + '(实习' + ee_d[:ee_d.find('个月') + 2] + ')'
   ee_d = '在校/应届' + ee_d.split('个月')[1]     #实习的经验要求修改
  l.append(ee.replace(' ',''))
  ee_l = ee_l.replace(' ','')
  if ee_l.find('·') != -1:        #薪水含有13薪
   ee_fl = ee_fl + ' ' + ee_l[ee_l.find('·')+1:]    #将13薪放在福利后面
  if ee_l.find('-') == -1:          #薪酬字符串中不含-
   if ee_l.find('元/天') != -1:
    sal_low = int(ee_l.split('元/天')[0].strip()) * 22              #统一标准单位元/月
    sal_hight = int(ee_l.split('元/天')[0].strip()) * 26
    sal_avg = (sal_low + sal_hight) // 2
    l.append(sal_low)
    l.append(sal_hight)
    l.append(sal_avg)
   elif ee_l.find('K') != -1:
    sal_low = int(float(ee_l.split('K')[0].strip()) * 1000)             #统一标准单位元/月
    sal_hight = sal_low
    sal_avg = sal_low
    l.append(sal_low)
    l.append(sal_hight)
    l.append(sal_avg)
   elif ee_l.find('元/时') != -1:
    sal_low = int(float(ee_l.split('元/时')[0].strip()) * 8 * 26)             #统一标准单位元/月
    sal_hight = sal_low
    sal_avg = sal_low
    l.append(sal_low)
    l.append(sal_hight)
    l.append(sal_avg)
   else:
    l.append(ee_l.strip())
    l.append("null")
    l.append("null")
  else:
   if ee_l.find('K') != -1:
    try:
     sal_low = int(float(ee_l.split('K')[0].split('-')[0]) * 1000)            #统一标准单位元/月
     sal_hight = int(float(ee_l.split('K')[0].split('-')[1]) * 1000)
     sal_avg = (sal_low + sal_hight) // 2
     l.append(sal_low)
     l.append(sal_hight)
     l.append(sal_avg)
    except:
     l.append(ee_l.strip())
     l.append("null")
     l.append("null")
   elif ee_l.find('元/天') != -1:
    try:
     sal_low = int(float(ee_l.split('元/天')[0].split('-')[0]) * 26)            #统一标准单位元/月
     sal_hight = int(float(ee_l.split('元/天')[0].split('-')[1]) * 26)
     sal_avg = (sal_low + sal_hight) // 2
     l.append(sal_low)
     l.append(sal_hight)
     l.append(sal_avg)
    except:
     l.append(ee_l.strip())
     l.append("null")
     l.append("null")
   elif ee_l.find('元/时') != -1:
    try:
     sal_low = int(float(ee_l.split('元/时')[0].split('-')[0]) * 8 * 26)            #统一标准单位元/月
     sal_hight = int(float(ee_l.split('元/时')[0].split('-')[1]) * 8 * 26)
     sal_avg = (sal_low + sal_hight) // 2
     l.append(sal_low)
     l.append(sal_hight)
     l.append(sal_avg)
    except:
     l.append(ee_l.strip())
     l.append("null")
     l.append("null")
   else:
    l.append(ee_l.strip())
    l.append("null")
    l.append("null")
        
  self.pause()         #提供暂停
  l.append(ee_a.strip().replace(' ',''))      #公司位置
  if ee_d.find('学历不限') != -1:
   l.append(ee_d.split('学历不限')[0])        #经验要求
   l.append("学历不限")                       #学历要求
  elif ee_d.find('大专') != -1:
   l.append(ee_d.split('大专')[0])
   l.append("大专")
  elif ee_d.find('中专') != -1:
   l.append(ee_d.split('中专')[0])
   l.append("中专/中技")
  elif ee_d.find('高中') != -1:
   l.append(ee_d.split('高中')[0])
   l.append("高中")
  elif ee_d.find('初中及以下') != -1:
   l.append(ee_d.split('初中及以下')[0])
   l.append("初中及以下")
  elif ee_d.find('本科') != -1:
   l.append(ee_d.split('本科')[0])
   l.append("本科")
  elif ee_d.find('硕士') != -1:
   l.append(ee_d.split('硕士')[0])
   l.append("硕士")
  elif ee_d.find('博士') != -1:
   l.append(ee_d.split('博士')[0])
   l.append("博士")
  else:
   l.append(ee_d.strip())
   l.append("null")
        
  self.pause()         #提供暂停
  l.append(ee_fl.strip())               #福利
  l.append(ee_c.strip())                #公司名称
  l.append('https://www.zhipin.com' + ee_link.strip())    #链接地址
  tmp_s = ee_t_s.split(ee_b.strip())[1]
  if tmp_s.find('已上市') != -1:          #判断是否已上市
   l.append('已上市')              #公司类型
  else:
   l.append('未上市')
  index = 0
  while index < len(tmp_s):
   if tmp_s[index].isdigit():      #判断数字字符最初出现的位置
    break
   else:
    index += 1
  if index < len(tmp_s):
   l.append(tmp_s[index:len(tmp_s)])   #通过第一个数字字符位置分割字符串   公司大小
  else:
   l.append("")
  l.append(ee_b.strip())                            #公司定位业务方向

  if ee_time.find('发布于') != -1:      #发布日期
   if ee_time.find('月') != -1 and ee_time.find('日') != -1:
    l.append(int(ee_time.replace('发布于','').replace('月','').replace('日','')))  #日期转整形
   elif ee_time.find(':') != -1:
    l.append(int(time.strftime('%m%d', time.localtime())))    #获取当前日期
   elif ee_time.find('昨天') != -1:
    l.append(int(time.strftime('%m%d', time.localtime(time.time() - 60*60*24))))   #昨天日期
   else:
    l.append(ee_time.strip())
  else:
   l.append(ee_time.strip())
  l.append(ee_msg.replace('\n',' ').replace('【',' ').replace('】',' ')) #职位描述和要求
  self.ui.log('---' + ee + '---')   #打印职位
  return l

 def get_boss(self):
  global sw
  global kw
  global tsw
  global driver
  global file_name
  global sleepTime
  global stopFlag
  global startFlag
  global runFlag
  global wb
  global ws
  self.current_url = driver.current_url
  while (self.isFirst or driver.current_url != self.current_url) and driver.current_url.find('zhipin.com') != -1:  #当前页面与上一页面不同，并且URL包含zhipin.com
   self.pause()            #提供暂停
   if self.mStop():       #提供停止
    return
   self.isFirst = False     #第一次循环设否
   self.ui.log("爬取当前页面中---------------------")
   self.current_url = driver.current_url
   while True:
    self.pause()     #提供暂停
    if self.mStop():       #提供停止
     return
    try:
     job_list = driver.find_element_by_class_name('job-list').find_elements_by_class_name('job-primary')   #获取列表
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("无法获取职位列表 休眠15秒后重试!")
     self.pause()     #提供暂停
     if self.mStop():       #提供停止
      return
     time.sleep(15)
    else:
     if len(job_list) > 0:                 #获取列表成功
      break
     else:                                 #列表长度有误
      self.ui.log("错误：职位列表长度为0!  休眠10秒后重试")
      self.pause()     #提供暂停
      if self.mStop():       #提供停止
       return
      time.sleep(10)
    
   for i in range(len(job_list)):
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    e = job_list[i].find_element_by_class_name("job-name").text      #获取工作名称
    e_a = job_list[i].find_element_by_class_name("job-area").text    #获取公司地区
    tmp = 0
    while tmp < len(sw):
     self.pause()         #提供暂停
     if self.mStop():       #提供停止
      return
     if e.find(sw[tmp]) != -1:
      break             #包含停用词，跳到下一工作
     else:
      tmp += 1
    if tmp < len(sw):
     continue           #包含停用词，跳到下一工作
    e_tag = ""
    try:
     e_tag = job_list[i].find_element_by_class_name("info-append").find_element_by_class_name("tags").text   #职位标签
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有职位标签可获取")
    except:
     self.ui.log("获取职位标签出现其它错误!")
    else:
     if e_tag != "":
      tmp = 0
      while tmp < len(tsw):
       self.pause()         #提供暂停
       if self.mStop():       #提供停止
        return
       if e_tag.find(tsw[tmp]) != -1:
        break             #包含标签停用词，跳到下一工作
       else:
        tmp += 1
      if tmp < len(tsw):
       continue           #包含标签停用词，跳到下一工作
    tmp = 0
    while tmp < len(kw):
     self.pause()         #提供暂停
     if self.mStop():       #提供停止
      return
     if e.find(kw[tmp]) != -1:
      break             #都不包含关键词，跳到下一工作
     else:
      tmp += 1
    if tmp == len(kw):
     continue           #都不包含关键词，跳到下一工作
    e_l = job_list[i].find_element_by_class_name("job-limit").find_element_by_class_name("red").text  #薪酬
    e_d = ""
    try:
     e_d = job_list[i].find_element_by_class_name("job-limit").find_element_by_tag_name('p').text   #经验与学历
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有经验与学历要求可获取")
    except:
     self.ui.log("获取经验与学历要求出现其它错误!")
    e_c = job_list[i].find_element_by_class_name("company-text").find_element_by_class_name("name").text   #公司名称
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    e_fl = ""
    try:
     e_fl = job_list[i].find_element_by_class_name("info-append").find_element_by_class_name("info-desc").text   #福利
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有福利信息可获取")
    except:
     self.ui.log("获取福利信息出现其它错误!")
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    e_link = job_list[i].find_element_by_class_name("primary-box").get_attribute('href')    #详情链接地址
    e_type_size = ""
    try:
     e_type_size = job_list[i].find_element_by_class_name("company-text").find_element_by_tag_name("p").text  #公司类型和大小
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有公司类型和大小信息可获取")
    except:
     self.ui.log("获取公司类型和大小信息出现其它错误!")
    e_business = ""
    try:
     e_business = job_list[i].find_element_by_class_name("company-text").find_element_by_tag_name("p").find_element_by_tag_name("a").text    #公司业务定位
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有公司业务定位信息可获取")
    except:
     self.ui.log("获取公司业务定位信息出现其它错误!")
    e_time = ""
    try:
     e_time = job_list[i].find_element_by_class_name("job-pub-time").text    #发布时间
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("没有发布时间可获取")
    except:
     self.ui.log("获取发布时间出现其它错误!")
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    tmp_index = 5           #获取职位描述和要求最多有5次尝试机会
    tmp_sleep_time = 0
    e_msg = ""
    while tmp_index > 0:
     primary_box = job_list[i].find_element_by_class_name("primary-box")                      #获取primary-box元素
     ActionChains(driver).move_to_element(primary_box).perform()                              #移动鼠标到primary-box元素
     time.sleep(1)
     try:
      e_msg = job_list[i].find_element_by_class_name("info-detail").find_element_by_class_name("detail-bottom-text").text      #获取职位描述和要求
     except se.common.exceptions.NoSuchElementException:
      tmp_sleep_time += 5
      self.ui.log('无法获取职位描述和要求!  ' + str(tmp_sleep_time) + '秒后重试')
      driver.execute_script('window.scrollTo(0,document.body.scrollHeight)')      #跳到网页底部
      time.sleep(tmp_sleep_time)    #休眠
      tmp_index -= 1              #次数减一
     else:
      break
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    ws.append(self.append_list_boss(self.de_illegal(e), self.de_illegal(e_l), self.de_illegal(e_a), self.de_illegal(e_d),
                                    self.de_illegal(e_fl), self.de_illegal(e_c), self.de_illegal(e_link), self.de_illegal(e_type_size),
                                    self.de_illegal(e_business), self.de_illegal(e_time), self.de_illegal(e_msg)))

   while True:
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    try:
     wb.save(file_name)       #保存数据到data文件
    except PermissionError:
     self.ui.log('表格被其它程序占用中，无法写入数据! 请关闭占用程序')
     tk.messagebox.showwarning("警告", "表格被其它程序占用中，请关闭占用程序")  #提示框
     self.ui.log('10秒后重试')
     self.pause()         #提供暂停
     if self.mStop():       #提供停止
      return
     time.sleep(10)
    else:
     break
   self.ui.log('保存成功,下一页')
   time.sleep(1)
   while True:
    self.pause()         #提供暂停
    if self.mStop():       #提供停止
     return
    try:
     driver.find_element_by_class_name("page").find_element_by_class_name("next").click()    #下一页
    except se.common.exceptions.NoSuchElementException:
     self.ui.log("找不到点击下一页的按钮  休眠15秒后重试!")
     self.pause()         #提供暂停
     if self.mStop():       #提供停止
      return
     time.sleep(15)
    else:
     break
   self.ui.log('准备爬取当前页面  休眠中zzz...')
   self.pause()         #提供暂停
   if self.mStop():       #提供停止
    return
   time.sleep(sleepTime)        #根据设定间隔休眠

  self.ui.log('结束爬取循环!')
  wb.save(file_name)
  wb.close()     #关闭工作本
  self.isFirst = True  #重置第一次循环标志
  startFlag = False
  stopFlag = False
  runFlag = False
  self.ui.openBossButton.config(state=NORMAL)
  self.ui.openJobButton.config(state=NORMAL)
  self.ui.startButton.config(text = '开始爬取')
  self.ui.spinbox.config(state=NORMAL)
  self.ui.log("爬虫已被停止")


#---------------------------------------------------------------------以上为爬取Boss直聘------------------------------------------------------------------------------------------------------------
class Checking():
 def __init__(self, ui):
  self.ui = ui

 def stop_check(self):
  global isStopChecking
  global isChecking
  global force_exit
  if force_exit == True:
   return True
  if isStopChecking == True and isChecking == True and tk.messagebox.askyesno(title = '提示', message = '是否要停止查重？'):  #查重运行中并且确定要停止
   return True
  else:
   return False

 def finish_check(self):
  global isChecking
  global isStopChecking
  global force_exit
  isChecking = False
  isStopChecking = False
  if force_exit == False:
   self.ui.checkButton.config(text = '开始查重')  #UI界面恢复正常
   self.ui.stopButton.config(state=NORMAL)
   self.ui.startButton.config(state=NORMAL)
   self.ui.openGoogleButton.config(state=NORMAL)
   self.ui.openBossButton.config(state=NORMAL)
   self.ui.openJobButton.config(state=NORMAL)
   self.ui.log('查重结束!')

 def match_rate(self, str1, str2):  #获取吻合度
  s_1 = copy.deepcopy(str1)
  s_2 = copy.deepcopy(str2)
  rate = 0 #吻合度
  del_char = [' ','、','“','”','；','，','（','）','。','《','》','【','】','：','！','￥','……','…',
            '——','—','·','(',')','{','}','|','[',']',':',';','"','\'','\\','<','>','?',',','.','/',
            '!','`','~',]  #要去除的字符列表
  for each_del_char in del_char:
   s_1 = s_1.replace(each_del_char, '')  #去除以上字符
   s_2 = s_2.replace(each_del_char, '')
  max_num = 0
  l_1 = jieba.lcut(s_1) #分词
  l_2 = jieba.lcut(s_2)
  set_1 = set(l_1) #去重复
  set_2 = set(l_2)
  unit = set_1 & set_2   #取交集
  difference_set = set_1 ^ set_2   #取差集
  if len(set_1) > len(set_2):
   max_num = len(set_1)
  else:
   max_num = len(set_2)
  rate = len(unit)/max_num
  #self.ui.log('吻合度: ' + str(rate*100) + '%')
  del s_1, s_2, set_1, set_2, unit, max_num
  return rate, str(difference_set)

 def do_check(self):
  global wb
  global ws
  global file_name
  max_r = ws.max_row
  now_time = int(time.strftime('%m%d', time.localtime()))   #获取当前日期
  tmp = 0
  self.ui.log('查重中...')
  for i in range(2, max_r + 1):
   for j in range(i+1, max_r + 1):
    #time.sleep(1)           #Debug用于耗时
    if self.stop_check():   #提供停止
     self.finish_check()    #结束查重
     return
    if ws['A'+str(i)].value == ws['A'+str(j)].value and ws['I'+str(i)].value == ws['I'+str(j)].value and ws['E'+str(i)].value == ws['E'+str(j)].value and ws['F'+str(i)].value == ws['F'+str(j)].value: #职位名称 公司名称 地区 经验要求均相同
     if ws['N'+str(i)].value == ws['N'+str(j)].value:   #发布日期相同，包括今日
      #if ws['O'+str(i)].value.replace('工作要求','').replace('岗位职责','').replace('职位描述','').replace(' ','') == ws['O'+str(j)].value.replace('工作要求','').replace('岗位职责','').replace('职位描述','').replace(' ',''):
       #ws['A'+str(i)] = 'delete'    #职位描述及要求相同
      tmp_rate, d_str = self.match_rate(ws['O'+str(i)].value, ws['O'+str(j)].value)  #获取匹配率
      if tmp_rate == 1.0:
       ws['A'+str(i)] = 'delete'    #职位描述及要求相同
       self.ui.log('重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '%')
       tmp += 1
      else:
       self.ui.log('重复行: ' + str(i) + ' 和 ' + str(j) + ' 职位描述和要求相似率 = ' + str(tmp_rate*100) + '% 差别字符为' + d_str)
      break
     elif ws['N'+str(i)].value <= now_time and ws['N'+str(j)].value <= now_time:     #小于等于今日日期 今年1月1日到今日且不相同
      if ws['N'+str(i)].value > ws['N'+str(j)].value:
       ws['A'+str(j)] = 'delete'
      else:
       ws['A'+str(i)] = 'delete'
     elif ws['N'+str(i)].value > now_time and ws['N'+str(j)].value > now_time:     #大于今日日期 去年明日到去年12月31日且不相同
      if ws['N'+str(i)].value > ws['N'+str(j)].value:
       ws['A'+str(j)] = 'delete'
      else:
       ws['A'+str(i)] = 'delete'
     elif ws['N'+str(i)].value <= now_time and ws['N'+str(j)].value > now_time:   #一个今年，另一个去年
      ws['A'+str(j)] = 'delete'
     elif ws['N'+str(i)].value > now_time and ws['N'+str(j)].value <= now_time:   #一个去年，另一个今年
      ws['A'+str(i)] = 'delete'
     self.ui.log('重复行: ' + str(i) + ' 和 ' + str(j))
     tmp += 1
     break
  self.ui.log('查找到重复'+str(tmp)+'个')
  if tmp > 0:
   if self.stop_check():   #提供停止
    self.finish_check()    #结束查重
    return
   self.ui.log('删除中...')
   ws_tmp = wb.create_sheet('tmp_sheet')
   l = []
   for k in range(1, max_r + 1):
    if self.stop_check():   #提供停止
     self.finish_check()    #结束查重
     return
    if ws['A'+str(k)].value != 'delete':
     l.clear()
     r = ws[str(k)]
     for z in r:
      l.append(z.value)
     ws_tmp.append(l)
   del wb['Sheet']
   ws_tmp.title = 'Sheet'
   ws_tmp.freeze_panes = 'A2'
   self.ui.log('删除成功!')
   tk.messagebox.showinfo("提示", "查重工作已完成!")
   self.ui.log('查重完成!')
   wb.save(file_name)
  self.finish_check()  #结束查重
  wb.close()
  
#界面自定义类
class UiRoot(Tk):
 def __init__(self):
  Tk.__init__(self)

class UiFrame(Frame):
 def run(self, func):
  self.master.run(func)
  
class BackgroundTask(object):
 """Similar to Android's AsyncTask"""
 def __init__(self, ui):
  self.ui = ui       #tk的root

 def doBefore(self):
  """Runs on the main thread, returns arg"""
  pass
 def do(self, arg):
  """Runs on a separate thread, returns result"""
  pass
 def doAfter(self, result):
  """Runs on the main thread again"""
  pass

 def run(self):
  """Invoke this on the main thread only"""
  arg = self.doBefore()
  threading.Thread(target=self._onThread, args=[arg]).start()     #新建一个线程处理去按钮触发的事件

 def _onThread(self, arg):
  result = self.do(arg)
  self.doAfter(result)
  #self.ui.run(lambda: self.doAfter(result))

class ShowTask(BackgroundTask):  #继承BackgroundTask类       打开谷歌浏览器触发事件
 def doBefore(self):
  self.ui.openGoogleButton.config(state=DISABLED)

 def do(self, arg):
  self.ui.log('打开谷歌浏览器')
  global driver
  global chromeOpen
  global startFlag
  global stopFlag
  driver = webdriver.Chrome()
  driver.set_page_load_timeout(10)      #设置超时限制时间
  while True:
   try:
    driver.execute(Command.STATUS)   #检查浏览器状态
    time.sleep(1)
   except urllib3.exceptions.MaxRetryError:   #出现异常说明浏览器已退出
    self.ui.log('谷歌浏览器已关闭')
    tk.messagebox.showinfo("提示", "Google浏览器已关闭")
    if startFlag == True:         #在运行或暂停中则停止爬虫
     stopFlag = True
     self.ui.log('与浏览器失去联系，爬虫被迫停止!')
     tk.messagebox.showerror("警告", "与浏览器失去联系，爬虫被迫停止!")
     tk.messagebox.showinfo("提示", "请重启爬虫程序")   #建议退出程序
     os._exit(0)         #退出程序
    break
   except Exception:
    self.ui.log('无法连接到浏览器，请关闭浏览器重试！')
    tk.messagebox.showwarning("警告", "无法连接到浏览器，请关闭浏览器重试！")
    break
   else:
    if chromeOpen == False:
     self.ui.log("浏览器打开成功")
     chromeOpen = True

 def doAfter(self, result):
  global chromeOpen
  self.ui.openGoogleButton.config(state=NORMAL)
  chromeOpen = False


class OpenBoss(BackgroundTask):            #打开Boss直聘按钮触发事件
 def doBefore(self):
  self.ui.openBossButton.config(state=DISABLED)

 def do(self, arg):
  global chromeOpen
  global driver
  if chromeOpen == False:                #判断浏览器是否打开
   self.ui.log("Google浏览器未打开")
   tk.messagebox.showwarning("警告", "Google浏览器未打开！")
   return
  self.ui.log("打开Boss直聘网页")
  while True:
   try:
    driver.get('https://www.zhipin.com/')   #打开Boss直聘
   except se.common.exceptions.WebDriverException:
    self.ui.log('连接超时，请检查网络')
    self.ui.log('5秒后重连')
    time.sleep(5)
   else:
    break
  self.ui.log('请在网页设置好要爬取的职业和地区!')

 def doAfter(self, result):
  self.ui.openBossButton.config(state=NORMAL)


class OpenJob(BackgroundTask):         #打开前程无忧按钮触发事件
 def doBefore(self):
  self.ui.openJobButton.config(state=DISABLED)

 def do(self, arg):
  global chromeOpen
  global driver
  if chromeOpen == False:                 #判断浏览器是否打开
   self.ui.log("Google浏览器未打开")
   tk.messagebox.showwarning("警告", "Google浏览器未打开！")
   return
  self.ui.log("打开前程无忧网页")
  while True:
   try:
    driver.get('https://www.51job.com/')   #打开前程无忧
   except se.common.exceptions.WebDriverException:
    self.ui.log('连接超时，请检查网络')
    self.ui.log('5秒后重连')
    time.sleep(5)
   else:
    break
  self.ui.log('请在网页设置好要爬取的职业和地区!')

 def doAfter(self, result):
  self.ui.openJobButton.config(state=NORMAL)


class Start_Pause(BackgroundTask):      #开始/暂停按钮触发事件
 def doBefore(self):
  pass

 def do(self, arg):
  global runFlag
  global startFlag
  global driver

  if chromeOpen == False:                           #判断浏览器是否打开
   self.ui.log("Google浏览器未打开")
   tk.messagebox.showwarning("警告", "Google浏览器未打开！")
   return

  url = driver.current_url
  if url.find('zhipin.com') == -1 and url.find('51job.com') == -1:     #判断网页是否打开正确
   tk.messagebox.showerror('错误', '无法识别该网页！')
   self.ui.log("无法识别网页！ 请打开Boss直聘或前程无忧网页")
   return

  self.ui.openBossButton.config(state=DISABLED)      #爬取期间禁止再次打开或更换网页
  self.ui.openJobButton.config(state=DISABLED)
  if runFlag == False:     #检查是否在运行
   self.ui.log("开始爬取")
   global sleepTime
   sleepTime = int(self.ui.spinbox.get())  #设置休眠时间
   self.ui.startButton.config(text = '暂停爬取')
   self.ui.spinbox.config(state=DISABLED)
   runFlag = True                 #运行中标志
   #while runFlag and not stopFlag:
    #self.ui.log('工作中!')
    #time.sleep(sleepTime)
   if startFlag == False:     #第一次开始
    startFlag = True               #已经按过开始键标志
    initWB(self.ui).open_data_file()  #初始化wb和ws，打开data文件
    if url.find('51job.com') != -1:
     Crawling(self.ui).get_job()          #执行爬取前程无忧网页
    else:
     Crawling(self.ui).get_boss()          #执行爬取Boss直聘网页
  else:
   self.ui.log("暂停爬取")
   self.ui.startButton.config(text = '开始爬取')
   self.ui.spinbox.config(state=NORMAL)
   runFlag = False            #暂停中标志

 def doAfter(self, result):
  pass

class Stop(BackgroundTask):          #停止按钮触发事件
 def doBefore(self):
  self.ui.stopButton.config(state=DISABLED)

 def do(self, arg):
  global runFlag
  global stopFlag
  global startFlag
  if startFlag == False:                 #爬虫未开始却按下停止键
   self.ui.log("爬虫未处于运行或暂停中")
   tk.messagebox.showinfo("提示", "爬虫未处于运行或暂停中！")
   return
  if startFlag == True and tk.messagebox.askyesno(title = '提示', message = '是否要停止爬取？'):          #没停止并且确定要停止
   self.ui.log("停止爬取中")
   self.ui.startButton.config(text = '开始爬取')
   self.ui.spinbox.config(state=NORMAL)
   #runFlag = False         #没有运行标志
   stopFlag = True         #停止运行中的爬虫标志
   #startFlag = False       #已经按过开始键标志
   #time.sleep(5)
   #stopFlag = False
   #self.ui.log("爬虫已被停止")
   self.ui.openBossButton.config(state=NORMAL)      #在非工作爬取期间可以打开或更换网页
   self.ui.openJobButton.config(state=NORMAL)

 def doAfter(self, result):
  self.ui.stopButton.config(state=NORMAL)

class Check(BackgroundTask):          #查重按钮触发事件
 def doBefore(self):
  pass

 def do(self, arg):
  global isChecking
  global isStopChecking
  if isChecking == False and isStopChecking == False:
   self.ui.checkButton.config(text = '停止查重')
   self.ui.stopButton.config(state=DISABLED)        #UI按钮不可用
   self.ui.startButton.config(state=DISABLED)
   self.ui.openGoogleButton.config(state=DISABLED)
   self.ui.openBossButton.config(state=DISABLED)
   self.ui.openJobButton.config(state=DISABLED)
   isChecking = True        #查重中标志
   initWB(self.ui).open_data_file()    #初始化wb和ws，打开data文件
   Checking(self.ui).do_check()        #执行查重
  else:
   isStopChecking = True    #结束标志

 def doAfter(self, result):
  pass


class Generate(UiFrame):     #生成图片界面
 def __init__(self, parent, **kwargs):
  UiFrame.__init__(self, parent, **kwargs)
  #mFrame = Labelframe(self)

class Search(UiFrame):     #推荐界面
 def __init__(self, parent, **kwargs):
  UiFrame.__init__(self, parent, **kwargs)
  #mFrame = Labelframe(self)


#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
#运行界面
class MainUi(UiRoot):
 """主窗口"""   
 def __init__(self):
  global root
  UiRoot.__init__(self)
  root = self
  self.protocol('WM_DELETE_WINDOW', Do_Destroy)
  self.geometry('400x650')   #窗口大小
  self.maxsize(width = 400, height = 650)      #窗口大小不可变
  self.minsize(width = 400, height = 650)      #窗口大小不可变
  self.title('Selenium控制型爬虫')         #窗口标题
  img1 = PhotoImage(file = 'google.gif')
  img2 = PhotoImage(file = 'boss.gif')
  img3 = PhotoImage(file = '51job.gif')
  
  tabs = ttk.Notebook(self, width = 388, height = 230)
  crawlerFrame = Frame(tabs)
  self.openGoogleButton = Button(crawlerFrame, text = '打开浏览器', image = img1, width = 60, height = 60, command = ShowTask(self).run)  #按钮 
  self.openGoogleButton.grid(row = 0, column = 0, padx = 40, pady = 10)
  self.openBossButton = Button(crawlerFrame, text = '打开Boss直聘', image = img2, width = 60, height = 60, command = OpenBoss(self).run)
  self.openBossButton.grid(row = 0, column = 1, padx = 0, pady = 10)
  self.openJobButton = Button(crawlerFrame, text = '打开前程无忧', image = img3, width = 60, height = 60, command = OpenJob(self).run)
  self.openJobButton.grid(row = 0, column = 2, padx = 40, pady = 10)

  Label(crawlerFrame, text="打开浏览器", font = '12').grid(row = 1, column = 0, padx = 0, pady = 0)    #按钮提示文字 
  Label(crawlerFrame, text="打开Boss直聘", font = '12').grid(row = 1, column = 1, padx = 0, pady = 0)
  Label(crawlerFrame, text="打开前程无忧", font = '12').grid(row = 1, column = 2, padx = 0, pady = 0)

  Label(crawlerFrame, text="爬取每一页面间隔休眠时间(单位:秒)", font = '3').place(x = 30, y = 130)
  self.spinbox = Spinbox(crawlerFrame, values = (5,10,15,20,25,30,35,40,45,50,60,70,80,90,100), state = 'readonly', font = '3', width = 4)
  self.spinbox.place(x = 305, y = 132)

  self.startButton = Button(crawlerFrame, text = '开始爬取', width = 16, height = 2,   font = '13', command = Start_Pause(self).run)  #按钮  
  self.startButton.place(x = 40, y= 174)
  self.stopButton = Button(crawlerFrame, text = '结束爬取', width = 16, height = 2,   font = '13', command = Stop(self).run)
  self.stopButton.place(x = 220, y= 174)
  tabs.add(crawlerFrame, text = '爬虫')
  #--------------------------------------------------------------以上为爬虫界面--------------------------------------------------------------------------------------------------
  checkFrame = Frame(tabs)
  Label(checkFrame, text="查出并去掉重复职位数据", font = ('宋体', 20)).place(x = 40, y = 50)
  self.checkButton = Button(checkFrame, text = '开始查重', width = 16, height = 2,   font = '13', command = Check(self).run)  #按钮
  self.checkButton.place(x = 120, y= 120)
  tabs.add(checkFrame, text = '查重')
  #--------------------------------------------------------------以上为查重界面--------------------------------------------------------------------------------------------------
  tabs.add(Generate(self), text = '数据可视化')
  tabs.add(Search(self), text = '职位推送')
  tabs.place(x = 5, y = 5)

  #-----------------------------------------------------------------------------------------------------------------------------------------------------------------
  self.text = Text(self, width = 55, height = 29)  #Debug显示框
  self.text.place(x = 5, y = 265)
  self.text.configure(state=DISABLED)

  self.mainloop()

 def log(self, msg):
  print(msg)
  self.text.configure(state=NORMAL)
  self.text.insert(END, msg + '\n')
  self.text.configure(state=DISABLED)
  self.text.see(END)



if __name__ == '__main__':
 MainUi()
